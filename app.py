from datetime import date, timedelta, datetime
from collections import defaultdict
import os
import io
from flask import Flask, render_template, redirect, url_for, request, session, flash, jsonify, make_response, send_file
from models import (
    Usuario, Circo, CategoriaColaborador, Colaborador, ColaboradorCategoria,
    Elenco, CategoriaFornecedor, Fornecedor, CategoriaReceita, Receita,
    CategoriaDespesa, Despesa, Evento, DespesaEvento, ReceitaEvento,
    CategoriaVeiculo, Veiculo, MultaVeiculo, IpvaVeiculo, LicenciamentoVeiculo, ManutencaoVeiculo,
    EquipeEvento, ElencoEvento, FornecedorEvento, 
    DespesaEmpresa, ReceitaEmpresa, TIPOS_DESPESA, VeiculoEvento, Parametro,
    Escola, VisitaEscola, LogSistema
)
from forms import (
    UsuarioForm, LoginForm, CircoForm, CategoriaColaboradorForm, ColaboradorForm, AutoCadastroForm,
    ElencoForm, CategoriaFornecedorForm, FornecedorForm, CategoriaReceitaForm, ReceitaForm,
    CategoriaDespesaForm, DespesaForm, EventoForm, CategoriaVeiculoForm, VeiculoForm,
    MultaVeiculoForm, IpvaVeiculoForm, LicenciamentoVeiculoForm, ManutencaoVeiculoForm,
    EquipeEventoForm, ElencoEventoForm, FornecedorEventoForm, DespesaEventoForm,
    DespesaEmpresaForm, ReceitaEmpresaForm, VeiculoEventoForm, ParametroForm,
    EscolaForm, VisitaEscolaForm
)
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
from extensions import db, login_manager
from sqlalchemy import func, text, or_, and_
from flask_migrate import Migrate
from flask_login import login_required, current_user
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import requests

load_dotenv()  # Carrega variáveis do .env

app = Flask(__name__)

# Configuração simples (usa DATABASE_URL ou fallback local)
app.config.from_object('config.Config')

# Criar pasta de upload se não existir
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)



# Inicialização das extensões
db.init_app(app)
login_manager.init_app(app)
migrate = Migrate(app, db)
mail = Mail(app)


# User loader para Flask-Login
@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# Função de contexto para templates
@app.context_processor
def inject_user_functions():
    return dict(
        is_root_user=is_root_user,
        is_admin_user=is_admin_user,
        is_promotor_user=is_promotor_user
    )

def allowed_file(filename):
    """Verifica se o arquivo tem uma extensão permitida"""
    ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'pdf', 'doc', 'docx'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def     calcular_lucro_evento(id_evento):
    """
    Função unificada para calcular o lucro real de um evento
    usando a mesma lógica do relatório de fechamento.
    
    Returns:
        dict: {
            'total_receitas': float,
            'despesas_cabeca': float,
            'total_liquido': float,
            'cinquenta_porcento_show': float,
            'reembolso_midias': float,
            'repasse_total': float,
            'total_despesas_socrates': float,
            'resultado_show': float (lucro real),
            'total_despesas': float (para exibição)
        }
    """
    # Calcular receitas do evento
    total_receitas = db.session.query(func.sum(ReceitaEvento.valor)).filter_by(id_evento=id_evento).scalar() or 0
    
    # Calcular despesas de cabeça (com flag despesa_cabeca=True)
    despesas_cabeca_flag = db.session.query(func.sum(DespesaEvento.valor)).filter_by(
        id_evento=id_evento, 
        despesa_cabeca=True
    ).scalar() or 0
    
    # Calcular despesas da categoria "PAGAS PELO CIRCO"
    despesas_pagas_pelo_circo = db.session.query(func.sum(DespesaEvento.valor)).join(
        Despesa, DespesaEvento.id_despesa == Despesa.id_despesa
    ).join(
        CategoriaDespesa, Despesa.id_categoria_despesa == CategoriaDespesa.id_categoria_despesa
    ).filter(
        DespesaEvento.id_evento == id_evento,
        or_(
            CategoriaDespesa.nome.ilike('%PAGAS PELO CIRCO%'),
            CategoriaDespesa.nome.ilike('%PAGO PELO CIRCO%')
        )
    ).scalar() or 0
    
    # Total de despesas de cabeça = despesas com flag + despesas pagas pelo circo
    despesas_cabeca = despesas_cabeca_flag + despesas_pagas_pelo_circo
    
    # Calcular total líquido (receitas - despesas de cabeça)
    total_liquido = total_receitas - despesas_cabeca
    
    # Calcular 50% show
    cinquenta_porcento_show = total_liquido / 2
    
    # Calcular reembolso mídias (despesas de cabeça EXCLUINDO categoria "PAGAS PELO CIRCO")
    reembolso_midias = db.session.query(func.sum(DespesaEvento.valor)).join(
        Despesa, DespesaEvento.id_despesa == Despesa.id_despesa
    ).join(
        CategoriaDespesa, Despesa.id_categoria_despesa == CategoriaDespesa.id_categoria_despesa
    ).filter(
        DespesaEvento.id_evento == id_evento,
        DespesaEvento.despesa_cabeca == True,
        ~CategoriaDespesa.nome.ilike('%PAGAS PELO CIRCO%'),
        ~CategoriaDespesa.nome.ilike('%PAGO PELO CIRCO%')
    ).scalar() or 0
    
    # Repasse total = 50% show + reembolso mídias
    repasse_total = cinquenta_porcento_show + reembolso_midias
    
    # Total de todas as despesas Sócrates Online (EXCLUINDO categoria "PAGAS PELO CIRCO")
    total_despesas_socrates = db.session.query(func.sum(DespesaEvento.valor)).join(
        Despesa, DespesaEvento.id_despesa == Despesa.id_despesa
    ).join(
        CategoriaDespesa, Despesa.id_categoria_despesa == CategoriaDespesa.id_categoria_despesa
    ).filter(
        DespesaEvento.id_evento == id_evento,
        ~CategoriaDespesa.nome.ilike('%PAGAS PELO CIRCO%'),
        ~CategoriaDespesa.nome.ilike('%PAGO PELO CIRCO%')
    ).scalar() or 0
    
    # Resultado do show = repasse total - total despesas sócrates (LUCRO REAL)
    resultado_show = repasse_total - total_despesas_socrates
    
    # Total de todas as despesas para exibição
    total_despesas = db.session.query(func.sum(DespesaEvento.valor)).filter_by(id_evento=id_evento).scalar() or 0
    
    return {
        'total_receitas': float(total_receitas),
        'despesas_cabeca': float(despesas_cabeca),
        'total_liquido': float(total_liquido),
        'cinquenta_porcento_show': float(cinquenta_porcento_show),
        'reembolso_midias': float(reembolso_midias),
        'repasse_total': float(repasse_total),
        'total_despesas_socrates': float(total_despesas_socrates),
        'resultado_show': float(resultado_show),
        'total_despesas': float(total_despesas)
    }

def calcular_lucro_simples(id_evento):
    """
    Função simplificada que retorna apenas o lucro do evento.
    Útil para casos onde só precisamos do valor final.
    
    Returns:
        float: Lucro real do evento (resultado do show)
    """
    calculo = calcular_lucro_evento(id_evento)
    return calculo['resultado_show']

def obter_datas_filtro_padrao(data_inicio_param=None, data_fim_param=None):
    """
    Função utilitária para definir datas padrão dos filtros.
    Se não fornecidas, usa os últimos 90 dias como padrão.
    
    Args:
        data_inicio_param: Data de início fornecida pelo usuário
        data_fim_param: Data de fim fornecida pelo usuário
        
    Returns:
        tuple: (data_inicio, data_fim) como strings no formato YYYY-MM-DD
    """
    if data_inicio_param and data_fim_param:
        return data_inicio_param, data_fim_param
    
    # Definir padrão: últimos 90 dias
    data_fim = date.today()
    data_inicio = data_fim - timedelta(days=90)
    
    return data_inicio.strftime('%Y-%m-%d'), data_fim.strftime('%Y-%m-%d')

@app.route('/')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # Obter informações do usuário logado
    usuario = Usuario.query.get(session['user_id'])
    if not usuario:
        flash('Erro ao carregar informações do usuário.', 'danger')
        return redirect(url_for('login'))
    
    # Para o usuário root, permitir acesso mesmo sem colaborador
    if not usuario.colaborador and not is_root_user():
        flash('Erro ao carregar informações do colaborador.', 'danger')
        return redirect(url_for('login'))
    
    # Verificar se é administrador ou produtor
    is_admin = is_admin_user()
    # Para root sem colaborador, não verificar categorias
    if usuario.colaborador:
        is_produtor = any(cat.nome.lower() == 'produtor' for cat in usuario.colaborador.categorias)
    else:
        is_produtor = False

    # TEMPORÁRIO: Lógica de eventos comentada
    """
    # Filtro de datas da lista de eventos do dashboard
    eventos_period = request.args.get('eventos_period', '90dias')  # Mudança: padrão para 90 dias
    if eventos_period == 'hoje':
        eventos_data_inicio = eventos_data_fim = date.today()
    elif eventos_period == 'ontem':
        eventos_data_inicio = eventos_data_fim = date.today() - timedelta(days=1)
    elif eventos_period == '7dias':
        eventos_data_fim = date.today()
        eventos_data_inicio = eventos_data_fim - timedelta(days=6)
    elif eventos_period == '30dias':
        eventos_data_fim = date.today()
        eventos_data_inicio = eventos_data_fim - timedelta(days=30)
    elif eventos_period == '90dias':
        # Incluir eventos futuros também
        eventos_data_fim = date.today() + timedelta(days=365)  # Eventos futuros até 1 ano
        eventos_data_inicio = date.today() - timedelta(days=90)  # Eventos passados até 90 dias
    elif eventos_period == 'mes':
        eventos_data_fim = date.today()
        eventos_data_inicio = eventos_data_fim.replace(day=1)
    elif eventos_period == 'custom':
        eventos_data_inicio = request.args.get('eventos_data_inicio')
        eventos_data_fim = request.args.get('eventos_data_fim')
        if eventos_data_inicio and eventos_data_fim:
            eventos_data_inicio = datetime.strptime(eventos_data_inicio, '%Y-%m-%d').date()
            eventos_data_fim = datetime.strptime(eventos_data_fim, '%Y-%m-%d').date()
        else:
            # Se custom mas sem datas, usar 90 dias
            eventos_data_fim = date.today() + timedelta(days=365)
            eventos_data_inicio = date.today() - timedelta(days=90)
    else:
        # Default: últimos 90 dias + eventos futuros
        eventos_data_fim = date.today() + timedelta(days=365)
        eventos_data_inicio = date.today() - timedelta(days=90)

    # Filtrar eventos baseado no tipo de usuário
    eventos_query = Evento.query.filter(Evento.status.in_(['planejamento', 'a realizar', 'em andamento', 'realizado']))
    
    if not is_admin:
        # Produtores veem apenas seus eventos
        eventos_query = eventos_query.filter_by(id_produtor=usuario.colaborador.id_colaborador)
    
    if eventos_data_inicio:
        eventos_query = eventos_query.filter(Evento.data_inicio >= eventos_data_inicio)
    if eventos_data_fim:
        eventos_query = eventos_query.filter(Evento.data_inicio <= eventos_data_fim)
    
    eventos = eventos_query.order_by(Evento.data_inicio.desc()).all()
    """

    return render_template(
        'dashboard.html',
        is_admin=is_admin,
        is_produtor=is_produtor,
        usuario=usuario
    )

def is_root_user():
    """Verifica se o usuário atual é o root"""
    return session.get('email') == 'root@socratesonline.com'

def is_admin_user():
    """Verifica se o usuário atual é administrador ou root"""
    if is_root_user():
        return True
    
    user_id = session.get('user_id')
    if not user_id:
        return False
    
    usuario = Usuario.query.get(user_id)
    if not usuario:
        return False
    
    # Se não tem colaborador, não é admin (exceto root que já foi verificado acima)
    if not usuario.colaborador:
        return False
    
    return any(cat.nome.lower() == 'administrativo' for cat in usuario.colaborador.categorias)

def is_promotor_user():
    """Verifica se o usuário atual é promotor de vendas"""
    user_id = session.get('user_id')
    if not user_id:
        return False
    
    usuario = Usuario.query.get(user_id)
    if not usuario or not usuario.colaborador:
        return False
    
    # Verificar se tem categoria que contenha 'promotor'
    return any('promotor' in cat.nome.lower() for cat in usuario.colaborador.categorias)

def registrar_log(acao, descricao=None, usuario_id=None):
    """
    Registrar log de operação no sistema
    
    Args:
        acao (str): Ação realizada (ex: 'Login', 'Exclusão de Usuário')
        descricao (str, optional): Descrição detalhada da ação
        usuario_id (int/str, optional): ID do usuário. Se None, pega do session
    """
    try:
        # Se não foi passado usuario_id, pegar do usuário logado
        if usuario_id is None:
            usuario_id = session.get('user_id')
        
        # Se ainda não tiver usuário, não registrar log
        if not usuario_id:
            print("Aviso: Tentativa de registrar log sem usuário logado")
            return
        
        # Buscar dados do usuário no momento do log
        usuario = Usuario.query.get(int(usuario_id))
        if not usuario:
            print(f"Aviso: Usuário ID {usuario_id} não encontrado para log")
            return
        
        # Criar registro de log com dados independentes
        novo_log = LogSistema(
            acao=acao,
            descricao=descricao,
            usuario_id=str(usuario_id),
            usuario_nome=usuario.nome,
            usuario_email=usuario.email,
            data_hora=datetime.now()
        )
        
        db.session.add(novo_log)
        db.session.commit()
        
    except Exception as e:
        # Em caso de erro no log, não interromper o fluxo principal
        print(f"Erro ao registrar log: {e}")
        try:
            db.session.rollback()
        except:
            pass

def enviar_whatsapp(numero_destino, mensagem):
    """
    Envia mensagem via WhatsApp Business API
    
    Args:
        numero_destino (str): Número do WhatsApp (pode ter formatação)
        mensagem (str): Texto da mensagem
        
    Returns:
        dict: {'success': bool, 'message_id': str, 'error': str}
    """
    api_url = app.config.get('WHATSAPP_API_URL')
    access_token = app.config.get('WHATSAPP_API_TOKEN')
    
    if not api_url or not access_token:
        return {
            'success': False,
            'error': 'WhatsApp API não configurada. Configure WHATSAPP_API_URL e WHATSAPP_API_TOKEN no .env'
        }
    
    # Limpar o número (remover caracteres especiais)
    numero_limpo = ''.join(filter(str.isdigit, numero_destino))
    
    # Garantir formato internacional brasileiro
    if len(numero_limpo) == 11 and numero_limpo.startswith('11'):
        numero_limpo = '55' + numero_limpo
    elif len(numero_limpo) == 10:
        numero_limpo = '5511' + numero_limpo
    elif not numero_limpo.startswith('55'):
        numero_limpo = '55' + numero_limpo
    
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }
    
    payload = {
        'messaging_product': 'whatsapp',
        'to': numero_limpo,
        'type': 'text',
        'text': {
            'body': mensagem
        }
    }
    
    try:
        response = requests.post(api_url, json=payload, headers=headers, timeout=30)
        
        if response.status_code == 200:
            response_data = response.json()
            message_id = response_data.get('messages', [{}])[0].get('id', 'unknown')
            return {
                'success': True,
                'message_id': message_id,
                'numero_enviado': numero_limpo
            }
        else:
            return {
                'success': False,
                'error': f'Erro da API: {response.status_code} - {response.text}'
            }
            
    except requests.exceptions.RequestException as e:
        return {
            'success': False,
            'error': f'Erro de conexão: {str(e)}'
        }
    except Exception as e:
        return {
            'success': False,
            'error': f'Erro inesperado: {str(e)}'
        }

def criar_mensagem_visita_escola(escola, visita, promotor):
    """
    Cria mensagem padrão para visita à escola
    
    Args:
        escola: Objeto Escola
        visita: Objeto VisitaEscola  
        promotor: Objeto Colaborador (promotor)
        
    Returns:
        str: Mensagem formatada
    """
    empresa_nome = os.environ.get('EMPRESA_NOME', 'Sócrates Online')
    empresa_contato = os.environ.get('EMPRESA_CONTATO', 'contato@socratesonline.com')
    empresa_whatsapp = os.environ.get('EMPRESA_WHATSAPP', '5511999887766')
    
    # Formatar número da empresa para exibição
    whatsapp_formatado = empresa_whatsapp
    if len(empresa_whatsapp) == 13 and empresa_whatsapp.startswith('55'):
        # Formato: 5511999887766 -> (11) 99988-7766
        ddd = empresa_whatsapp[2:4]
        numero = empresa_whatsapp[4:]
        whatsapp_formatado = f"({ddd}) {numero[:5]}-{numero[5:]}"
    
    mensagem = f"""🎪 *{empresa_nome}*

Olá, {escola.nome}!

Temos o prazer de informar que foi agendada uma visita do nosso circo em sua escola.

📅 *Data da Visita:* {visita.data_visita.strftime('%d/%m/%Y às %H:%M')}
👤 *Promotor Responsável:* {promotor.nome}
📍 *Escola:* {escola.nome} - {escola.cidade}/{escola.estado}

Nossa equipe entrará em contato para alinhar todos os detalhes do espetáculo.

Para mais informações:
📧 {empresa_contato}
📱 {whatsapp_formatado}

Aguardamos vocês para um espetáculo inesquecível! 🎭✨

_Mensagem enviada automaticamente pelo sistema {empresa_nome}_"""
    
    return mensagem

def criar_usuario_root():
    """Criar usuário ROOT se não existir"""
    try:
        # Verificar se já existe usuário ROOT
        usuario_root = Usuario.query.filter_by(email='root@socratesonline.com').first()
        if usuario_root:
            return  # Usuário ROOT já existe
        
        print("🔄 Criando usuário ROOT do sistema...")
        
        # Criar categorias básicas se não existirem
        categorias_basicas = [
            'Administrativo',
            'Operacional', 
            'Promotor de Vendas',
            'Produtor',
            'Motorista',
            'Técnico'
        ]
        
        categorias_criadas = []
        for nome_categoria in categorias_basicas:
            categoria_existente = CategoriaColaborador.query.filter_by(nome=nome_categoria).first()
            if not categoria_existente:
                nova_categoria = CategoriaColaborador(nome=nome_categoria)
                db.session.add(nova_categoria)
                db.session.flush()  # Para obter o ID
                categorias_criadas.append(nova_categoria)
                print(f"📝 Categoria '{nome_categoria}' criada")
            else:
                categorias_criadas.append(categoria_existente)
        
        # Verificar se já existe colaborador ROOT
        colaborador_root = Colaborador.query.filter_by(email='root@socratesonline.com').first()
        if not colaborador_root:
            # Criar colaborador ROOT
            colaborador_root = Colaborador(
                nome='ROOT - Administrador do Sistema',
                telefone='(00) 00000-0000',
                email='root@socratesonline.com'
            )
            db.session.add(colaborador_root)
            db.session.flush()  # Para obter o ID do colaborador
            print("👤 Colaborador ROOT criado")
        
        # Verificar e associar todas as categorias ao colaborador ROOT
        categorias_associadas = [assoc.categoria for assoc in colaborador_root.categorias_associacao]
        for categoria in categorias_criadas:
            if categoria not in categorias_associadas:
                associacao = ColaboradorCategoria(
                    id_colaborador=colaborador_root.id_colaborador,
                    id_categoria_colaborador=categoria.id_categoria_colaborador
                )
                db.session.add(associacao)
        
        # Criar usuário ROOT
        senha_hash = generate_password_hash('Admin@2025')
        usuario_root = Usuario(
            nome='ROOT',
            email='root@socratesonline.com',
            senha_hash=senha_hash,
            id_colaborador=colaborador_root.id_colaborador
        )
        db.session.add(usuario_root)
        
        db.session.commit()
        print("✅ Usuário ROOT criado com sucesso!")
        print("📧 Email: root@socratesonline.com")
        print("🔒 Senha: Admin@2025")
        print("🏷️ Categorias:", ', '.join([cat.nome for cat in categorias_criadas]))
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao criar usuário ROOT: {e}")
        import traceback
        traceback.print_exc()

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Verificar e criar usuário ROOT se necessário
    criar_usuario_root()
    
    # Verificar se é o primeiro acesso (apenas usuário ROOT existe)
    total_usuarios = Usuario.query.count()
    primeiro_acesso = total_usuarios <= 1
    
    form = LoginForm()
    if form.validate_on_submit():
        usuario = Usuario.query.filter_by(email=form.email.data).first()
        if usuario and check_password_hash(usuario.senha_hash, form.password.data):
            session['user_id'] = usuario.id
            session['email'] = usuario.email  # Adicionar email na sessão para verificação de root
            # Usar a primeira categoria do colaborador para compatibilidade
            if usuario.colaborador and usuario.colaborador.categorias:
                session['categoria'] = usuario.colaborador.categorias[0].nome.lower()
            else:
                session['categoria'] = 'administrativo'  # padrão
            
            # Registrar log de login
            registrar_log('Login', f'Usuário "{usuario.nome}" fez login no sistema', usuario.id)
            
            return redirect(url_for('dashboard'))
        else:
            # Credenciais incorretas
            flash('Email ou senha incorretos. Tente novamente.', 'danger')
    elif form.errors:
        # Erros de validação do formulário
        flash('Por favor, verifique os dados informados.', 'warning')
    
    return render_template('login.html', form=form, primeiro_acesso=primeiro_acesso)

@app.route('/logout')
def logout():
    # Registrar log de logout antes de limpar a sessão
    user_id = session.get('user_id')
    if user_id:
        usuario = Usuario.query.get(user_id)
        if usuario:
            registrar_log('Logout', f'Usuário "{usuario.nome}" fez logout do sistema', user_id)
    
    session.pop('user_id', None)
    session.pop('categoria', None)
    session.pop('email', None)
    return redirect(url_for('login'))

@app.route('/auto-cadastro', methods=['GET', 'POST'])
def auto_cadastro():
    # Verificar se existem categorias (exceto administrativo)
    todas_categorias = CategoriaColaborador.query.all()
    categorias_existentes = [cat for cat in todas_categorias if 'administrativo' not in cat.nome.lower()]
    if not categorias_existentes:
        flash('Sistema indisponível no momento. Entre em contato com o administrador.', 'warning')
        return redirect(url_for('login'))
    
    form = AutoCadastroForm()
    form.categoria.choices = [(c.id_categoria_colaborador, c.nome) for c in categorias_existentes]
    
    if form.validate_on_submit():
        try:
            # Criar colaborador
            novo_colaborador = Colaborador(
                nome=form.nome.data,
                telefone=form.telefone.data,
                email=form.email.data
            )
            db.session.add(novo_colaborador)
            db.session.flush()  # Para obter o ID do colaborador
            
            # Associar categoria selecionada
            nova_associacao = ColaboradorCategoria(
                id_colaborador=novo_colaborador.id_colaborador,
                id_categoria_colaborador=form.categoria.data
            )
            db.session.add(nova_associacao)
            
            # Criar usuário de login
            hashed_password = generate_password_hash(form.password.data)
            novo_usuario = Usuario(
                nome=form.nome.data,
                email=form.email.data,
                senha_hash=hashed_password,
                id_colaborador=novo_colaborador.id_colaborador
            )
            db.session.add(novo_usuario)
            
            db.session.commit()
            
            flash('Cadastro realizado com sucesso! Você já pode fazer login.', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar cadastro: {str(e)}', 'danger')
    
    return render_template('auto_cadastro.html', form=form)

@app.route('/cadastros/circos', methods=['GET', 'POST'])
def cadastrar_circo():
    form = CircoForm()
    if form.validate_on_submit():
        novo = Circo(
            nome=form.nome.data,
            contato_responsavel=form.contato_responsavel.data,
            telefone_contato=form.telefone_contato.data,
            observacoes=form.observacoes.data
        )
        db.session.add(novo)
        db.session.commit()
        flash('Circo cadastrado com sucesso!', 'success')
        return redirect(url_for('cadastrar_circo'))
    circos = Circo.query.all()
    return render_template('cadastrar_circo.html', form=form, circos=circos)

@app.route('/cadastros/circos/editar/<int:id>', methods=['GET', 'POST'])
def editar_circo(id):
    circo = Circo.query.get_or_404(id)
    form = CircoForm(obj=circo)
    if form.validate_on_submit():
        circo.nome = form.nome.data
        circo.contato_responsavel = form.contato_responsavel.data
        circo.telefone_contato = form.telefone_contato.data
        circo.observacoes = form.observacoes.data
        db.session.commit()
        flash('Circo atualizado com sucesso!', 'success')
        return redirect(url_for('cadastrar_circo'))
    circos = Circo.query.all()
    return render_template('cadastrar_circo.html', form=form, circos=circos)

@app.route('/cadastros/circos/excluir/<int:id>')
def excluir_circo(id):
    circo = Circo.query.get_or_404(id)
    
    # Verificar se existem eventos usando este circo
    eventos_usando = Evento.query.filter_by(id_circo=id).count()
    if eventos_usando > 0:
        flash(f'Não é possível excluir este circo pois existem {eventos_usando} evento(s) associado(s) a ele.', 'danger')
        return redirect(url_for('cadastrar_circo'))
    
    db.session.delete(circo)
    db.session.commit()
    flash('Circo excluído com sucesso!', 'success')
    return redirect(url_for('cadastrar_circo'))

@app.route('/cadastros/categorias-colaborador', methods=['GET', 'POST'])
def cadastrar_categoria_colaborador():
    form = CategoriaColaboradorForm()
    if form.validate_on_submit():
        nova = CategoriaColaborador(nome=form.nome.data)
        db.session.add(nova)
        db.session.commit()
        flash('Categoria cadastrada com sucesso!', 'success')
        return redirect(url_for('cadastrar_categoria_colaborador'))
    categorias = CategoriaColaborador.query.all()
    return render_template('categorias_colaborador.html', form=form, categorias=categorias)

@app.route('/cadastros/colaboradores', methods=['GET', 'POST'])
def cadastrar_colaborador():
    # Verificar se existem categorias
    categorias_existentes = CategoriaColaborador.query.all()
    if not categorias_existentes:
        flash('É necessário cadastrar pelo menos uma categoria de colaborador antes de cadastrar colaboradores.', 'warning')
        return redirect(url_for('cadastrar_categoria_colaborador'))
    
    form = ColaboradorForm()
    form.categorias.choices = [(c.id_categoria_colaborador, c.nome) for c in categorias_existentes]
    
    if request.method == 'POST':
        print(f"DEBUG: Dados recebidos - Nome: {form.nome.data}, Email: {form.email.data}, Telefone: {form.telefone.data}")
        print(f"DEBUG: Categorias selecionadas: {form.categorias.data}")
        print(f"DEBUG: Formulário válido: {form.validate_on_submit()}")
        
        if form.errors:
            print(f"DEBUG: Erros do formulário: {form.errors}")
        
    if form.validate_on_submit():
        try:
            print("DEBUG: Iniciando criação do colaborador...")
            
            # Limpar e validar dados antes de salvar
            nome_limpo = form.nome.data.strip() if form.nome.data else ""
            telefone_limpo = form.telefone.data.strip() if form.telefone.data else None
            email_limpo = form.email.data.strip().lower() if form.email.data else ""
            
            if not nome_limpo:
                flash('Nome é obrigatório.', 'danger')
                return render_template('colaboradores.html', form=form, colaboradores=Colaborador.query.all())
            
            if not email_limpo:
                flash('Email é obrigatório.', 'danger')
                return render_template('colaboradores.html', form=form, colaboradores=Colaborador.query.all())
            
            if not form.categorias.data:
                flash('É necessário selecionar pelo menos uma categoria.', 'danger')
                return render_template('colaboradores.html', form=form, colaboradores=Colaborador.query.all())
            
            print(f"DEBUG: Dados limpos - Nome: {nome_limpo}, Email: {email_limpo}, Telefone: {telefone_limpo}")
            
            novo = Colaborador(nome=nome_limpo, telefone=telefone_limpo, email=email_limpo)
            db.session.add(novo)
            db.session.flush()  # Para obter o ID do colaborador
            print(f"DEBUG: Colaborador criado com ID: {novo.id_colaborador}")
            
            # Adicionar as categorias selecionadas
            categorias_adicionadas = 0
            for categoria_id in form.categorias.data:
                print(f"DEBUG: Adicionando categoria {categoria_id}")
                nova_associacao = ColaboradorCategoria(
                    id_colaborador=novo.id_colaborador,
                    id_categoria_colaborador=categoria_id
                )
                db.session.add(nova_associacao)
                categorias_adicionadas += 1
            
            print(f"DEBUG: {categorias_adicionadas} categorias adicionadas")
            
            # Criar usuário automaticamente se senha foi fornecida
            if form.password.data and form.password.data.strip():
                print("DEBUG: Criando usuário para o colaborador...")
                hashed_password = generate_password_hash(form.password.data)
                novo_usuario = Usuario(
                    nome=nome_limpo,
                    email=email_limpo,
                    senha_hash=hashed_password,
                    id_colaborador=novo.id_colaborador
                )
                db.session.add(novo_usuario)
                print("DEBUG: Usuário criado!")
            
            db.session.commit()
            print("DEBUG: Commit realizado com sucesso!")
            flash('Colaborador cadastrado com sucesso!', 'success')
            # Redirecionar para limpar o formulário
            return redirect(url_for('cadastrar_colaborador'))
            
        except Exception as e:
            print(f"DEBUG: Erro ao salvar colaborador: {str(e)}")
            import traceback
            traceback.print_exc()
            db.session.rollback()
            flash(f'Erro ao cadastrar colaborador: {str(e)}', 'danger')
    
    colaboradores = Colaborador.query.all()
    return render_template('colaboradores.html', form=form, colaboradores=colaboradores)

@app.route('/cadastros/categorias-colaborador/editar/<int:id>', methods=['GET', 'POST'])
def editar_categoria_colaborador(id):
    categoria = CategoriaColaborador.query.get_or_404(id)
    form = CategoriaColaboradorForm(obj=categoria)
    if form.validate_on_submit():
        categoria.nome = form.nome.data
        db.session.commit()
        flash('Categoria atualizada com sucesso!', 'success')
        return redirect(url_for('cadastrar_categoria_colaborador'))
    categorias = CategoriaColaborador.query.all()
    return render_template('categorias_colaborador.html', form=form, categorias=categorias)

@app.route('/cadastros/categorias-colaborador/excluir/<int:id>')
def excluir_categoria_colaborador(id):
    categoria = CategoriaColaborador.query.get_or_404(id)
    
    # Verificar se existem colaboradores usando esta categoria
    colaboradores_usando = ColaboradorCategoria.query.filter_by(id_categoria_colaborador=id).count()
    if colaboradores_usando > 0:
        flash(f'Não é possível excluir esta categoria pois existem {colaboradores_usando} colaborador(es) associado(s) a ela.', 'danger')
        return redirect(url_for('cadastrar_categoria_colaborador'))
    
    db.session.delete(categoria)
    db.session.commit()
    flash('Categoria excluída com sucesso!', 'success')
    return redirect(url_for('cadastrar_categoria_colaborador'))

@app.route('/cadastros/colaboradores/editar/<int:id>', methods=['GET', 'POST'])
def editar_colaborador(id):
    # Verificar se existem categorias
    categorias_existentes = CategoriaColaborador.query.all()
    if not categorias_existentes:
        flash('É necessário cadastrar pelo menos uma categoria de colaborador antes de editar colaboradores.', 'warning')
        return redirect(url_for('cadastrar_categoria_colaborador'))
    
    colaborador = Colaborador.query.get_or_404(id)
    form = ColaboradorForm(colaborador_id=id, is_edit_mode=True)
    
    # Configurar as choices sempre, antes de qualquer validação
    form.categorias.choices = [(c.id_categoria_colaborador, c.nome) for c in categorias_existentes]
    
    # Verificar se existe usuário associado ao colaborador
    usuario_associado = Usuario.query.filter_by(id_colaborador=id).first()
    email_original = colaborador.email
    
    # Se for GET, preencher com os dados atuais do colaborador
    if request.method == 'GET':
        form.nome.data = colaborador.nome
        form.telefone.data = colaborador.telefone
        form.email.data = colaborador.email
        # Pré-selecionar as categorias atuais do colaborador
        categorias_atuais = [assoc.id_categoria_colaborador for assoc in colaborador.categorias_associacao]
        form.categorias.data = categorias_atuais
    
    if form.validate_on_submit():
        email_novo = form.email.data
        email_mudou = email_original != email_novo
        
        # Se o email mudou e existe usuário associado, verificar confirmação
        if email_mudou and usuario_associado:
            confirmacao = request.form.get('confirmar_sync_email')
            if not confirmacao:
                # Mostrar tela de confirmação
                colaboradores = Colaborador.query.all()
                return render_template('colaboradores.html', 
                                     form=form, 
                                     colaboradores=colaboradores,
                                     mostrar_confirmacao_email=True,
                                     colaborador_editando=colaborador,
                                     usuario_associado=usuario_associado,
                                     email_novo=email_novo)
        
        # Atualizar dados do colaborador
        colaborador.nome = form.nome.data
        colaborador.telefone = form.telefone.data
        colaborador.email = email_novo
        
        # Se email mudou e existe usuário associado, sincronizar
        if email_mudou and usuario_associado:
            usuario_associado.email = email_novo
        
        # Obter categorias selecionadas do formulário
        categorias_selecionadas = form.categorias.data
        
        # Se o campo do formulário estiver vazio, tentar pegar do request.form
        if not categorias_selecionadas:
            categorias_selecionadas = request.form.getlist('categorias')
            categorias_selecionadas = [int(cat_id) for cat_id in categorias_selecionadas if cat_id.isdigit()]
        
        if not categorias_selecionadas:
            flash('É necessário selecionar pelo menos uma categoria.', 'danger')
            colaboradores = Colaborador.query.all()
            return render_template('colaboradores.html', form=form, colaboradores=colaboradores)
        
        try:
            # Remover todas as associações existentes
            ColaboradorCategoria.query.filter_by(id_colaborador=colaborador.id_colaborador).delete()
            
            # Adicionar as novas categorias selecionadas
            for categoria_id in categorias_selecionadas:
                nova_associacao = ColaboradorCategoria(
                    id_colaborador=colaborador.id_colaborador,
                    id_categoria_colaborador=categoria_id
                )
                db.session.add(nova_associacao)
            
            db.session.commit()
            
            # Registrar log da operação
            registrar_log('Editar Colaborador', f'Colaborador "{colaborador.nome}" editado')
            
            if email_mudou and usuario_associado:
                flash(f'Colaborador atualizado com sucesso! O email do usuário de login também foi atualizado para {email_novo}.', 'success')
            else:
                flash('Colaborador atualizado com sucesso!', 'success')
            return redirect(url_for('cadastrar_colaborador'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar colaborador: {str(e)}', 'danger')
    else:
        # Se o formulário não validou e é POST, mostrar erros
        if request.method == 'POST' and form.errors:
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f'Erro no campo {field}: {error}', 'danger')
    
    colaboradores = Colaborador.query.all()
    return render_template('colaboradores.html', form=form, colaboradores=colaboradores, colaborador_editando=colaborador)

@app.route('/cadastros/colaboradores/excluir/<int:id>')
def excluir_colaborador(id):
    colaborador = Colaborador.query.get_or_404(id)
    
    # Verificar se é produtor de eventos
    eventos_produzindo = Evento.query.filter_by(id_produtor=id).count()
    if eventos_produzindo > 0:
        flash(f'Não é possível excluir este colaborador pois ele é produtor de {eventos_produzindo} evento(s).', 'danger')
        return redirect(url_for('cadastrar_colaborador'))
    
    # Verificar se participa de equipe de evento
    equipe_usando = EquipeEvento.query.filter_by(id_colaborador=id).count()
    if equipe_usando > 0:
        flash(f'Não é possível excluir este colaborador pois ele está vinculado a {equipe_usando} registro(s) de equipe em evento(s).', 'danger')
        return redirect(url_for('cadastrar_colaborador'))
    
    # Verificar se está vinculado como motorista em algum uso de veículo
    motorista_usando = VeiculoEvento.query.filter_by(id_motorista=id).count()
    if motorista_usando > 0:
        flash(f'Não é possível excluir este colaborador pois ele é motorista em {motorista_usando} registro(s) de veículo em evento(s).', 'danger')
        return redirect(url_for('cadastrar_colaborador'))
    
    # Verificar se tem usuário associado
    usuario_associado = Usuario.query.filter_by(id_colaborador=id).first()
    if usuario_associado:
        flash(f'Não é possível excluir este colaborador pois ele possui um usuário associado. Exclua o usuário primeiro.', 'danger')
        return redirect(url_for('cadastrar_colaborador'))
    
    # Registrar log da operação antes da exclusão
    nome_colaborador = colaborador.nome
    registrar_log('Excluir Colaborador', f'Colaborador "{nome_colaborador}" excluído')
    
    db.session.delete(colaborador)
    db.session.commit()
    flash('Colaborador excluído com sucesso!', 'success')
    return redirect(url_for('cadastrar_colaborador'))

@app.route('/colaboradores/<int:id>/criar-usuario', methods=['GET', 'POST'])
def criar_usuario(id):
    if session.get('categoria', '').lower() != 'administrativo':
        flash('Apenas administradores podem criar usuários.', 'danger')
        return redirect(url_for('cadastrar_colaborador'))
    
    colaborador = Colaborador.query.get_or_404(id)
    
    # Verificar se o colaborador tem email cadastrado
    if not colaborador.email:
        flash(f'O colaborador {colaborador.nome} não possui email cadastrado. Edite o colaborador e adicione um email primeiro.', 'warning')
        return redirect(url_for('cadastrar_colaborador'))
    
    # Verificar se já existe usuário para este colaborador
    usuario_existente = Usuario.query.filter_by(id_colaborador=id).first()
    if usuario_existente:
        flash(f'Já existe um usuário para o colaborador {colaborador.nome}.', 'warning')
        return redirect(url_for('cadastrar_colaborador'))
    
    form = UsuarioForm(colaborador_email=colaborador.email)
    
    # Pré-preencher o nome com o nome do colaborador
    if request.method == 'GET':
        form.nome.data = colaborador.nome
        form.email.data = colaborador.email
    
    if form.validate_on_submit():
        # Verificar se email já existe (usando email do colaborador)
        email_existente = Usuario.query.filter_by(email=colaborador.email).first()
        if email_existente:
            flash('Este email já está sendo usado por outro usuário.', 'danger')
            return render_template('criar_usuario.html', form=form, colaborador=colaborador)
        
        # Validar se as senhas coincidem
        if form.password.data != form.confirm_password.data:
            flash('As senhas não coincidem.', 'danger')
            return render_template('criar_usuario.html', form=form, colaborador=colaborador)
        
        # Criar usuário
        hashed_password = generate_password_hash(form.password.data)
        novo_usuario = Usuario(
            nome=form.nome.data,
            email=colaborador.email,  # Usar email do colaborador
            senha_hash=hashed_password,
            id_colaborador=colaborador.id_colaborador
        )
        
        db.session.add(novo_usuario)
        db.session.commit()
        flash(f'Usuário criado com sucesso para o colaborador {colaborador.nome}!', 'success')
        return redirect(url_for('cadastrar_colaborador'))
    
    return render_template('criar_usuario.html', form=form, colaborador=colaborador)

@app.route('/colaboradores/<int:id>/editar-usuario', methods=['GET', 'POST'])
def editar_usuario(id):
    if session.get('categoria', '').lower() != 'administrativo':
        flash('Apenas administradores podem editar usuários.', 'danger')
        return redirect(url_for('cadastrar_colaborador'))
    
    colaborador = Colaborador.query.get_or_404(id)
    usuario = Usuario.query.filter_by(id_colaborador=id).first()
    
    if not usuario:
        flash(f'Não existe usuário para o colaborador {colaborador.nome}.', 'warning')
        return redirect(url_for('cadastrar_colaborador'))
    
    form = UsuarioForm(obj=usuario, is_edit=True)
    
    if form.validate_on_submit():
        # Verificar se email já existe (exceto o próprio usuário)
        email_existente = Usuario.query.filter(Usuario.email == form.email.data, Usuario.id != usuario.id).first()
        if email_existente:
            flash('Este email já está sendo usado por outro usuário.', 'danger')
            return render_template('editar_usuario.html', form=form, colaborador=colaborador, usuario=usuario)
        
        # Atualizar usuário
        usuario.nome = form.nome.data
        usuario.email = form.email.data
        if form.password.data:  # Só atualiza senha se foi preenchida
            usuario.senha_hash = generate_password_hash(form.password.data)
        
        db.session.commit()
        flash(f'Usuário do colaborador {colaborador.nome} atualizado com sucesso!', 'success')
        return redirect(url_for('cadastrar_colaborador'))
    
    return render_template('editar_usuario.html', form=form, colaborador=colaborador, usuario=usuario)

@app.route('/colaboradores/<int:id>/excluir-usuario')
def excluir_usuario(id):
    if session.get('categoria', '').lower() != 'administrativo':
        flash('Apenas administradores podem excluir usuários.', 'danger')
        return redirect(url_for('cadastrar_colaborador'))
    
    colaborador = Colaborador.query.get_or_404(id)
    usuario = Usuario.query.filter_by(id_colaborador=id).first()
    
    if not usuario:
        flash(f'Não existe usuário para o colaborador {colaborador.nome}.', 'warning')
        return redirect(url_for('cadastrar_colaborador'))
    
    try:
        # Registrar log de exclusão antes de excluir (dados ficam independentes)
        registrar_log('Exclusão de Usuário', 
                     f'Usuário "{usuario.nome}" (ID: {usuario.id}) do colaborador "{colaborador.nome}" foi excluído')
        
        # Excluir o usuário (logs são independentes, não há problema de FK)
        db.session.delete(usuario)
        db.session.commit()
        
        flash(f'Usuário do colaborador {colaborador.nome} excluído com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao excluir usuário: {e}")
        flash(f'Erro ao excluir usuário: {str(e)}', 'danger')
    
    return redirect(url_for('cadastrar_colaborador'))

@app.route('/cadastros/elenco', methods=['GET', 'POST'])
def cadastrar_elenco():
    form = ElencoForm()
    if form.validate_on_submit():
        novo = Elenco(
            nome=form.nome.data,
            cpf=form.cpf.data,
            endereco=form.endereco.data,
            cidade=form.cidade.data,
            estado=form.estado.data,
            telefone=form.telefone.data,
            email=form.email.data,
            observacoes=form.observacoes.data
        )
        db.session.add(novo)
        db.session.commit()
        flash('Integrante do elenco cadastrado com sucesso!', 'success')
        return redirect(url_for('cadastrar_elenco'))
    elenco = Elenco.query.all()
    return render_template('elenco.html', form=form, elenco=elenco)

@app.route('/cadastros/elenco/editar/<int:id>', methods=['GET', 'POST'])
def editar_elenco(id):
    integrante = Elenco.query.get_or_404(id)
    form = ElencoForm(obj=integrante)
    if form.validate_on_submit():
        integrante.nome = form.nome.data
        integrante.cpf = form.cpf.data
        integrante.endereco = form.endereco.data
        integrante.cidade = form.cidade.data
        integrante.estado = form.estado.data
        integrante.telefone = form.telefone.data
        integrante.email = form.email.data
        integrante.observacoes = form.observacoes.data
        db.session.commit()
        flash('Integrante atualizado com sucesso!', 'success')
        return redirect(url_for('cadastrar_elenco'))
    elenco = Elenco.query.all()
    return render_template('elenco.html', form=form, elenco=elenco)

@app.route('/cadastros/elenco/excluir/<int:id>')
def excluir_elenco(id):
    integrante = Elenco.query.get_or_404(id)
    db.session.delete(integrante)
    db.session.commit()
    flash('Integrante excluído com sucesso!', 'success')
    return redirect(url_for('cadastrar_elenco'))

@app.route('/cadastros/categorias-fornecedor', methods=['GET', 'POST'])
def cadastrar_categoria_fornecedor():
    form = CategoriaFornecedorForm()
    if form.validate_on_submit():
        nova = CategoriaFornecedor(nome=form.nome.data)
        db.session.add(nova)
        db.session.commit()
        flash('Categoria cadastrada com sucesso!', 'success')
        return redirect(url_for('cadastrar_categoria_fornecedor'))
    categorias = CategoriaFornecedor.query.all()
    return render_template('categorias_fornecedor.html', form=form, categorias=categorias)

@app.route('/cadastros/fornecedores', methods=['GET', 'POST'])
def cadastrar_fornecedor():
    # Verificar se existem categorias
    categorias_existentes = CategoriaFornecedor.query.all()
    if not categorias_existentes:
        flash('É necessário cadastrar pelo menos uma categoria de fornecedor antes de cadastrar fornecedores.', 'warning')
        return redirect(url_for('cadastrar_categoria_fornecedor'))
    
    form = FornecedorForm()
    form.id_categoria_fornecedor.choices = [(c.id_categoria_fornecedor, c.nome) for c in categorias_existentes]
    if form.validate_on_submit():
        novo = Fornecedor(
            nome=form.nome.data,
            telefone=form.telefone.data,
            cidade=form.cidade.data,
            estado=form.estado.data,
            id_categoria_fornecedor=form.id_categoria_fornecedor.data
        )
        db.session.add(novo)
        db.session.commit()
        flash('Fornecedor cadastrado com sucesso!', 'success')
        return redirect(url_for('cadastrar_fornecedor'))
    fornecedores = Fornecedor.query.all()
    return render_template('fornecedores.html', form=form, fornecedores=fornecedores)

@app.route('/cadastros/categorias-fornecedor/editar/<int:id>', methods=['GET', 'POST'])
def editar_categoria_fornecedor(id):
    categoria = CategoriaFornecedor.query.get_or_404(id)
    form = CategoriaFornecedorForm(obj=categoria)
    if form.validate_on_submit():
        categoria.nome = form.nome.data
        db.session.commit()
        flash('Categoria atualizada com sucesso!', 'success')
        return redirect(url_for('cadastrar_categoria_fornecedor'))
    categorias = CategoriaFornecedor.query.all()
    return render_template('categorias_fornecedor.html', form=form, categorias=categorias)

@app.route('/cadastros/categorias-fornecedor/excluir/<int:id>')
def excluir_categoria_fornecedor(id):
    categoria = CategoriaFornecedor.query.get_or_404(id)
    
    # Verificar se existem fornecedores usando esta categoria
    fornecedores_usando = Fornecedor.query.filter_by(id_categoria_fornecedor=id).count()
    if fornecedores_usando > 0:
        flash(f'Não é possível excluir esta categoria pois existem {fornecedores_usando} fornecedor(es) associado(s) a ela.', 'danger')
        return redirect(url_for('cadastrar_categoria_fornecedor'))
    
    db.session.delete(categoria)
    db.session.commit()
    flash('Categoria excluída com sucesso!', 'success')
    return redirect(url_for('cadastrar_categoria_fornecedor'))

@app.route('/cadastros/fornecedores/editar/<int:id>', methods=['GET', 'POST'])
def editar_fornecedor(id):
    # Verificar se existem categorias
    categorias_existentes = CategoriaFornecedor.query.all()
    if not categorias_existentes:
        flash('É necessário cadastrar pelo menos uma categoria de fornecedor antes de editar fornecedores.', 'warning')
        return redirect(url_for('cadastrar_categoria_fornecedor'))
    
    fornecedor = Fornecedor.query.get_or_404(id)
    form = FornecedorForm(obj=fornecedor)
    form.id_categoria_fornecedor.choices = [(c.id_categoria_fornecedor, c.nome) for c in categorias_existentes]
    if form.validate_on_submit():
        fornecedor.nome = form.nome.data
        fornecedor.telefone = form.telefone.data
        fornecedor.cidade = form.cidade.data
        fornecedor.estado = form.estado.data
        fornecedor.id_categoria_fornecedor = form.id_categoria_fornecedor.data
        db.session.commit()
        flash('Fornecedor atualizado com sucesso!', 'success')
        return redirect(url_for('cadastrar_fornecedor'))
    fornecedores = Fornecedor.query.all()
    return render_template('fornecedores.html', form=form, fornecedores=fornecedores)

@app.route('/cadastros/fornecedores/excluir/<int:id>')
def excluir_fornecedor(id):
    fornecedor = Fornecedor.query.get_or_404(id)
    # Fornecedores normalmente não têm dependências diretas no sistema atual
    # Mas se houver tabelas de relacionamento no futuro, adicionar aqui
    db.session.delete(fornecedor)
    db.session.commit()
    flash('Fornecedor excluído com sucesso!', 'success')
    return redirect(url_for('cadastrar_fornecedor'))

@app.route('/cadastros/receitas', methods=['GET', 'POST'])
def cadastrar_receita():
    # Verificar se existem categorias
    categorias_existentes = CategoriaReceita.query.all()
    if not categorias_existentes:
        flash('É necessário cadastrar pelo menos uma categoria de receita antes de cadastrar receitas.', 'warning')
        return redirect(url_for('cadastrar_categoria_receita'))
    
    form = ReceitaForm()
    form.id_categoria_receita.choices = [(c.id_categoria_receita, c.nome) for c in categorias_existentes]
    if form.validate_on_submit():
        nova = Receita(nome=form.nome.data, id_categoria_receita=form.id_categoria_receita.data)
        db.session.add(nova)
        db.session.commit()
        flash('Receita cadastrada com sucesso!', 'success')
        return redirect(url_for('cadastrar_receita'))
    receitas = Receita.query.all()
    return render_template('receitas.html', form=form, receitas=receitas)

@app.route('/cadastros/receitas/editar/<int:id>', methods=['GET', 'POST'])
def editar_receita(id):
    # Verificar se existem categorias
    categorias_existentes = CategoriaReceita.query.all()
    if not categorias_existentes:
        flash('É necessário cadastrar pelo menos uma categoria de receita antes de editar receitas.', 'warning')
        return redirect(url_for('cadastrar_categoria_receita'))
    
    receita = Receita.query.get_or_404(id)
    form = ReceitaForm(obj=receita)
    form.id_categoria_receita.choices = [(c.id_categoria_receita, c.nome) for c in categorias_existentes]
    if form.validate_on_submit():
        receita.nome = form.nome.data
        receita.id_categoria_receita = form.id_categoria_receita.data
        db.session.commit()
        flash('Receita atualizada com sucesso!', 'success')
        return redirect(url_for('cadastrar_receita'))
    receitas = Receita.query.all()
    return render_template('receitas.html', form=form, receitas=receitas)

@app.route('/cadastros/receitas/excluir/<int:id>')
def excluir_receita(id):
    receita = Receita.query.get_or_404(id)
    
    # Verificar se existem eventos usando esta receita
    eventos_usando = ReceitaEvento.query.filter_by(id_receita=id).count()
    if eventos_usando > 0:
        flash(f'Não é possível excluir esta receita pois existem {eventos_usando} registro(s) de receita em evento(s) associado(s) a ela.', 'danger')
        return redirect(url_for('cadastrar_receita'))
    
    db.session.delete(receita)
    db.session.commit()
    flash('Receita excluída com sucesso!', 'success')
    return redirect(url_for('cadastrar_receita'))

@app.route('/cadastros/categorias-receita', methods=['GET', 'POST'])
def cadastrar_categoria_receita():
    form = CategoriaReceitaForm()
    if form.validate_on_submit():
        nova = CategoriaReceita(nome=form.nome.data)
        db.session.add(nova)
        db.session.commit()
        flash('Categoria cadastrada com sucesso!', 'success')
        return redirect(url_for('cadastrar_categoria_receita'))
    categorias = CategoriaReceita.query.all()
    return render_template('categorias_receita.html', form=form, categorias=categorias)

@app.route('/cadastros/categorias-receita/editar/<int:id>', methods=['GET', 'POST'])
def editar_categoria_receita(id):
    categoria = CategoriaReceita.query.get_or_404(id)
    form = CategoriaReceitaForm(obj=categoria)
    if form.validate_on_submit():
        categoria.nome = form.nome.data
        db.session.commit()
        flash('Categoria atualizada com sucesso!', 'success')
        return redirect(url_for('cadastrar_categoria_receita'))
    categorias = CategoriaReceita.query.all()
    return render_template('categorias_receita.html', form=form, categorias=categorias)

@app.route('/cadastros/categorias-receita/excluir/<int:id>')
def excluir_categoria_receita(id):
    categoria = CategoriaReceita.query.get_or_404(id)
    
    # Verificar se existem receitas usando esta categoria
    receitas_usando = Receita.query.filter_by(id_categoria_receita=id).count()
    if receitas_usando > 0:
        flash(f'Não é possível excluir esta categoria pois existem {receitas_usando} receita(s) associada(s) a ela.', 'danger')
        return redirect(url_for('cadastrar_categoria_receita'))
    
    db.session.delete(categoria)
    db.session.commit()
    flash('Categoria excluída com sucesso!', 'success')
    return redirect(url_for('cadastrar_categoria_receita'))

@app.route('/cadastros/categorias-despesa', methods=['GET', 'POST'])
def cadastrar_categoria_despesa():
    form = CategoriaDespesaForm()
    if form.validate_on_submit():
        nova = CategoriaDespesa(nome=form.nome.data)
        db.session.add(nova)
        db.session.commit()
        flash('Categoria cadastrada com sucesso!', 'success')
        return redirect(url_for('cadastrar_categoria_despesa'))
    categorias = CategoriaDespesa.query.all()
    return render_template('categorias_despesa.html', form=form, categorias=categorias)

@app.route('/cadastros/categorias-despesa/editar/<int:id>', methods=['GET', 'POST'])
def editar_categoria_despesa(id):
    categoria = CategoriaDespesa.query.get_or_404(id)
    form = CategoriaDespesaForm(obj=categoria)
    if form.validate_on_submit():
        categoria.nome = form.nome.data
        db.session.commit()
        flash('Categoria atualizada com sucesso!', 'success')
        return redirect(url_for('cadastrar_categoria_despesa'))
    categorias = CategoriaDespesa.query.all()
    return render_template('categorias_despesa.html', form=form, categorias=categorias)

@app.route('/cadastros/categorias-despesa/excluir/<int:id>')
def excluir_categoria_despesa(id):
    categoria = CategoriaDespesa.query.get_or_404(id)
    
    # Proteger categoria "PAGAS PELO CIRCO" e "PAGO PELO CIRCO"
    categoria_nome_upper = categoria.nome.upper()
    if 'PAGAS PELO CIRCO' in categoria_nome_upper or 'PAGO PELO CIRCO' in categoria_nome_upper:
        flash('Não é possível excluir esta categoria pois ela é protegida pelo sistema.', 'danger')
        return redirect(url_for('cadastrar_categoria_despesa'))
    
    # Verificar se existem despesas usando esta categoria
    despesas_usando = Despesa.query.filter_by(id_categoria_despesa=id).count()
    if despesas_usando > 0:
        flash(f'Não é possível excluir esta categoria pois existem {despesas_usando} despesa(s) associada(s) a ela.', 'danger')
        return redirect(url_for('cadastrar_categoria_despesa'))
    
    db.session.delete(categoria)
    db.session.commit()
    flash('Categoria excluída com sucesso!', 'success')
    return redirect(url_for('cadastrar_categoria_despesa'))

@app.route('/cadastros/parametros', methods=['GET', 'POST'])
def cadastrar_parametro():
    form = ParametroForm()
    if form.validate_on_submit():
        novo = Parametro(
            parametro=form.parametro.data,
            valor=form.valor.data,
            observacoes=form.observacoes.data
        )
        db.session.add(novo)
        db.session.commit()
        flash('Parâmetro cadastrado com sucesso!', 'success')
        return redirect(url_for('cadastrar_parametro'))
    parametros = Parametro.query.all()
    return render_template('parametros.html', form=form, parametros=parametros)

@app.route('/cadastros/parametros/editar/<int:id>', methods=['GET', 'POST'])
def editar_parametro(id):
    parametro = Parametro.query.get_or_404(id)
    form = ParametroForm(obj=parametro)
    if form.validate_on_submit():
        parametro.parametro = form.parametro.data
        parametro.valor = form.valor.data
        parametro.observacoes = form.observacoes.data
        db.session.commit()
        flash('Parâmetro atualizado com sucesso!', 'success')
        return redirect(url_for('cadastrar_parametro'))
    parametros = Parametro.query.all()
    return render_template('parametros.html', form=form, parametros=parametros)

@app.route('/cadastros/parametros/excluir/<int:id>')
def excluir_parametro(id):
    parametro = Parametro.query.get_or_404(id)
    
    # Verificar se é o parâmetro 'gasolina' (não pode ser excluído)
    if parametro.parametro.lower() == 'gasolina':
        flash('O parâmetro "gasolina" não pode ser excluído.', 'danger')
        return redirect(url_for('cadastrar_parametro'))
    
    db.session.delete(parametro)
    db.session.commit()
    flash('Parâmetro excluído com sucesso!', 'success')
    return redirect(url_for('cadastrar_parametro'))

@app.route('/cadastros/escolas', methods=['GET', 'POST'])
def cadastrar_escola():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    form = EscolaForm()
    if form.validate_on_submit():
        nova = Escola(
            nome=form.nome.data,
            endereco=form.endereco.data,
            cidade=form.cidade.data,
            estado=form.estado.data,
            email=form.email.data,
            whatsapp=form.whatsapp.data,
            nome_contato=form.nome_contato.data,
            cargo_contato=form.cargo_contato.data,
            observacoes=form.observacoes.data
        )
        db.session.add(nova)
        db.session.commit()
        
        # Registrar log da operação
        registrar_log('Cadastrar Escola', f'Escola "{nova.nome}" cadastrada - {nova.cidade}/{nova.estado}')
        
        flash('Escola cadastrada com sucesso!', 'success')
        return redirect(url_for('cadastrar_escola'))
    
    escolas = Escola.query.order_by(Escola.nome).all()
    return render_template('escolas.html', form=form, escolas=escolas)

@app.route('/cadastros/escolas/editar/<int:id>', methods=['GET', 'POST'])
def editar_escola(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    escola = Escola.query.get_or_404(id)
    form = EscolaForm(obj=escola)
    
    if form.validate_on_submit():
        escola.nome = form.nome.data
        escola.endereco = form.endereco.data
        escola.cidade = form.cidade.data
        escola.estado = form.estado.data
        escola.email = form.email.data
        escola.whatsapp = form.whatsapp.data
        escola.nome_contato = form.nome_contato.data
        escola.cargo_contato = form.cargo_contato.data
        escola.observacoes = form.observacoes.data
        
        db.session.commit()
        
        # Registrar log da operação
        registrar_log('Editar Escola', f'Escola "{escola.nome}" editada - {escola.cidade}/{escola.estado}')
        
        flash('Escola atualizada com sucesso!', 'success')
        return redirect(url_for('cadastrar_escola'))
    
    escolas = Escola.query.order_by(Escola.nome).all()
    return render_template('escolas.html', form=form, escolas=escolas, escola_editando=escola)

@app.route('/cadastros/escolas/excluir/<int:id>')
def excluir_escola(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    escola = Escola.query.get_or_404(id)
    
    # Verificar se existem visitas associadas
    visitas_count = VisitaEscola.query.filter_by(id_escola=id).count()
    if visitas_count > 0:
        flash(f'Não é possível excluir esta escola pois existem {visitas_count} visita(s) associada(s) a ela.', 'danger')
        return redirect(url_for('cadastrar_escola'))
    
    # Salvar dados para o log antes de excluir
    nome_escola = escola.nome
    cidade_escola = escola.cidade
    estado_escola = escola.estado
    
    db.session.delete(escola)
    db.session.commit()
    
    # Registrar log da operação
    registrar_log('Excluir Escola', f'Escola "{nome_escola}" excluída - {cidade_escola}/{estado_escola}')
    
    flash('Escola excluída com sucesso!', 'success')
    return redirect(url_for('cadastrar_escola'))

@app.route('/visitas/escolas', methods=['GET', 'POST'])
def cadastrar_visita_escola():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verificar se existem escolas cadastradas
    escolas_existentes = Escola.query.all()
    if not escolas_existentes:
        flash('É necessário cadastrar pelo menos uma escola antes de agendar visitas.', 'warning')
        return redirect(url_for('cadastrar_escola'))
    
    form = VisitaEscolaForm()
    form.id_escola.choices = [(e.id_escola, f"{e.nome} - {e.cidade}/{e.estado}") for e in escolas_existentes]
    
    # Buscar colaboradores com categoria "Promotor de Vendas"
    promotores = db.session.query(Colaborador).join(ColaboradorCategoria).join(CategoriaColaborador).filter(
        CategoriaColaborador.nome.ilike('%promotor de vendas%')
    ).order_by(Colaborador.nome).all()
    
    form.id_promotor.choices = [(p.id_colaborador, p.nome) for p in promotores]
    
    # Definir promotor padrão como o usuário logado se ele for promotor
    if request.method == 'GET':
        usuario = Usuario.query.get(session['user_id'])
        if usuario and usuario.colaborador:
            # Verificar se o usuário logado é promotor de vendas
            is_promotor = any(cat.nome.lower() == 'promotor de vendas' for cat in usuario.colaborador.categorias)
            if is_promotor:
                form.id_promotor.data = usuario.colaborador.id_colaborador
    
    if form.validate_on_submit():
        nova_visita = VisitaEscola(
            id_escola=form.id_escola.data,
            data_visita=form.data_visita.data,
            id_promotor=form.id_promotor.data,
            observacoes_visita=form.observacoes_visita.data,
            status_visita=form.status_visita.data
        )
        db.session.add(nova_visita)
        db.session.commit()
        
        # Registrar log da operação
        escola = Escola.query.get(form.id_escola.data)
        registrar_log('Cadastrar Visita', f'Visita agendada para escola "{escola.nome}" - {nova_visita.data_visita.strftime("%d/%m/%Y %H:%M")}')
        
        flash('Visita agendada com sucesso!', 'success')
        return redirect(url_for('cadastrar_visita_escola'))
    
    # Listar visitas ordenadas por data
    visitas = db.session.query(VisitaEscola).join(Escola).join(Colaborador).order_by(VisitaEscola.data_visita.desc()).all()
    return render_template('visitas_escola.html', form=form, visitas=visitas)

@app.route('/visitas/escolas/editar/<int:id>', methods=['GET', 'POST'])
def editar_visita_escola(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    visita = VisitaEscola.query.get_or_404(id)
    
    # Verificar se o usuário pode editar esta visita (apenas o próprio promotor ou admin)
    usuario = Usuario.query.get(session['user_id'])
    if not usuario or not usuario.colaborador:
        flash('Erro ao carregar informações do usuário.', 'danger')
        return redirect(url_for('login'))
    
    if session.get('categoria') != 'administrativo' and visita.id_promotor != usuario.colaborador.id_colaborador:
        flash('Você não tem permissão para editar esta visita.', 'danger')
        return redirect(url_for('cadastrar_visita_escola'))
    
    escolas_existentes = Escola.query.all()
    form = VisitaEscolaForm(obj=visita)
    form.id_escola.choices = [(e.id_escola, f"{e.nome} - {e.cidade}/{e.estado}") for e in escolas_existentes]
    
    # Buscar colaboradores com categoria "Promotor de Vendas"
    promotores = db.session.query(Colaborador).join(ColaboradorCategoria).join(CategoriaColaborador).filter(
        CategoriaColaborador.nome.ilike('%promotor de vendas%')
    ).order_by(Colaborador.nome).all()
    
    form.id_promotor.choices = [(p.id_colaborador, p.nome) for p in promotores]
    
    if form.validate_on_submit():
        visita.id_escola = form.id_escola.data
        visita.id_promotor = form.id_promotor.data
        visita.data_visita = form.data_visita.data
        visita.observacoes_visita = form.observacoes_visita.data
        visita.status_visita = form.status_visita.data
        
        db.session.commit()
        
        # Registrar log da operação
        escola = Escola.query.get(visita.id_escola)
        registrar_log('Editar Visita', f'Visita editada para escola "{escola.nome}" - {visita.data_visita.strftime("%d/%m/%Y %H:%M")}')
        
        flash('Visita atualizada com sucesso!', 'success')
        return redirect(url_for('cadastrar_visita_escola'))
    
    visitas = db.session.query(VisitaEscola).join(Escola).join(Colaborador).order_by(VisitaEscola.data_visita.desc()).all()
    return render_template('visitas_escola.html', form=form, visitas=visitas, visita_editando=visita)

@app.route('/visitas/escolas/excluir/<int:id>')
def excluir_visita_escola(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    visita = VisitaEscola.query.get_or_404(id)
    
    # Verificar se o usuário pode excluir esta visita (apenas o próprio promotor ou admin)
    usuario = Usuario.query.get(session['user_id'])
    if not usuario or not usuario.colaborador:
        flash('Erro ao carregar informações do usuário.', 'danger')
        return redirect(url_for('login'))
    
    if session.get('categoria') != 'administrativo' and visita.id_promotor != usuario.colaborador.id_colaborador:
        flash('Você não tem permissão para excluir esta visita.', 'danger')
        return redirect(url_for('cadastrar_visita_escola'))
    
    # Salvar dados para o log antes de excluir
    escola = Escola.query.get(visita.id_escola)
    data_visita = visita.data_visita.strftime("%d/%m/%Y %H:%M")
    
    db.session.delete(visita)
    db.session.commit()
    
    # Registrar log da operação
    registrar_log('Excluir Visita', f'Visita excluída para escola "{escola.nome}" - {data_visita}')
    
    flash('Visita excluída com sucesso!', 'success')
    return redirect(url_for('cadastrar_visita_escola'))

@app.route('/escolas/<int:id>/historico')
def historico_visitas_escola(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    escola = Escola.query.get_or_404(id)
    visitas = VisitaEscola.query.filter_by(id_escola=id).join(Colaborador).order_by(VisitaEscola.data_visita.desc()).all()
    
    return render_template('historico_visitas_escola.html', escola=escola, visitas=visitas)

@app.route('/marketing/dashboard')
def marketing_dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verificar permissão (apenas administradores)
    if not is_admin_user():
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Estatísticas gerais
    total_escolas = Escola.query.count()
    total_visitas = VisitaEscola.query.count()
    
    # Escolas com visitas pendentes de envio de material
    escolas_pendentes_email = db.session.query(Escola).join(VisitaEscola).filter(
        and_(
            Escola.email.isnot(None),
            Escola.email != '',
            VisitaEscola.email_enviado == False
        )
    ).distinct().count()
    
    escolas_pendentes_whatsapp = db.session.query(Escola).join(VisitaEscola).filter(
        and_(
            Escola.whatsapp.isnot(None),
            Escola.whatsapp != '',
            VisitaEscola.whatsapp_enviado == False
        )
    ).distinct().count()
    
    # Escolas já contactadas
    escolas_email_enviado = db.session.query(Escola).join(VisitaEscola).filter(
        VisitaEscola.email_enviado == True
    ).distinct().count()
    
    escolas_whatsapp_enviado = db.session.query(Escola).join(VisitaEscola).filter(
        VisitaEscola.whatsapp_enviado == True
    ).distinct().count()
    
    # Visitas por status
    visitas_agendadas = VisitaEscola.query.filter_by(status_visita='agendada').count()
    visitas_realizadas = VisitaEscola.query.filter_by(status_visita='realizada').count()
    visitas_canceladas = VisitaEscola.query.filter_by(status_visita='cancelada').count()
    
    # Últimas visitas realizadas
    ultimas_visitas = db.session.query(VisitaEscola).join(Escola).join(Colaborador).order_by(VisitaEscola.data_visita.desc()).limit(5).all()
    
    # Escolas mais visitadas
    escolas_mais_visitadas = db.session.query(
        Escola.nome,
        Escola.cidade,
        Escola.estado,
        func.count(VisitaEscola.id_visita).label('total_visitas')
    ).join(VisitaEscola).group_by(Escola.id_escola).order_by(func.count(VisitaEscola.id_visita).desc()).limit(5).all()
    
    return render_template('marketing_dashboard.html',
                         total_escolas=total_escolas,
                         total_visitas=total_visitas,
                         escolas_pendentes_email=escolas_pendentes_email,
                         escolas_pendentes_whatsapp=escolas_pendentes_whatsapp,
                         escolas_email_enviado=escolas_email_enviado,
                         escolas_whatsapp_enviado=escolas_whatsapp_enviado,
                         visitas_agendadas=visitas_agendadas,
                         visitas_realizadas=visitas_realizadas,
                         visitas_canceladas=visitas_canceladas,
                         ultimas_visitas=ultimas_visitas,
                         escolas_mais_visitadas=escolas_mais_visitadas)

def enviar_email_marketing(escola, colaborador):
    """
    Função para enviar email de marketing para uma escola
    """
    try:
        # Verificar se as configurações de email estão definidas
        if not app.config.get('MAIL_USERNAME') or not app.config.get('MAIL_PASSWORD'):
            print("❌ Configurações de email não definidas")
            return False
        
        # Criar mensagem de email
        assunto = f"Circo Stankowich - Proposta de Apresentação para {escola.nome}"
        
        # Template do email em HTML
        corpo_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background-color: #e74c3c; color: white; padding: 20px; text-align: center; }}
                .content {{ padding: 20px; background-color: #f9f9f9; }}
                .footer {{ padding: 20px; text-align: center; color: #666; }}
                .btn {{ background-color: #e74c3c; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🎪 Circo Stankowich</h1>
                    <p>Alegria e Diversão para Toda a Família!</p>
                </div>
                
                <div class="content">
                    <h2>Olá, {escola.nome_contato or 'Responsável'}!</h2>
                    
                    <p>Esperamos que esteja tudo bem! Somos do <strong>Circo Stankowich</strong> e gostaríamos de apresentar uma proposta especial para levar alegria e diversão diretamente para a <strong>{escola.nome}</strong>.</p>
                    
                    <h3>🎭 O que oferecemos:</h3>
                    <ul>
                        <li>Espetáculos circenses profissionais</li>
                        <li>Apresentações educativas e divertidas</li>
                        <li>Experiência inesquecível para alunos e familiares</li>
                        <li>Preços especiais para instituições de ensino</li>
                    </ul>
                    
                    <h3>📍 Informações da Escola:</h3>
                    <p><strong>Escola:</strong> {escola.nome}<br>
                    <strong>Cidade:</strong> {escola.cidade}/{escola.estado}</p>
                    
                    <p>Gostaríamos muito de agendar uma conversa para apresentar nossa proposta detalhada e discutir as melhores opções para vocês.</p>
                    
                    <p>Ficamos no aguardo do seu contato!</p>
                </div>
                
                <div class="footer">
                    <p><strong>Atenciosamente,</strong><br>
                    {colaborador.nome}<br>
                    Circo Stankowich</p>
                    
                    <p>📧 contato@socratesonline.com<br>
                    📱 {colaborador.telefone or 'Telefone disponível via email'}</p>
                    
                    <p style="font-size: 12px; color: #999;">
                        Este email foi enviado automaticamente pelo sistema de marketing do Circo Stankowich.
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Versão em texto simples
        corpo_texto = f"""
        Circo Stankowich - Proposta de Apresentação
        
        Olá, {escola.nome_contato or 'Responsável'}!
        
        Somos do Circo Stankowich e gostaríamos de apresentar uma proposta especial para levar alegria e diversão diretamente para a {escola.nome}.
        
        O que oferecemos:
        - Espetáculos circenses profissionais
        - Apresentações educativas e divertidas
        - Experiência inesquecível para alunos e familiares
        - Preços especiais para instituições de ensino
        
        Escola: {escola.nome}
        Cidade: {escola.cidade}/{escola.estado}
        
        Gostaríamos muito de agendar uma conversa para apresentar nossa proposta detalhada.
        
        Atenciosamente,
        {colaborador.nome}
        Circo Stankowich
        
        contato@socratesonline.com
        {colaborador.telefone or 'Telefone disponível via email'}
        """
        
        # Criar e enviar mensagem
        msg = Message(
            subject=assunto,
            recipients=[escola.email],
            html=corpo_html,
            body=corpo_texto
        )
        
        mail.send(msg)
        
        # Registrar log do sistema
        log = LogSistema(
            usuario_id=session.get('user_id'),
            acao='ENVIO_EMAIL_MARKETING',
            detalhes=f'Email enviado para escola {escola.nome} ({escola.email})',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        
        return True
        
    except Exception as e:
        print(f"❌ Erro ao enviar email: {str(e)}")
        
        # Registrar log de erro
        try:
            log = LogSistema(
                usuario_id=session.get('user_id'),
                acao='ERRO_ENVIO_EMAIL_MARKETING',
                detalhes=f'Erro ao enviar email para escola {escola.nome}: {str(e)}',
                ip_address=request.remote_addr
            )
            db.session.add(log)
        except:
            pass
        
        return False

@app.route('/marketing/enviar-material', methods=['POST'])
def enviar_material_marketing():
    if 'user_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    # Verificar permissão
    if session.get('categoria', '').lower() not in ['administrativo', 'promotor de vendas']:
        return jsonify({'error': 'Acesso restrito a administradores e promotores de vendas'}), 403
    
    try:
        tipo_envio = request.json.get('tipo')  # 'email' ou 'whatsapp'
        
        if tipo_envio not in ['email', 'whatsapp']:
            return jsonify({'error': 'Tipo de envio inválido'}), 400
        
        # Obter usuário atual
        usuario = Usuario.query.get(session['user_id'])
        if not usuario or not usuario.colaborador:
            return jsonify({'error': 'Usuário inválido'}), 400
        
        # Buscar visitas pendentes de envio
        if tipo_envio == 'email':
            visitas_pendentes = db.session.query(VisitaEscola).join(Escola).filter(
                and_(
                    Escola.email.isnot(None),
                    Escola.email != '',
                    VisitaEscola.email_enviado == False
                )
            ).all()
        else:  # whatsapp
            visitas_pendentes = db.session.query(VisitaEscola).join(Escola).filter(
                and_(
                    Escola.whatsapp.isnot(None),
                    Escola.whatsapp != '',
                    VisitaEscola.whatsapp_enviado == False
                )
            ).all()
        
        if not visitas_pendentes:
            return jsonify({'message': f'Nenhuma escola pendente de envio por {tipo_envio}', 'count': 0})
        
        # Simular envio e marcar como enviado
        contador_enviados = 0
        data_envio = datetime.now()
        
        for visita in visitas_pendentes:
            try:
                # Aqui seria implementada a lógica real de envio
                # Por enquanto, apenas simulamos o envio
                
                if tipo_envio == 'email':
                    # Envio real de email
                    sucesso_email = enviar_email_marketing(visita.escola, usuario.colaborador)
                    if sucesso_email:
                        print(f"📧 Email enviado com sucesso para: {visita.escola.email} - Escola: {visita.escola.nome}")
                        visita.email_enviado = True
                        visita.data_email_enviado = data_envio
                    else:
                        print(f"❌ Falha ao enviar email para: {visita.escola.email} - Escola: {visita.escola.nome}")
                        continue
                else:
                    # Envio real de WhatsApp
                    promotor = Colaborador.query.get(visita.id_promotor)
                    mensagem = criar_mensagem_visita_escola(visita.escola, visita, promotor)
                    
                    resultado = enviar_whatsapp(visita.escola.whatsapp, mensagem)
                    
                    if resultado['success']:
                        print(f"📱 WhatsApp enviado com sucesso para: {visita.escola.whatsapp} - Escola: {visita.escola.nome}")
                        print(f"   Message ID: {resultado['message_id']}")
                        visita.whatsapp_enviado = True
                        visita.data_whatsapp_enviado = data_envio
                        
                        # Registrar log do envio
                        registrar_log('Envio WhatsApp', 
                                    f'WhatsApp enviado para escola "{visita.escola.nome}" - {resultado["numero_enviado"]}')
                    else:
                        print(f"❌ Erro ao enviar WhatsApp para {visita.escola.nome}: {resultado['error']}")
                        continue
                
                contador_enviados += 1
                
            except Exception as e:
                print(f"❌ Erro ao enviar para escola {visita.escola.nome}: {e}")
                continue
        
        # Salvar alterações
        db.session.commit()
        
        return jsonify({
            'message': f'Material enviado com sucesso por {tipo_envio}!',
            'count': contador_enviados,
            'tipo': tipo_envio
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@app.route('/eventos')
def listar_eventos():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    try:
        # Obter informações do usuário logado
        usuario = Usuario.query.get(session['user_id'])
        if not usuario or not usuario.colaborador:
            flash('Erro ao carregar informações do usuário.', 'danger')
            return redirect(url_for('login'))
        
        # Verificar se é administrador
        is_admin = is_admin_user()

        # Processar filtro de período
        period = request.args.get('period', '90dias')  # Default para 90 dias
        data_inicio = None
        data_fim = None
        
        if period == 'hoje':
            data_inicio = data_fim = date.today()
        elif period == 'ontem':
            data_inicio = data_fim = date.today() - timedelta(days=1)
        elif period == '7dias':
            data_fim = date.today()
            data_inicio = data_fim - timedelta(days=6)
        elif period == 'mes':
            data_fim = date.today()
            data_inicio = data_fim.replace(day=1)
        elif period == 'custom':
            data_inicio_str = request.args.get('data_inicio')
            data_fim_str = request.args.get('data_fim')
            if data_inicio_str and data_fim_str:
                try:
                    data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
                    data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
                except ValueError:
                    flash('Datas inválidas fornecidas.', 'warning')
                    # Fallback para 90 dias
                    data_fim = date.today()
                    data_inicio = data_fim - timedelta(days=90)
            else:
                # Se custom mas sem datas, usar 90 dias
                data_fim = date.today()
                data_inicio = data_fim - timedelta(days=90)
        else:
            # Default: últimos 90 dias (incluindo eventos futuros)
            data_fim = date.today() + timedelta(days=365)  # Incluir eventos futuros
            data_inicio = date.today() - timedelta(days=90)
        

        
        # Filtrar eventos baseado no tipo de usuário
        eventos_query = Evento.query.filter(
            Evento.status.in_(['planejamento', 'a realizar', 'em andamento', 'realizado'])
        )
        
        # Aplicar filtro de data se definido
        if data_inicio:
            eventos_query = eventos_query.filter(Evento.data_inicio >= data_inicio)
        if data_fim:
            eventos_query = eventos_query.filter(Evento.data_inicio <= data_fim)
        
        if not is_admin:
            # Produtores veem apenas seus eventos
            eventos_query = eventos_query.filter_by(id_produtor=usuario.colaborador.id_colaborador)

        
        eventos = eventos_query.order_by(Evento.data_inicio.desc()).all()
        

        
        # Buscar dados para os modais de adição rápida
        categorias_receita = CategoriaReceita.query.all()
        categorias_despesa = CategoriaDespesa.query.filter(
            CategoriaDespesa.id_categoria_despesa.in_(
                db.session.query(Despesa.id_categoria_despesa).filter(
                    Despesa.id_tipo_despesa.in_([1, 2])
                ).distinct()
            )
        ).all()
        fornecedores = Fornecedor.query.order_by(Fornecedor.nome).all()
        current_date = date.today().strftime('%Y-%m-%d')
        
        # Gravar log da ação
        try:
            log_entry = LogSistema(
                usuario_id=str(session['user_id']),
                usuario_nome=usuario.nome,
                usuario_email=usuario.email,
                acao='Visualizar Eventos',
                descricao=f'Usuário acessou a listagem de eventos com filtro: {period}'
            )
            db.session.add(log_entry)
            db.session.commit()
        except Exception as log_error:
            print(f"Erro ao gravar log: {log_error}")
        
        # Converter objetos para dicionários simples para evitar problemas de serialização
        categorias_receita_dict = [{'id': cat.id_categoria_receita, 'nome': cat.nome} for cat in categorias_receita]
        categorias_despesa_dict = [{'id': cat.id_categoria_despesa, 'nome': cat.nome} for cat in categorias_despesa]
        fornecedores_dict = [{'id': forn.id_fornecedor, 'nome': forn.nome} for forn in fornecedores]
        
        return render_template('eventos.html', 
                             eventos=eventos, 
                             is_admin=is_admin, 
                             usuario=usuario,
                             categorias_receita=categorias_receita_dict,
                             categorias_despesa=categorias_despesa_dict,
                             fornecedores=fornecedores_dict,
                             current_date=current_date,
                             period=period,
                             data_inicio=data_inicio.strftime('%Y-%m-%d') if data_inicio else '',
                             data_fim=data_fim.strftime('%Y-%m-%d') if data_fim else '')
    
    except Exception as e:
        print(f"Erro na rota de eventos: {str(e)}")
        flash('Erro interno do servidor. Tente novamente.', 'danger')
        return redirect(url_for('dashboard'))

# Endpoints de API para carregamento dinÃ¢mico nos modais
@app.route('/api/despesas-por-categoria/<int:categoria_id>')
def api_despesas_por_categoria(categoria_id):
    try:
        # Filtrar apenas despesas de evento (tipos 1 e 2)
        despesas = Despesa.query.filter_by(id_categoria_despesa=categoria_id).filter(
            Despesa.id_tipo_despesa.in_([1, 2])
        ).all()
        despesas_data = []
        
        for despesa in despesas:
            # Usar o valor médio já cadastrado na despesa ou calcular dinamicamente
            valor_medio = despesa.valor_medio_despesa
            
            # Se não tem valor médio cadastrado, calcular baseado nos eventos
            if not valor_medio:
                despesas_evento = DespesaEvento.query.filter_by(id_despesa=despesa.id_despesa).all()
                if despesas_evento:
                    valores = [de.valor for de in despesas_evento if de.valor]
                    if valores:
                        valor_medio = sum(valores) / len(valores)
            
            despesas_data.append({
                'id_despesa': despesa.id_despesa,
                'nome': despesa.nome,
                'valor_medio': float(valor_medio) if valor_medio else None,
                'id_tipo_despesa': despesa.id_tipo_despesa
            })
        
        return jsonify(despesas_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/despesas-empresa-por-categoria/<int:categoria_id>')
def api_despesas_empresa_por_categoria(categoria_id):
    try:
        # Filtrar apenas despesas da empresa (tipos 3 e 4)
        despesas = Despesa.query.filter_by(id_categoria_despesa=categoria_id).filter(
            Despesa.id_tipo_despesa.in_([3, 4])
        ).all()
        despesas_data = []
        
        for despesa in despesas:
            # Usar o valor médio já cadastrado na despesa ou calcular dinamicamente
            valor_medio = despesa.valor_medio_despesa
            
            # Se não tem valor médio cadastrado, calcular baseado nas despesas da empresa
            if not valor_medio:
                despesas_empresa = DespesaEmpresa.query.filter_by(id_despesa=despesa.id_despesa).all()
                if despesas_empresa:
                    valores = [de.valor for de in despesas_empresa if de.valor]
                    if valores:
                        valor_medio = sum(valores) / len(valores)
            
            despesas_data.append({
                'id_despesa': despesa.id_despesa,
                'nome': despesa.nome,
                'valor_medio': float(valor_medio) if valor_medio else None,
                'id_tipo_despesa': despesa.id_tipo_despesa
            })
        
        return jsonify(despesas_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/receitas-por-categoria/<int:categoria_id>')
def api_receitas_por_categoria(categoria_id):
    try:
        receitas = Receita.query.filter_by(id_categoria_receita=categoria_id).all()
        receitas_data = []
        
        for receita in receitas:
            receitas_data.append({
                'id_receita': receita.id_receita,
                'nome': receita.nome
            })
        
        return jsonify(receitas_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/despesa-detalhes/<int:despesa_id>')
def api_despesa_detalhes(despesa_id):
    """API para buscar detalhes de uma despesa específica"""
    try:
        despesa = Despesa.query.get(despesa_id)
        if not despesa:
            return jsonify({'error': 'Despesa não encontrada'}), 404
        
        return jsonify({
            'id_despesa': despesa.id_despesa,
            'nome': despesa.nome,
            'flag_alimentacao': despesa.flag_alimentacao,
            'flag_combustivel': despesa.flag_combustivel,
            'categoria_nome': despesa.categoria.nome if despesa.categoria else None,
            'tipo_despesa': despesa.id_tipo_despesa,
            'valor_medio': float(despesa.valor_medio_despesa) if despesa.valor_medio_despesa else None
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/despesa-valor-medio/<int:despesa_id>')
def api_despesa_valor_medio(despesa_id):
    """API para buscar valor médio de uma despesa e validar percentuais"""
    try:
        despesa = Despesa.query.get(despesa_id)
        if not despesa:
            return jsonify({'error': 'Despesa não encontrada'}), 404
        
        valor_medio = float(despesa.valor_medio_despesa) if despesa.valor_medio_despesa else None
        
        return jsonify({
            'success': True,
            'valor_medio': valor_medio,
            'valor_medio_formatado': f"{valor_medio:.2f}".replace('.', ',') if valor_medio else None,
            'nome_despesa': despesa.nome,
            'has_valor_medio': valor_medio is not None and valor_medio > 0
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/eventos/<int:id_evento>/salvar-receita', methods=['POST'])
def salvar_receita_individual(id_evento):
    try:
        data = request.get_json()
        
        # Validar dados obrigatórios
        if not data.get('receita_id') or not data.get('valor'):
            return jsonify({'success': False, 'message': 'Receita e valor são obrigatórios'})
        
        # Verificar se o evento existe
        evento = Evento.query.get(id_evento)
        if not evento:
            return jsonify({'success': False, 'message': 'Evento não encontrado'})
        
        # Converter valor - tratar tanto formato brasileiro quanto americano
        try:
            valor_str = str(data['valor']).strip()
            
            # Se contém ponto e vírgula, é formato brasileiro (ex: 1.000,50)
            if '.' in valor_str and ',' in valor_str:
                # Remover pontos de milhares e trocar vírgula por ponto
                valor_str = valor_str.replace('.', '').replace(',', '.')
            # Se contém apenas vírgula, trocar por ponto
            elif ',' in valor_str and '.' not in valor_str:
                valor_str = valor_str.replace(',', '.')
            
            valor = float(valor_str)
            
            if valor <= 0:
                return jsonify({'success': False, 'message': 'Valor deve ser maior que zero'})
                
        except (ValueError, TypeError) as e:

            return jsonify({'success': False, 'message': 'Valor inválido'})
        
        # Converter data
        try:
            data_receita = datetime.strptime(data['data'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': 'Data inválida'})
        
        # Criar receita do evento
        receita_evento = ReceitaEvento(
            id_evento=id_evento,
            id_receita=int(data['receita_id']),
            valor=valor,
            data=data_receita,
            observacoes=data.get('observacoes', '')
        )
        
        db.session.add(receita_evento)
        db.session.commit()
        
        # Buscar nome da receita para retorno
        receita = Receita.query.get(data['receita_id'])
        receita_nome = receita.nome if receita else 'Receita'
        
        print(f"âœ… Receita salva: {receita_nome} - R$ {valor:,.2f}")
        
        return jsonify({
            'success': True, 
            'message': 'Receita adicionada com sucesso',
            'receita_evento_id': receita_evento.id_receita_evento,
            'receita_nome': receita_nome,
            'valor_salvo': valor
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"âŒ Erro ao salvar receita: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'})

@app.route('/eventos/novo', methods=['GET', 'POST'])
def novo_evento():
    form = EventoForm()
    form.id_circo.choices = [(c.id_circo, c.nome) for c in Circo.query.all()]
    
    # Filtrar apenas colaboradores que são produtores
    produtores = [p for p in Colaborador.query.all() if p.tem_categoria_produtor]
    if not produtores:
        flash('É necessário cadastrar pelo menos um colaborador como produtor antes de criar eventos.', 'warning')
        return redirect(url_for('cadastrar_colaborador'))
    
    form.id_produtor.choices = [(p.id_colaborador, p.nome) for p in produtores]
    
    categorias_receita = CategoriaReceita.query.all()
    
    # Filtrar apenas categorias que tÃªm despesas de evento (tipos 1 e 2)
    categorias_despesa = CategoriaDespesa.query.filter(
        CategoriaDespesa.id_categoria_despesa.in_(
            db.session.query(Despesa.id_categoria_despesa).filter(
                Despesa.id_tipo_despesa.in_([1, 2])
            ).distinct()
        )
    ).all()

    if form.validate_on_submit():
        # Debug: Verificar o tipo dos dados recebidos
        print(f"Tipo data_inicio: {type(form.data_inicio.data)}, valor: {form.data_inicio.data}")
        print(f"Tipo data_fim: {type(form.data_fim.data)}, valor: {form.data_fim.data}")
        
        # Processamento manual das datas se necessário
        data_inicio = form.data_inicio.data
        data_fim = form.data_fim.data
        
        # Converter strings para objetos date se necessário
        if isinstance(data_inicio, str):
            if '/' in data_inicio:
                # Formato brasileiro DD/MM/YYYY
                data_inicio = datetime.strptime(data_inicio, '%d/%m/%Y').date()
            else:
                # Formato ISO YYYY-MM-DD
                data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        
        if isinstance(data_fim, str):
            if '/' in data_fim:
                # Formato brasileiro DD/MM/YYYY
                data_fim = datetime.strptime(data_fim, '%d/%m/%Y').date()
            else:
                # Formato ISO YYYY-MM-DD
                data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        
        novo = Evento(
            nome=form.nome.data,
            data_inicio=data_inicio,
            data_fim=data_fim,
            cidade=form.cidade.data,
            estado=form.estado.data,
            endereco=form.endereco.data,
            id_circo=form.id_circo.data,
            id_produtor=form.id_produtor.data,
            status=form.status.data,
            observacoes=form.observacoes.data
        )
        
        print(f"Evento criado: {novo.nome}, ID antes do add: {novo.id_evento}")
        db.session.add(novo)
        print(f"Evento adicionado Ã  sessão")
        
        try:
            db.session.flush()
            print(f"Flush executado, ID do evento: {novo.id_evento}")
        except Exception as e:
            print(f"Erro no flush: {e}")
            raise

        # Adicionar automaticamente o produtor na equipe do evento
        if form.id_produtor.data:
            equipe_produtor = EquipeEvento(
                id_evento=novo.id_evento,
                id_colaborador=form.id_produtor.data,
                funcao='Produtor',
                observacoes='Adicionado automaticamente como produtor do evento'
            )
            db.session.add(equipe_produtor)
            print(f"Produtor {form.id_produtor.data} adicionado automaticamente à equipe do evento")

        # 1. Processar receitas
        receitas_ids = request.form.getlist('receita_id[]')
        receita_datas = request.form.getlist('receita_data[]')
        valores = request.form.getlist('valor[]')
        observacoes = request.form.getlist('obs[]')

        for i in range(len(receitas_ids)):
            try:
                valor = float(valores[i].replace(',', '.'))
                data_receita = receita_datas[i] if i < len(receita_datas) else None
                data = datetime.strptime(data_receita, '%Y-%m-%d').date() if data_receita else None
            except:
                continue
            if receitas_ids[i] and valor:
                db.session.add(ReceitaEvento(
                    id_evento=novo.id_evento,
                    id_receita=int(receitas_ids[i]),
                    valor=valor,
                    data=data,
                    observacoes=observacoes[i]
                ))

        # 2. Processar despesas variáveis do formulário
        despesa_ids = request.form.getlist('despesa_id[]')
        datas_desp = request.form.getlist('despesa_data[]')
        valores_desp = request.form.getlist('despesa_valor[]')
        status_pag = request.form.getlist('despesa_status_pagamento[]')
        forma_pag = request.form.getlist('despesa_forma_pagamento[]')
        pago_por = request.form.getlist('despesa_pago_por[]')
        obs_desp = request.form.getlist('despesa_obs[]')

        for i in range(len(despesa_ids)):
            try:
                # Verificar se tem ID da despesa (obrigatório)
                if not despesa_ids[i]:
                    continue
                    
                # Verificar se tem valor preenchido
                if not valores_desp[i] or valores_desp[i].strip() == '':
                    continue
                    
                valor = float(valores_desp[i].replace(',', '.'))
                data = datetime.strptime(datas_desp[i], '%Y-%m-%d').date() if datas_desp[i] else date.today()
            except (ValueError, IndexError):
                continue
                
            # Adicionar despesa ao evento
            db.session.add(DespesaEvento(
                id_evento=novo.id_evento,
                id_despesa=int(despesa_ids[i]),
                data_vencimento=data,
                valor=valor,
                status_pagamento=status_pag[i] if i < len(status_pag) else 'pendente',
                forma_pagamento=forma_pag[i] if i < len(forma_pag) else 'débito',
                pago_por=pago_por[i] if i < len(pago_por) else '',
                observacoes=obs_desp[i] if i < len(obs_desp) else ''
            ))

        # 3. Cadastrar TODAS as despesas fixas automaticamente (UMA VEZ)
        despesas_fixas = Despesa.query.filter_by(id_tipo_despesa=1).all()
        
        for despesa_fixa in despesas_fixas:
            valor_automatico = despesa_fixa.valor_medio_despesa or 0
            db.session.add(DespesaEvento(
                id_evento=novo.id_evento,
                id_despesa=despesa_fixa.id_despesa,
                data_vencimento=novo.data_inicio or date.today(),
                valor=valor_automatico,
                status_pagamento='pendente',
                forma_pagamento='débito',
                pago_por='',
                observacoes='Despesa fixa adicionada automaticamente'
            ))

        try:
            db.session.commit()
            print(f"âœ… COMMIT executado com sucesso! Evento ID: {novo.id_evento}")
            
            # Verificar se o evento foi realmente salvo
            evento_verificacao = Evento.query.get(novo.id_evento)
            if evento_verificacao:
                print(f"âœ… Evento verificado no banco: {evento_verificacao.nome}")
            else:
                print("âŒ ERRO: Evento não encontrado no banco após commit!")
            
            # Verificar as despesas APÃ"S o commit
            despesas_depois = DespesaEvento.query.filter_by(id_evento=novo.id_evento).all()
            print(f"\n=== DESPESAS NO EVENTO APÓS O COMMIT ===")
            for desp in despesas_depois:
                print(f"ID Despesa: {desp.id_despesa}, Valor SALVO: {desp.valor}, Tipo: {desp.despesa.id_tipo_despesa}")
            
            # Registrar log da operação
            registrar_log('Criar Evento', f'Evento "{novo.nome}" criado - {novo.cidade}/{novo.estado} ({novo.data_inicio.strftime("%d/%m/%Y") if novo.data_inicio else "sem data"})')
            
            flash('Evento e despesas/receitas cadastrados com sucesso!', 'success')
            return redirect(url_for('editar_evento', id=novo.id_evento))
            
        except Exception as e:
            print(f"âŒ ERRO no commit: {e}")
            db.session.rollback()
            flash(f'Erro ao salvar evento: {str(e)}', 'danger')
            # Não fazer redirect em caso de erro para debugar

    # Para a exibiÃ§ão inicial (antes de salvar), mostrar todas as despesas disponíveis
    categorias_receita_dict = {
        c.id_categoria_receita: [{'id_receita': r.id_receita, 'nome': r.nome}
                                 for r in Receita.query.filter_by(id_categoria_receita=c.id_categoria_receita)]
        for c in categorias_receita
    }

    # Criar estrutura de despesas organizada por tipo para exibiÃ§ão
    categorias_despesa_dict = {}
    for categoria in categorias_despesa:
        despesas_categoria = Despesa.query.filter_by(
            id_categoria_despesa=categoria.id_categoria_despesa
        ).filter(Despesa.id_tipo_despesa.in_([1, 2])).all()
        
        despesas_fixas = [d for d in despesas_categoria if d.id_tipo_despesa == 1]
        despesas_variaveis = [d for d in despesas_categoria if d.id_tipo_despesa == 2]
        
        categorias_despesa_dict[categoria.id_categoria_despesa] = {
            'fixas': [{
                'id_despesa': d.id_despesa, 
                'nome': d.nome,
                'valor_medio': float(d.valor_medio_despesa) if d.valor_medio_despesa else None,
                'tipo': 1,
                'ja_cadastrada': False  # No cadastro, nenhuma está cadastrada ainda
            } for d in despesas_fixas],
            'variaveis': [{
                'id_despesa': d.id_despesa, 
                'nome': d.nome,
                'valor_medio': float(d.valor_medio_despesa) if d.valor_medio_despesa else None,
                'tipo': 2,
                'ja_cadastrada': False  # No cadastro, nenhuma está cadastrada ainda
            } for d in despesas_variaveis]
        }

    # Criar estrutura vazia para despesas_evento_por_categoria (novo evento não tem despesas ainda)
    despesas_evento_por_categoria = {}
    for categoria in categorias_despesa:
        despesas_evento_por_categoria[categoria.id_categoria_despesa] = {
            'fixas': [],
            'variaveis': []
        }

    return render_template(
        'novo_evento.html',
        form=form,
        categorias_receita=categorias_receita,
        categorias_receita_dict=categorias_receita_dict,
        categorias_despesa=categorias_despesa,
        categorias_despesa_dict=categorias_despesa_dict,
        despesas_evento_por_categoria=despesas_evento_por_categoria,  # Adicionando a variável faltante
        fornecedores=[{'id_fornecedor': f.id_fornecedor, 'nome': f.nome} for f in Fornecedor.query.all()],
        receitas_salvas=[],
        despesas_salvas=[],
        current_date=date.today().isoformat(),
        is_editing=False
    )

@app.route('/eventos/editar/<int:id>', methods=['GET', 'POST'])
def editar_evento(id):
    evento = Evento.query.get_or_404(id)
    form = EventoForm(obj=evento)
    form.id_circo.choices = [(c.id_circo, c.nome) for c in Circo.query.all()]
    
    # Filtrar apenas colaboradores que são produtores
    produtores = [p for p in Colaborador.query.all() if p.tem_categoria_produtor]
    if not produtores:
        flash('É necessário cadastrar pelo menos um colaborador como produtor antes de editar eventos.', 'warning')
        return redirect(url_for('cadastrar_colaborador'))
    
    form.id_produtor.choices = [(p.id_colaborador, p.nome) for p in produtores]
    
    categorias_receita = CategoriaReceita.query.all()
    
    # Filtrar apenas categorias que tÃªm despesas de evento (tipos 1 e 2)
    categorias_despesa = CategoriaDespesa.query.filter(
        CategoriaDespesa.id_categoria_despesa.in_(
            db.session.query(Despesa.id_categoria_despesa).filter(
                Despesa.id_tipo_despesa.in_([1, 2])
            ).distinct()
        )
    ).all()

    # Obter receitas e despesas já cadastradas
    receitas_salvas = ReceitaEvento.query.filter_by(id_evento=id).all()
    despesas_salvas = DespesaEvento.query.filter_by(id_evento=id).all()

    if form.validate_on_submit():
        # Verificar se o produtor mudou
        produtor_anterior = evento.id_produtor
        novo_produtor = form.id_produtor.data
        
        # Processamento manual das datas se necessário
        data_inicio = form.data_inicio.data
        data_fim = form.data_fim.data
        
        # Converter strings para objetos date se necessário
        if isinstance(data_inicio, str):
            if '/' in data_inicio:
                # Formato brasileiro DD/MM/YYYY
                data_inicio = datetime.strptime(data_inicio, '%d/%m/%Y').date()
            else:
                # Formato ISO YYYY-MM-DD
                data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        
        if isinstance(data_fim, str):
            if '/' in data_fim:
                # Formato brasileiro DD/MM/YYYY
                data_fim = datetime.strptime(data_fim, '%d/%m/%Y').date()
            else:
                # Formato ISO YYYY-MM-DD
                data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        
        evento.nome = form.nome.data
        evento.data_inicio = data_inicio
        evento.data_fim = data_fim
        evento.cidade = form.cidade.data
        evento.estado = form.estado.data
        evento.endereco = form.endereco.data
        evento.id_circo = form.id_circo.data
        evento.id_produtor = form.id_produtor.data
        evento.status = form.status.data
        evento.observacoes = form.observacoes.data

        # Gerenciar mudança de produtor na equipe do evento
        if produtor_anterior != novo_produtor:
            # Obter nomes dos produtores para as observações
            nome_produtor_anterior = None
            nome_novo_produtor = None
            
            if produtor_anterior:
                colaborador_anterior = Colaborador.query.get(produtor_anterior)
                nome_produtor_anterior = colaborador_anterior.nome if colaborador_anterior else f"ID {produtor_anterior}"
            
            if novo_produtor:
                colaborador_novo = Colaborador.query.get(novo_produtor)
                nome_novo_produtor = colaborador_novo.nome if colaborador_novo else f"ID {novo_produtor}"
            
            data_alteracao = datetime.now().strftime('%d/%m/%Y às %H:%M')
            
            # Atualizar o registro do produtor anterior com observação sobre a modificação
            if produtor_anterior:
                equipe_anterior = EquipeEvento.query.filter_by(
                    id_evento=evento.id_evento,
                    id_colaborador=produtor_anterior,
                    funcao='Produtor'
                ).first()
                if equipe_anterior:
                    # Atualizar observações do produtor anterior
                    observacao_modificacao = f"Produtor substituído por {nome_novo_produtor} em {data_alteracao}"
                    if equipe_anterior.observacoes:
                        equipe_anterior.observacoes += f" | {observacao_modificacao}"
                    else:
                        equipe_anterior.observacoes = observacao_modificacao
                    
                    # Alterar função para indicar que não é mais o produtor ativo
                    equipe_anterior.funcao = 'Ex-Produtor'
                    print(f"Produtor anterior {nome_produtor_anterior} marcado como Ex-Produtor com observação")
            
            # Adicionar o novo produtor à equipe (se não existir já)
            if novo_produtor:
                equipe_existente = EquipeEvento.query.filter_by(
                    id_evento=evento.id_evento,
                    id_colaborador=novo_produtor,
                    funcao='Produtor'
                ).first()
                
                if not equipe_existente:
                    # Criar observação para o novo produtor
                    observacao_novo = f"Assumiu como produtor em {data_alteracao}"
                    if nome_produtor_anterior:
                        observacao_novo += f", substituindo {nome_produtor_anterior}"
                    
                    nova_equipe_produtor = EquipeEvento(
                        id_evento=evento.id_evento,
                        id_colaborador=novo_produtor,
                        funcao='Produtor',
                        observacoes=observacao_novo
                    )
                    db.session.add(nova_equipe_produtor)
                    print(f"Novo produtor {nome_novo_produtor} adicionado à equipe do evento")
                else:
                    # Se já existe na equipe, apenas atualizar para Produtor e adicionar observação
                    observacao_promocao = f"Promovido a Produtor em {data_alteracao}"
                    if nome_produtor_anterior:
                        observacao_promocao += f", substituindo {nome_produtor_anterior}"
                    
                    if equipe_existente.observacoes:
                        equipe_existente.observacoes += f" | {observacao_promocao}"
                    else:
                        equipe_existente.observacoes = observacao_promocao
                    
                    equipe_existente.funcao = 'Produtor'
                    print(f"Colaborador {nome_novo_produtor} promovido a Produtor na equipe do evento")

        # Exclui os itens marcados
        excluir_receitas = request.form.get('excluir_receita_ids', '')
        excluir_despesas = request.form.get('excluir_despesa_ids', '')

        if excluir_receitas:
            for rid in excluir_receitas.split(','):
                if rid.strip():
                    ReceitaEvento.query.filter_by(id_receita_evento=int(rid)).delete()

        if excluir_despesas:
            for did in excluir_despesas.split(','):
                if did.strip():
                    DespesaEvento.query.filter_by(id_despesa_evento=int(did)).delete()

        # Processar novas receitas
        receitas_ids = request.form.getlist('receita_id[]')
        receita_datas = request.form.getlist('receita_data[]')
        valores = request.form.getlist('valor[]')
        observacoes = request.form.getlist('obs[]')
        for i in range(len(receitas_ids)):
            try:
                valor = float(valores[i].replace(',', '.'))
                data_receita = receita_datas[i] if i < len(receita_datas) else None
                data = datetime.strptime(data_receita, '%Y-%m-%d').date() if data_receita else None
            except:
                continue
            if receitas_ids[i] and valor:
                db.session.add(ReceitaEvento(
                    id_evento=evento.id_evento,
                    id_receita=int(receitas_ids[i]),
                    valor=valor,
                    data=data,
                    observacoes=observacoes[i]
                ))

        # Processar despesas (novas e editadas)
        despesa_ids = request.form.getlist('despesa_id[]')
        datas_desp = request.form.getlist('despesa_data[]')
        valores_desp = request.form.getlist('despesa_valor[]')
        status_pag = request.form.getlist('despesa_status_pagamento[]')
        forma_pag = request.form.getlist('despesa_forma_pagamento[]')
        pago_por = request.form.getlist('despesa_pago_por[]')
        obs_desp = request.form.getlist('despesa_obs[]')
        
        print(f"=== DEBUG DESPESAS COMPLETO ===")
        print(f"Total de despesas recebidas: {len(despesa_ids)}")
        print(f"despesa_ids: {despesa_ids}")
        print(f"valores_desp: {valores_desp}")
        print(f"datas_desp: {datas_desp}")
        print(f"status_pag: {status_pag}")
        print(f"forma_pag: {forma_pag}")
        print(f"pago_por: {pago_por}")
        print(f"obs_desp: {obs_desp}")
        
        # Verificar todas as despesas já cadastradas no evento ANTES do processamento
        despesas_antes = DespesaEvento.query.filter_by(id_evento=evento.id_evento).all()
        print(f"\n=== DESPESAS NO EVENTO ANTES DO PROCESSAMENTO ===")
        for desp in despesas_antes:
            print(f"ID Despesa: {desp.id_despesa}, Valor atual: {desp.valor}, Tipo: {desp.despesa.id_tipo_despesa}")
        
        for i in range(len(despesa_ids)):
            try:
                # Verificar se tem ID da despesa (obrigatório)
                if not despesa_ids[i]:
                    print(f"Ãndice {i}: ID da despesa vazio, pulando...")
                    continue
                    
                # Verificar se tem valor preenchido
                if not valores_desp[i] or valores_desp[i].strip() == '':
                    print(f"Ãndice {i}: Valor da despesa vazio, pulando...")
                    continue
                    
                valor = float(valores_desp[i].replace(',', '.'))
                data = datetime.strptime(datas_desp[i], '%Y-%m-%d').date() if datas_desp[i] else date.today()
                id_despesa = int(despesa_ids[i])
                
                print(f"\n--- Processando ÃNDICE {i} ---")
                print(f"ID Despesa: {id_despesa}")
                print(f"Valor recebido: '{valores_desp[i]}'")
                print(f"Valor convertido: {valor}")
                print(f"Data: {data}")
                print(f"Status: {status_pag[i] if i < len(status_pag) else 'pendente'}")
                print(f"Forma: {forma_pag[i] if i < len(forma_pag) else 'débito'}")
                
            except (ValueError, IndexError) as e:
                print(f"ERRO ao processar despesa índice {i}: {e}")
                continue
            
            # Verificar se já existe uma DespesaEvento para esta despesa neste evento
            # PERMITIR MÚLTIPLAS INSTÂNCIAS DA MESMA DESPESA - Comentado para sempre criar nova
            # despesa_evento_existente = DespesaEvento.query.filter_by(
            #     id_evento=evento.id_evento,
            #     id_despesa=id_despesa
            # ).first()
            
            if False:  # Sempre criar nova despesa
                # ATUALIZAR despesa existente
                print(f"✅ ENCONTROU despesa existente!")
                print(f"   - Valor antigo: {despesa_evento_existente.valor}")
                print(f"   - Valor novo: {valor}")
                print(f"   - ID DespesaEvento: {despesa_evento_existente.id_despesa_evento}")
                
                despesa_evento_existente.data_vencimento = data
                despesa_evento_existente.valor = valor
                despesa_evento_existente.status_pagamento = status_pag[i] if i < len(status_pag) else 'pendente'
                despesa_evento_existente.forma_pagamento = forma_pag[i] if i < len(forma_pag) else 'débito'
                despesa_evento_existente.pago_por = pago_por[i] if i < len(pago_por) else ''
                despesa_evento_existente.observacoes = obs_desp[i] if i < len(obs_desp) else ''
                
                print(f"   - Despesa atualizada na sessão!")
            else:
                # CRIAR nova despesa no evento
                print(f"🆕 NÃO encontrou despesa existente, criando nova...")
                nova_despesa_evento = DespesaEvento(
                    id_evento=evento.id_evento,
                    id_despesa=id_despesa,
                    data_vencimento=data,
                    valor=valor,
                    status_pagamento=status_pag[i] if i < len(status_pag) else 'pendente',
                    forma_pagamento=forma_pag[i] if i < len(forma_pag) else 'débito',
                    pago_por=pago_por[i] if i < len(pago_por) else '',
                    observacoes=obs_desp[i] if i < len(obs_desp) else ''
                )
                db.session.add(nova_despesa_evento)
                print(f"   - Nova despesa adicionada à sessão!")

        # Processar checkboxes de despesa_cabeca
        despesas_cabeca_ids = request.form.getlist('despesa_cabeca_evento[]')
        print(f"=== DEBUG DESPESAS CABEÇA ===")
        print(f"IDs das despesas marcadas como cabeça: {despesas_cabeca_ids}")
        
        # Primeiro, desmarcar todas as despesas do evento
        todas_despesas_evento = DespesaEvento.query.filter_by(id_evento=evento.id_evento).all()
        for despesa_evento in todas_despesas_evento:
            despesa_evento.despesa_cabeca = False
        
        # Marcar apenas as selecionadas
        for despesa_evento_id in despesas_cabeca_ids:
            if despesa_evento_id:
                despesa_evento = DespesaEvento.query.get(int(despesa_evento_id))
                if despesa_evento and despesa_evento.id_evento == evento.id_evento:
                    despesa_evento.despesa_cabeca = True
                    print(f"✅ Despesa evento ID {despesa_evento_id} marcada como cabeça")

        try:
            db.session.commit()
            print(f"âœ… COMMIT executado com sucesso! Evento ID: {evento.id_evento}")
            
            # Verificar se o evento foi realmente salvo
            evento_verificacao = Evento.query.get(evento.id_evento)
            if evento_verificacao:
                print(f"âœ… Evento verificado no banco: {evento_verificacao.nome}")
            else:
                print("âŒ ERRO: Evento não encontrado no banco após commit!")
            
            # Verificar as despesas APÃ"S o commit
            despesas_depois = DespesaEvento.query.filter_by(id_evento=evento.id_evento).all()
            print(f"\n=== DESPESAS NO EVENTO APÓS O COMMIT ===")
            for desp in despesas_depois:
                print(f"ID Despesa: {desp.id_despesa}, Valor SALVO: {desp.valor}, Tipo: {desp.despesa.id_tipo_despesa}")
            
            # Registrar log da operação
            registrar_log('Editar Evento', f'Evento "{evento.nome}" editado - {evento.cidade}/{evento.estado}')
            
            flash('Evento atualizado com sucesso!', 'success')
            return redirect(url_for('editar_evento', id=evento.id_evento))
            
        except Exception as e:
            print(f"âŒ ERRO no commit: {e}")
            db.session.rollback()
            flash(f'Erro ao salvar evento: {str(e)}', 'danger')
            # Não fazer redirect em caso de erro para debugar

    categorias_receita_dict = {
        c.id_categoria_receita: [{'id_receita': r.id_receita, 'nome': r.nome}
                                 for r in Receita.query.filter_by(id_categoria_receita=c.id_categoria_receita)]
        for c in categorias_receita
    }

    # Criar estrutura de despesas para ediÃ§ão - mostrar apenas as que faltam cadastrar
    categorias_despesa_dict = {}
    for categoria in categorias_despesa:
        # Obter todas as despesas disponíveis desta categoria
        despesas_categoria = Despesa.query.filter_by(
            id_categoria_despesa=categoria.id_categoria_despesa
        ).filter(Despesa.id_tipo_despesa.in_([1, 2])).all()
        
        despesas_fixas = [d for d in despesas_categoria if d.id_tipo_despesa == 1]
        despesas_variaveis = [d for d in despesas_categoria if d.id_tipo_despesa == 2]
        
        # IDs das despesas já cadastradas neste evento
        despesas_ja_cadastradas = set(d.id_despesa for d in despesas_salvas)
        
        # Criar um dicionário para acesso rápido aos dados das despesas já cadastradas
        despesas_evento_dict = {d.id_despesa: d for d in despesas_salvas}
        
        # Para as despesas fixas: sempre mostrar TODAS, mas com valores reais se já cadastradas
        categorias_despesa_dict[categoria.id_categoria_despesa] = {
            'fixas': [{
                'id_despesa': d.id_despesa, 
                'nome': d.nome,
                'valor_medio': float(d.valor_medio_despesa) if d.valor_medio_despesa else None,
                'tipo': 1,
                'ja_cadastrada': d.id_despesa in despesas_ja_cadastradas,
                'categoria_id': categoria.id_categoria_despesa,
                'categoria_nome': categoria.nome,
                # Se já foi cadastrada, usar valores reais do evento
                'valor_atual': float(despesas_evento_dict[d.id_despesa].valor) if d.id_despesa in despesas_ja_cadastradas else (float(d.valor_medio_despesa) if d.valor_medio_despesa else None),
                'data_atual': despesas_evento_dict[d.id_despesa].data_vencimento.strftime('%Y-%m-%d') if d.id_despesa in despesas_ja_cadastradas else None,
                'status_atual': despesas_evento_dict[d.id_despesa].status_pagamento if d.id_despesa in despesas_ja_cadastradas else 'pendente',
                'forma_atual': despesas_evento_dict[d.id_despesa].forma_pagamento if d.id_despesa in despesas_ja_cadastradas else 'débito',
                'fornecedor_atual': despesas_evento_dict[d.id_despesa].id_fornecedor if d.id_despesa in despesas_ja_cadastradas else None,
                'fornecedor_nome_atual': despesas_evento_dict[d.id_despesa].fornecedor.nome if d.id_despesa in despesas_ja_cadastradas and despesas_evento_dict[d.id_despesa].fornecedor else '',
                'pago_por_atual': despesas_evento_dict[d.id_despesa].pago_por if d.id_despesa in despesas_ja_cadastradas else '',
                'obs_atual': despesas_evento_dict[d.id_despesa].observacoes if d.id_despesa in despesas_ja_cadastradas else 'Despesa fixa automática',
                'despesa_cabeca_atual': despesas_evento_dict[d.id_despesa].despesa_cabeca if d.id_despesa in despesas_ja_cadastradas else False,
                'id_despesa_evento': despesas_evento_dict[d.id_despesa].id_despesa_evento if d.id_despesa in despesas_ja_cadastradas else None,
                'comprovante_atual': despesas_evento_dict[d.id_despesa].comprovante if d.id_despesa in despesas_ja_cadastradas else '',
                'qtd_dias_atual': despesas_evento_dict[d.id_despesa].qtd_dias if d.id_despesa in despesas_ja_cadastradas else None,
                'qtd_pessoas_atual': despesas_evento_dict[d.id_despesa].qtd_pessoas if d.id_despesa in despesas_ja_cadastradas else None,
                'valor_pago_socrates_atual': float(despesas_evento_dict[d.id_despesa].valor_pago_socrates) if d.id_despesa in despesas_ja_cadastradas and despesas_evento_dict[d.id_despesa].valor_pago_socrates else None
            } for d in despesas_fixas],  # TODAS as despesas fixas
            'variaveis': [{
                'id_despesa': d.id_despesa, 
                'nome': d.nome,
                'valor_medio': float(d.valor_medio_despesa) if d.valor_medio_despesa else None,
                'tipo': 2,
                'ja_cadastrada': d.id_despesa in despesas_ja_cadastradas,
                'categoria_id': categoria.id_categoria_despesa,
                'categoria_nome': categoria.nome,
                # Se já foi cadastrada, usar valores reais do evento
                'valor_atual': float(despesas_evento_dict[d.id_despesa].valor) if d.id_despesa in despesas_ja_cadastradas else (float(d.valor_medio_despesa) if d.valor_medio_despesa else None),
                'data_atual': despesas_evento_dict[d.id_despesa].data_vencimento.strftime('%Y-%m-%d') if d.id_despesa in despesas_ja_cadastradas else None,
                'status_atual': despesas_evento_dict[d.id_despesa].status_pagamento if d.id_despesa in despesas_ja_cadastradas else 'pendente',
                'forma_atual': despesas_evento_dict[d.id_despesa].forma_pagamento if d.id_despesa in despesas_ja_cadastradas else 'débito',
                'fornecedor_atual': despesas_evento_dict[d.id_despesa].id_fornecedor if d.id_despesa in despesas_ja_cadastradas else None,
                'fornecedor_nome_atual': despesas_evento_dict[d.id_despesa].fornecedor.nome if d.id_despesa in despesas_ja_cadastradas and despesas_evento_dict[d.id_despesa].fornecedor else '',
                'pago_por_atual': despesas_evento_dict[d.id_despesa].pago_por if d.id_despesa in despesas_ja_cadastradas else '',
                'obs_atual': despesas_evento_dict[d.id_despesa].observacoes if d.id_despesa in despesas_ja_cadastradas else '',
                'despesa_cabeca_atual': despesas_evento_dict[d.id_despesa].despesa_cabeca if d.id_despesa in despesas_ja_cadastradas else False,
                'id_despesa_evento': despesas_evento_dict[d.id_despesa].id_despesa_evento if d.id_despesa in despesas_ja_cadastradas else None,
                'comprovante_atual': despesas_evento_dict[d.id_despesa].comprovante if d.id_despesa in despesas_ja_cadastradas else '',
                'qtd_dias_atual': despesas_evento_dict[d.id_despesa].qtd_dias if d.id_despesa in despesas_ja_cadastradas else None,
                'qtd_pessoas_atual': despesas_evento_dict[d.id_despesa].qtd_pessoas if d.id_despesa in despesas_ja_cadastradas else None,
                'valor_pago_socrates_atual': float(despesas_evento_dict[d.id_despesa].valor_pago_socrates) if d.id_despesa in despesas_ja_cadastradas and despesas_evento_dict[d.id_despesa].valor_pago_socrates else None
            } for d in despesas_variaveis]  # TODAS as despesas variáveis (cadastradas e não cadastradas)
        }

    # Nova estrutura: listar TODAS as instâncias das despesas do evento por categoria
    despesas_evento_por_categoria = {}
    for categoria in categorias_despesa:
        # Filtrar despesas do evento por categoria
        despesas_desta_categoria = [
            de for de in despesas_salvas 
            if de.despesa.id_categoria_despesa == categoria.id_categoria_despesa
        ]
        
        despesas_evento_por_categoria[categoria.id_categoria_despesa] = {
            'fixas': [],
            'variaveis': [],
            'categoria_nome': categoria.nome
        }
        
        # Separar por tipo
        for despesa_evento in despesas_desta_categoria:
            despesa_data = {
                'id_despesa': despesa_evento.id_despesa,
                'id_despesa_evento': despesa_evento.id_despesa_evento,
                'nome': despesa_evento.despesa.nome,
                'categoria_id': categoria.id_categoria_despesa,
                'categoria_nome': categoria.nome,
                'ja_cadastrada': True,
                'valor_atual': float(despesa_evento.valor),
                'data_atual': despesa_evento.data_vencimento.strftime('%Y-%m-%d'),
                'status_atual': despesa_evento.status_pagamento,
                'forma_atual': despesa_evento.forma_pagamento,
                'fornecedor_atual': despesa_evento.id_fornecedor,
                'fornecedor_nome_atual': despesa_evento.fornecedor.nome if despesa_evento.fornecedor else '',
                'pago_por_atual': despesa_evento.pago_por or '',
                'obs_atual': despesa_evento.observacoes or '',
                'despesa_cabeca_atual': despesa_evento.despesa_cabeca,
                'comprovante_atual': despesa_evento.comprovante or '',
                'qtd_dias_atual': despesa_evento.qtd_dias,
                'qtd_pessoas_atual': despesa_evento.qtd_pessoas,
                'valor_pago_socrates_atual': float(despesa_evento.valor_pago_socrates) if despesa_evento.valor_pago_socrates else None,
                'tipo': despesa_evento.despesa.id_tipo_despesa
            }
            
            if despesa_evento.despesa.id_tipo_despesa == 1:  # Fixa
                despesas_evento_por_categoria[categoria.id_categoria_despesa]['fixas'].append(despesa_data)
            else:  # Variável
                despesas_evento_por_categoria[categoria.id_categoria_despesa]['variaveis'].append(despesa_data)

    return render_template(
        'novo_evento.html',
        form=form,
        categorias_receita=categorias_receita,
        categorias_receita_dict=categorias_receita_dict,
        categorias_despesa=categorias_despesa,
        categorias_despesa_dict=categorias_despesa_dict,
        despesas_evento_por_categoria=despesas_evento_por_categoria,  # Nova estrutura
        fornecedores=[{'id_fornecedor': f.id_fornecedor, 'nome': f.nome} for f in Fornecedor.query.all()],
        receitas_salvas=receitas_salvas,
        despesas_salvas=despesas_salvas,
        current_date=date.today().isoformat(),
        is_editing=True
    )

@app.route('/eventos/excluir/<int:id>')
def excluir_evento(id):
    try:
        print(f"Tentando excluir evento com ID: {id}")
        
        # Buscar o evento primeiro
        evento = Evento.query.get(id)
        if not evento:
            flash('Evento não encontrado!', 'error')
            return redirect(url_for('listar_eventos'))
        
        nome_evento = evento.nome
        print(f"Deletando dependências do evento {nome_evento}...")
        
        # Fechar a sessão atual para liberar locks
        db.session.close()
        
        # Criar uma nova sessão para as operações de exclusão
        # PRAGMAs são específicos do SQLite; proteger para outros bancos (ex.: PostgreSQL)
        if db.engine.dialect.name == 'sqlite':
            db.session.execute(text("PRAGMA foreign_keys=OFF"))
            db.session.execute(text("PRAGMA journal_mode=DELETE"))
        
        # Deletar todas as dependências
        print("Iniciando exclusão das dependências...")
        
        # 1. Despesas do evento
        despesas_deletadas = db.session.execute(
            text("DELETE FROM despesas_evento WHERE id_evento = :id_evento"),
            {"id_evento": id}
        ).rowcount
        print(f"Despesas deletadas: {despesas_deletadas}")
        
        # 2. Receitas do evento  
        receitas_deletadas = db.session.execute(
            text("DELETE FROM receitas_evento WHERE id_evento = :id_evento"),
            {"id_evento": id}
        ).rowcount
        print(f"Receitas deletadas: {receitas_deletadas}")
        
        # 3. Equipe do evento
        equipe_deletada = db.session.execute(
            text("DELETE FROM equipe_evento WHERE id_evento = :id_evento"),
            {"id_evento": id}
        ).rowcount
        print(f"Equipe deletada: {equipe_deletada}")
        
        # 4. Elenco do evento
        elenco_deletado = db.session.execute(
            text("DELETE FROM elenco_evento WHERE id_evento = :id_evento"),
            {"id_evento": id}
        ).rowcount
        print(f"Elenco deletado: {elenco_deletado}")
        
        # 5. Fornecedores do evento
        fornecedores_deletados = db.session.execute(
            text("DELETE FROM fornecedor_evento WHERE id_evento = :id_evento"),
            {"id_evento": id}
        ).rowcount
        print(f"Fornecedores deletados: {fornecedores_deletados}")
        
        # 6. Veículos do evento
        veiculos_deletados = db.session.execute(
            text("DELETE FROM veiculos_evento WHERE id_evento = :id_evento"),
            {"id_evento": id}
        ).rowcount
        print(f"Veículos deletados: {veiculos_deletados}")
        
        # 7. Deletar o evento principal
        evento_deletado = db.session.execute(
            text("DELETE FROM evento WHERE id_evento = :id_evento"),
            {"id_evento": id}
        ).rowcount
        print(f"Evento deletado: {evento_deletado}")
        
        # Commit das alterações
        db.session.commit()
        
        # Reabilitar foreign keys (apenas SQLite)
        if db.engine.dialect.name == 'sqlite':
            db.session.execute(text("PRAGMA foreign_keys=ON"))
            db.session.commit()
        
        # Registrar log da operação
        registrar_log('Excluir Evento', f'Evento "{nome_evento}" excluído')
        
        print(f"Evento '{nome_evento}' excluído com sucesso!")
        flash(f'Evento "{nome_evento}" excluído com sucesso!', 'success')
        
    except Exception as e:
        print(f"Erro ao deletar evento: {e}")
        try:
            db.session.rollback()
        except:
            pass
        flash(f'Erro ao excluir evento: {str(e)}', 'error')
    
    return redirect(url_for('listar_eventos'))

@app.route('/relatorios/lucratividade-mensal')
def relatorios_lucratividade_mensal():
    if session.get('categoria', '').lower() != 'administrativo':
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Obter mês e ano dos parâmetros ou usar mês atual como padrão
    mes_param = request.args.get('mes')
    ano_param = request.args.get('ano')
    
    # Se não foram fornecidos, usar mês atual
    hoje = date.today()
    mes_selecionado = int(mes_param) if mes_param else hoje.month
    ano_selecionado = int(ano_param) if ano_param else hoje.year
    
    # Calcular primeiro e último dia do mês selecionado
    primeiro_dia = date(ano_selecionado, mes_selecionado, 1)
    if mes_selecionado == 12:
        ultimo_dia = date(ano_selecionado + 1, 1, 1) - timedelta(days=1)
    else:
        ultimo_dia = date(ano_selecionado, mes_selecionado + 1, 1) - timedelta(days=1)
        
    # Buscar receitas do mês (eventos + empresa)
    # 1. Receitas de eventos com data no período
    receitas_evento = db.session.query(func.sum(ReceitaEvento.valor)).filter(
            ReceitaEvento.data >= primeiro_dia,
            ReceitaEvento.data <= ultimo_dia
        ).scalar() or 0
        
    # 2. Receitas da empresa com data no período
    receitas_empresa = db.session.query(func.sum(ReceitaEmpresa.valor)).filter(
        ReceitaEmpresa.data >= primeiro_dia,
        ReceitaEmpresa.data <= ultimo_dia
    ).scalar() or 0
    
    # Total de receitas
    receitas_mes = receitas_evento + receitas_empresa
    
    # Buscar despesas do mês (eventos pagas + empresa pagas)
    # 1. Despesas de eventos com status pago e data de pagamento no período
    # (usa data_vencimento como fallback se data_pagamento for NULL)
    despesas_evento_pagas = db.session.query(func.sum(DespesaEvento.valor)).filter(
        DespesaEvento.status_pagamento == 'pago',
        or_(
            # Prioridade 1: data_pagamento preenchida e no período
            and_(
                DespesaEvento.data_pagamento.isnot(None),
                DespesaEvento.data_pagamento >= primeiro_dia,
                DespesaEvento.data_pagamento <= ultimo_dia
            ),
            # Fallback: data_pagamento NULL mas data_vencimento no período
            and_(
                DespesaEvento.data_pagamento.is_(None),
                DespesaEvento.data_vencimento >= primeiro_dia,
                DespesaEvento.data_vencimento <= ultimo_dia
            )
        )
    ).scalar() or 0
    
    # 2. Despesas da empresa com status pago e data de pagamento no período
    # (usa data_vencimento como fallback se data_pagamento for NULL)
    despesas_empresa_pagas = db.session.query(func.sum(DespesaEmpresa.valor)).filter(
        DespesaEmpresa.status_pagamento == 'pago',
        or_(
            # Prioridade 1: data_pagamento preenchida e no período
            and_(
                DespesaEmpresa.data_pagamento.isnot(None),
                DespesaEmpresa.data_pagamento >= primeiro_dia,
                DespesaEmpresa.data_pagamento <= ultimo_dia
            ),
            # Fallback: data_pagamento NULL mas data_vencimento no período
            and_(
                DespesaEmpresa.data_pagamento.is_(None),
                DespesaEmpresa.data_vencimento >= primeiro_dia,
                DespesaEmpresa.data_vencimento <= ultimo_dia
            )
        )
    ).scalar() or 0
    
    # Total de despesas
    despesas_mes = despesas_evento_pagas + despesas_empresa_pagas
    
    lucro_mes = receitas_mes - despesas_mes
    
    # Buscar receitas detalhadas por categoria (eventos + empresa)
    # 1. Receitas de eventos por categoria
    receitas_evento_categoria = db.session.query(
        CategoriaReceita.nome.label('categoria'),
        func.sum(ReceitaEvento.valor).label('total')
    ).join(
        Receita, ReceitaEvento.id_receita == Receita.id_receita
    ).join(
        CategoriaReceita, Receita.id_categoria_receita == CategoriaReceita.id_categoria_receita
    ).filter(
        ReceitaEvento.data >= primeiro_dia,
        ReceitaEvento.data <= ultimo_dia
    ).group_by(CategoriaReceita.nome).all()
    
    # 2. Receitas da empresa por categoria
    receitas_empresa_categoria = db.session.query(
        CategoriaReceita.nome.label('categoria'),
        func.sum(ReceitaEmpresa.valor).label('total')
    ).join(
        Receita, ReceitaEmpresa.id_receita == Receita.id_receita
    ).join(
        CategoriaReceita, Receita.id_categoria_receita == CategoriaReceita.id_categoria_receita
    ).filter(
        ReceitaEmpresa.data >= primeiro_dia,
        ReceitaEmpresa.data <= ultimo_dia
    ).group_by(CategoriaReceita.nome).all()
    
    # Combinar receitas de eventos e empresa por categoria
    receitas_combinadas = {}
    for item in receitas_evento_categoria:
        if item.categoria not in receitas_combinadas:
            receitas_combinadas[item.categoria] = 0
        receitas_combinadas[item.categoria] += float(item.total)
    
    for item in receitas_empresa_categoria:
        if item.categoria not in receitas_combinadas:
            receitas_combinadas[item.categoria] = 0
        receitas_combinadas[item.categoria] += float(item.total)
    
    # Converter para formato esperado pelo template
    receitas_por_categoria = [
        type('obj', (object,), {'categoria': categoria, 'total': total})()
        for categoria, total in receitas_combinadas.items()
    ]
    
    # Buscar despesas detalhadas por categoria (eventos pagas + empresa pagas)
    # 1. Despesas de eventos pagas por categoria
    # (usa data_vencimento como fallback se data_pagamento for NULL)
    despesas_evento_categoria = db.session.query(
        CategoriaDespesa.nome.label('categoria'),
        func.sum(DespesaEvento.valor).label('total')
    ).join(
            Despesa, DespesaEvento.id_despesa == Despesa.id_despesa
        ).join(
            CategoriaDespesa, Despesa.id_categoria_despesa == CategoriaDespesa.id_categoria_despesa
        ).filter(
        DespesaEvento.status_pagamento == 'pago',
        or_(
            # Prioridade 1: data_pagamento preenchida e no período
            and_(
                DespesaEvento.data_pagamento.isnot(None),
                DespesaEvento.data_pagamento >= primeiro_dia,
                DespesaEvento.data_pagamento <= ultimo_dia
            ),
            # Fallback: data_pagamento NULL mas data_vencimento no período
            and_(
                DespesaEvento.data_pagamento.is_(None),
                DespesaEvento.data_vencimento >= primeiro_dia,
                DespesaEvento.data_vencimento <= ultimo_dia
            )
        )
    ).group_by(CategoriaDespesa.nome).all()
    
    # 2. Despesas da empresa pagas por categoria
    # (usa data_vencimento como fallback se data_pagamento for NULL)
    despesas_empresa_categoria = db.session.query(
        CategoriaDespesa.nome.label('categoria'),
        func.sum(DespesaEmpresa.valor).label('total')
    ).join(
        Despesa, DespesaEmpresa.id_despesa == Despesa.id_despesa
    ).join(
        CategoriaDespesa, Despesa.id_categoria_despesa == CategoriaDespesa.id_categoria_despesa
    ).filter(
        DespesaEmpresa.status_pagamento == 'pago',
        or_(
            # Prioridade 1: data_pagamento preenchida e no período
            and_(
                DespesaEmpresa.data_pagamento.isnot(None),
                DespesaEmpresa.data_pagamento >= primeiro_dia,
                DespesaEmpresa.data_pagamento <= ultimo_dia
            ),
            # Fallback: data_pagamento NULL mas data_vencimento no período
            and_(
                DespesaEmpresa.data_pagamento.is_(None),
                DespesaEmpresa.data_vencimento >= primeiro_dia,
                DespesaEmpresa.data_vencimento <= ultimo_dia
            )
        )
    ).group_by(CategoriaDespesa.nome).all()
    
    # Combinar despesas de eventos e empresa por categoria
    despesas_combinadas = {}
    for item in despesas_evento_categoria:
        if item.categoria not in despesas_combinadas:
            despesas_combinadas[item.categoria] = 0
        despesas_combinadas[item.categoria] += float(item.total)
    
    for item in despesas_empresa_categoria:
        if item.categoria not in despesas_combinadas:
            despesas_combinadas[item.categoria] = 0
        despesas_combinadas[item.categoria] += float(item.total)

    # Converter para formato esperado pelo template
    despesas_por_categoria = [
        type('obj', (object,), {'categoria': categoria, 'total': total})()
        for categoria, total in despesas_combinadas.items()
    ]
    
    # Buscar eventos do mês
    eventos_mes = Evento.query.filter(
        Evento.data_inicio >= primeiro_dia,
        Evento.data_inicio <= ultimo_dia
    ).all()
    
    # Lista de meses para o seletor
    meses_opcoes = [
        (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
        (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
        (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro')
    ]
    
    # Lista de anos (últimos 5 anos + próximos 2)
    ano_atual = hoje.year
    anos_opcoes = list(range(ano_atual - 5, ano_atual + 3))

    return render_template(
        'relatorios_lucratividade_mensal.html',
        mes_selecionado=mes_selecionado,
        ano_selecionado=ano_selecionado,
        primeiro_dia=primeiro_dia,
        ultimo_dia=ultimo_dia,
        receitas_mes=float(receitas_mes),
        despesas_mes=float(despesas_mes),
        lucro_mes=float(lucro_mes),
        receitas_por_categoria=receitas_por_categoria,
        despesas_por_categoria=despesas_por_categoria,
        eventos_mes=eventos_mes,
        meses_opcoes=meses_opcoes,
        anos_opcoes=anos_opcoes,
        total_eventos=len(eventos_mes),
        hoje=hoje  # Adicionar data atual para o template
    )

@app.route('/relatorios/faturamento-evento')
def relatorios_faturamento_evento():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # Obter informações do usuário logado
    usuario = Usuario.query.get(session['user_id'])
    if not usuario or not usuario.colaborador:
        flash('Erro ao carregar informações do usuário.', 'danger')
        return redirect(url_for('login'))
    
    # Verificar se é administrador
    is_admin = is_admin_user()
    
    if not is_admin:
        # Verificar se é produtor
        is_produtor = any(cat.nome.lower() == 'produtor' for cat in usuario.colaborador.categorias)
        if not is_produtor:
            flash('Acesso restrito a administradores e produtores.', 'danger')
            return redirect(url_for('dashboard'))

    # Processar período selecionado
    period = request.args.get('period', '90dias')
    
    if period == 'custom':
        data_inicio, data_fim = obter_datas_filtro_padrao(
            request.args.get('data_inicio'),
            request.args.get('data_fim')
        )
    else:
        # Definir datas baseado no período selecionado
        data_fim = date.today()
        if period == 'hoje':
            data_inicio = data_fim
        elif period == 'ontem':
            data_inicio = data_fim = data_fim - timedelta(days=1)
        elif period == '7dias':
            data_inicio = data_fim - timedelta(days=7)
        elif period == '30dias':
            data_inicio = data_fim - timedelta(days=30)
        else:  # 90dias (padrão)
            data_inicio = data_fim - timedelta(days=90)
        
        data_inicio = data_inicio.strftime('%Y-%m-%d')
        data_fim = data_fim.strftime('%Y-%m-%d')
    
    # Filtrar eventos baseado no tipo de usuário
    eventos_query = Evento.query
    
    if not is_admin:
        # Produtores veem apenas seus eventos
        eventos_query = eventos_query.filter_by(id_produtor=usuario.colaborador.id_colaborador)
    
    # Aplicar filtros de data
    eventos_query = eventos_query.filter(Evento.data_inicio >= datetime.strptime(data_inicio, '%Y-%m-%d').date())
    eventos_query = eventos_query.filter(Evento.data_fim <= datetime.strptime(data_fim, '%Y-%m-%d').date())
    
    eventos = eventos_query.order_by(Evento.data_inicio.desc()).all()
    
    # Calcular lucratividade dos eventos usando a função unificada
    eventos_lucratividade = []
    for evento in eventos:
        calculo = calcular_lucro_evento(evento.id_evento)
        
        eventos_lucratividade.append({
            'nome': evento.nome,
            'receitas': calculo['total_receitas'],
            'despesas': calculo['total_despesas'],
            'lucro': calculo['resultado_show'],
            'evento': evento
        })
    
    # Ordenar por lucro (do maior para o menor) e pegar top 10
    eventos_mais_lucrativos = sorted(eventos_lucratividade, key=lambda x: x['lucro'], reverse=True)[:10]
    
    # Preparar dados para o gráfico (apenas lucro)
    nomes_eventos = [e['nome'][:25] + '...' if len(e['nome']) > 25 else e['nome'] for e in eventos_mais_lucrativos]
    lucros_eventos = [e['lucro'] for e in eventos_mais_lucrativos]
    
    return render_template(
        'relatorios_faturamento_evento.html', 
        eventos=eventos, 
        data_inicio=data_inicio, 
        data_fim=data_fim,
        period=period,
        nomes_eventos=nomes_eventos,
        lucros_eventos=lucros_eventos,
        total_eventos=len(eventos),
        is_admin=is_admin,
        usuario=usuario
    )

@app.route('/relatorios/faturamento-evento/<int:id_evento>')
def relatorio_faturamento_evento(id_evento):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # Obter informações do usuário logado
    usuario = Usuario.query.get(session['user_id'])
    if not usuario or not usuario.colaborador:
        flash('Erro ao carregar informações do usuário.', 'danger')
        return redirect(url_for('login'))
    
    # Verificar se é administrador
    is_admin = is_admin_user()
    
    evento = Evento.query.get_or_404(id_evento)
    
    # Verificar se o usuário tem permissão para ver este evento
    if not is_admin:
        # Verificar se é produtor
        is_produtor = any(cat.nome.lower() == 'produtor' for cat in usuario.colaborador.categorias)
        if not is_produtor:
            flash('Acesso restrito a administradores e produtores.', 'danger')
            return redirect(url_for('dashboard'))
        
        # Verificar se é o produtor deste evento
        if evento.id_produtor != usuario.colaborador.id_colaborador:
            flash('Você só pode visualizar relatórios dos seus próprios eventos.', 'danger')
            return redirect(url_for('relatorios_faturamento_evento'))
    
    receitas = ReceitaEvento.query.filter_by(id_evento=id_evento).all()
    despesas = DespesaEvento.query.filter_by(id_evento=id_evento).all()
    receitas_labels = [r.receita.nome for r in receitas]
    receitas_values = [float(r.valor) for r in receitas]
    despesas_labels = [d.despesa.nome for d in despesas]
    despesas_values = [float(d.valor) for d in despesas]
    
    return render_template(
        'relatorio_faturamento_evento.html', 
        evento=evento, 
        receitas_labels=receitas_labels, 
        receitas_values=receitas_values, 
        despesas_labels=despesas_labels, 
        despesas_values=despesas_values,
        is_admin=is_admin,
        usuario=usuario
    )

@app.route('/relatorios/fechamento-evento')
def relatorios_fechamento_evento():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # Obter informações do usuário logado
    usuario = Usuario.query.get(session['user_id'])
    if not usuario or not usuario.colaborador:
        flash('Erro ao carregar informações do usuário.', 'danger')
        return redirect(url_for('login'))
    
    # Verificar se é administrador
    is_admin = is_admin_user()
    
    # Verificar se tem permissão (administrador ou produtor)
    if not is_admin:
        is_produtor = any(cat.nome.lower() == 'produtor' for cat in usuario.colaborador.categorias)
        if not is_produtor:
            flash('Acesso restrito a administradores e produtores.', 'danger')
            return redirect(url_for('dashboard'))

    # Obter datas do filtro com padrão de 90 dias
    data_inicio, data_fim = obter_datas_filtro_padrao(
        request.args.get('data_inicio'),
        request.args.get('data_fim')
    )

    # Query base de eventos
    eventos_query = Evento.query.filter(Evento.status.in_(['planejamento', 'a realizar', 'em andamento', 'realizado']))
    
    if not is_admin:
        # Produtores veem apenas seus eventos
        eventos_query = eventos_query.filter_by(id_produtor=usuario.colaborador.id_colaborador)
    
    # Aplicar filtros de data
    eventos_query = eventos_query.filter(Evento.data_inicio >= datetime.strptime(data_inicio, '%Y-%m-%d').date())
    eventos_query = eventos_query.filter(Evento.data_inicio <= datetime.strptime(data_fim, '%Y-%m-%d').date())
    
    eventos = eventos_query.order_by(Evento.data_inicio.desc()).all()
    
    return render_template('relatorios_fechamento_evento.html', 
                         eventos=eventos,
                         total_eventos=len(eventos),
                         data_inicio=data_inicio,
                         data_fim=data_fim,
                         is_admin=is_admin,
                         usuario=usuario)

@app.route('/relatorios/fechamento-evento/<int:id_evento>')
def relatorio_fechamento_evento(id_evento):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # Obter informações do usuário logado
    usuario = Usuario.query.get(session['user_id'])
    if not usuario or not usuario.colaborador:
        flash('Erro ao carregar informações do usuário.', 'danger')
        return redirect(url_for('login'))
    
    # Verificar se é administrador
    is_admin = is_admin_user()

    # Usar a função unificada para obter dados
    dados = obter_dados_completos_evento(id_evento)
    evento = dados['evento']
    
    # Verificar se o usuário tem permissão para ver este evento
    if not is_admin:
        # Verificar se é produtor
        is_produtor = any(cat.nome.lower() == 'produtor' for cat in usuario.colaborador.categorias)
        if not is_produtor:
            flash('Acesso restrito a administradores e produtores.', 'danger')
            return redirect(url_for('dashboard'))
        
        # Verificar se é o produtor deste evento
        if evento.id_produtor != usuario.colaborador.id_colaborador:
            flash('Você só pode visualizar relatórios dos seus próprios eventos.', 'danger')
            return redirect(url_for('relatorios_fechamento_evento'))
    
    # Obter dados da função unificada
    despesas_cabeca = dados['despesas_cabeca_agrupadas']
    receitas_evento = dados['receitas_agrupadas']
    todas_despesas = dados['despesas_agrupadas']
    
    # Usar valores da função unificada para garantir consistência
    calculo = dados['calculo_financeiro']
    total_despesas_bruto = dados['totais_calculados']['despesas_cabeca_total']
    reembolso_midias = dados['totais_calculados']['reembolso_midias']
    total_receitas = calculo['total_receitas']
    total_liquido = calculo['total_liquido']
    cinquenta_porcento_show = calculo['cinquenta_porcento_show']
    repasse_total = calculo['repasse_total']
    # Usar o total baseado nos valores exibidos ao invés do cálculo da função
    total_despesas_socrates = dados['totais_calculados']['total_despesas_socrates_exibidas']
    # Recalcular o resultado do show com os valores corretos
    resultado_show = repasse_total - total_despesas_socrates
    
    return render_template('relatorio_fechamento_evento.html', 
                         evento=evento,
                         usuario=usuario,
                         is_admin=is_admin,
                         despesas_cabeca=despesas_cabeca,
                         total_despesas_bruto=total_despesas_bruto,
                         receitas_evento=receitas_evento,
                         total_receitas=total_receitas,
                         total_liquido=total_liquido,
                         cinquenta_porcento_show=cinquenta_porcento_show,
                         reembolso_midias=reembolso_midias,
                         repasse_total=repasse_total,
                         todas_despesas=todas_despesas,
                         total_despesas_socrates=total_despesas_socrates,
                         resultado_show=resultado_show)

@app.route('/cadastros/categorias-veiculo', methods=['GET', 'POST'])
def cadastrar_categoria_veiculo():
    form = CategoriaVeiculoForm()
    if form.validate_on_submit():
        nova = CategoriaVeiculo(nome=form.nome.data)
        db.session.add(nova)
        db.session.commit()
        flash('Categoria cadastrada com sucesso!', 'success')
        return redirect(url_for('cadastrar_categoria_veiculo'))
    categorias = CategoriaVeiculo.query.all()
    return render_template('categorias_veiculo.html', form=form, categorias=categorias)

@app.route('/cadastros/categorias-veiculo/editar/<int:id>', methods=['GET', 'POST'])
def editar_categoria_veiculo(id):
    categoria = CategoriaVeiculo.query.get_or_404(id)
    form = CategoriaVeiculoForm(obj=categoria)
    if form.validate_on_submit():
        categoria.nome = form.nome.data
        db.session.commit()
        flash('Categoria atualizada com sucesso!', 'success')
        return redirect(url_for('cadastrar_categoria_veiculo'))
    categorias = CategoriaVeiculo.query.all()
    return render_template('categorias_veiculo.html', form=form, categorias=categorias)

@app.route('/cadastros/categorias-veiculo/excluir/<int:id>')
def excluir_categoria_veiculo(id):
    categoria = CategoriaVeiculo.query.get_or_404(id)
    
    # Verificar se existem veículos usando esta categoria
    veiculos_usando = Veiculo.query.filter_by(id_categoria_veiculo=id).count()
    if veiculos_usando > 0:
        flash(f'Não é possível excluir esta categoria pois existem {veiculos_usando} veículo(s) associado(s) a ela.', 'danger')
        return redirect(url_for('cadastrar_categoria_veiculo'))
    
    db.session.delete(categoria)
    db.session.commit()
    flash('Categoria excluída com sucesso!', 'success')
    return redirect(url_for('cadastrar_categoria_veiculo'))

@app.route('/cadastros/veiculos', methods=['GET', 'POST'])
def cadastrar_veiculo():
    # Verificar se existem categorias
    categorias_existentes = CategoriaVeiculo.query.all()
    if not categorias_existentes:
        flash('É necessário cadastrar pelo menos uma categoria de veículo antes de cadastrar veículos.', 'warning')
        return redirect(url_for('cadastrar_categoria_veiculo'))
    
    form = VeiculoForm()
    form.id_categoria_veiculo.choices = [(c.id_categoria_veiculo, c.nome) for c in categorias_existentes]
    if form.validate_on_submit():
        novo = Veiculo(
            nome=form.nome.data,
            modelo=form.modelo.data,
            marca=form.marca.data,
            ano=form.ano.data,
            placa=form.placa.data,
            cor=form.cor.data,
            combustivel=form.combustivel.data,
            capacidade_passageiros=form.capacidade_passageiros.data,
            media_km_litro=form.media_km_litro.data,
            observacoes=form.observacoes.data,
            id_categoria_veiculo=form.id_categoria_veiculo.data
        )
        db.session.add(novo)
        db.session.commit()
        flash('Veículo cadastrado com sucesso!', 'success')
        return redirect(url_for('cadastrar_veiculo'))
    veiculos = Veiculo.query.all()
    return render_template('veiculos.html', form=form, veiculos=veiculos)

@app.route('/cadastros/veiculos/editar/<int:id>', methods=['GET', 'POST'])
def editar_veiculo(id):
    # Verificar se existem categorias
    categorias_existentes = CategoriaVeiculo.query.all()
    if not categorias_existentes:
        flash('É necessário cadastrar pelo menos uma categoria de veículo antes de editar veículos.', 'warning')
        return redirect(url_for('cadastrar_categoria_veiculo'))
    
    veiculo = Veiculo.query.get_or_404(id)
    form = VeiculoForm(obj=veiculo)
    form.id_categoria_veiculo.choices = [(c.id_categoria_veiculo, c.nome) for c in categorias_existentes]
    if form.validate_on_submit():
        veiculo.nome = form.nome.data
        veiculo.modelo = form.modelo.data
        veiculo.marca = form.marca.data
        veiculo.ano = form.ano.data
        veiculo.placa = form.placa.data
        veiculo.cor = form.cor.data
        veiculo.combustivel = form.combustivel.data
        veiculo.capacidade_passageiros = form.capacidade_passageiros.data
        veiculo.media_km_litro = form.media_km_litro.data
        veiculo.observacoes = form.observacoes.data
        veiculo.id_categoria_veiculo = form.id_categoria_veiculo.data
        db.session.commit()
        flash('Veículo atualizado com sucesso!', 'success')
        return redirect(url_for('cadastrar_veiculo'))
    veiculos = Veiculo.query.all()
    return render_template('veiculos.html', form=form, veiculos=veiculos)

@app.route('/cadastros/veiculos/excluir/<int:id>')
def excluir_veiculo(id):
    veiculo = Veiculo.query.get_or_404(id)
    # Veículos normalmente não tÃªm dependÃªncias diretas no sistema atual
    # Mas se houver tabelas de relacionamento no futuro, adicionar aqui
    db.session.delete(veiculo)
    db.session.commit()
    flash('Veículo excluído com sucesso!', 'success')
    return redirect(url_for('cadastrar_veiculo'))

# =============================================================================
# ROTAS PARA SERVIÇOS DE VEÍCULOS (MULTAS, IPVA, LICENCIAMENTO, MANUTENÇÃO)
# =============================================================================

# ==================== MULTAS ====================

@app.route('/cadastros/veiculos/<int:id_veiculo>/multas')
def listar_multas_veiculo(id_veiculo):
    """Lista todas as multas de um veículo específico"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    multas = MultaVeiculo.query.filter_by(id_veiculo=id_veiculo).order_by(MultaVeiculo.data_infracao.desc()).all()
    
    # Se for requisição para modal, retornar apenas o conteúdo
    if request.args.get('modal') == '1':
        return render_template('modal_content_servicos.html', 
                             veiculo=veiculo, 
                             servicos=multas, 
                             tipo_servico='multas',
                             titulo='Multas')
    
    return render_template('servicos_veiculo.html', 
                         veiculo=veiculo, 
                         servicos=multas, 
                         tipo_servico='multas',
                         titulo='Multas')

@app.route('/cadastros/veiculos/<int:id_veiculo>/multas/nova', methods=['GET', 'POST'])
def nova_multa_veiculo(id_veiculo):
    """Cadastra nova multa para o veículo"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    form = MultaVeiculoForm()
    
    if form.validate_on_submit():
        try:
            nova_multa = MultaVeiculo(
                id_veiculo=id_veiculo,
                numero_ait=form.numero_ait.data,
                data_infracao=form.data_infracao.data,
                data_vencimento=form.data_vencimento.data,
                data_pagamento=form.data_pagamento.data,
                valor_original=float(form.valor_original.data),
                valor_pago=float(form.valor_pago.data) if form.valor_pago.data else None,
                local_infracao=form.local_infracao.data,
                tipo_infracao=form.tipo_infracao.data,
                orgao_autuador=form.orgao_autuador.data,
                status=form.status.data,
                observacoes=form.observacoes.data
            )
            
            db.session.add(nova_multa)
            db.session.commit()
            
            # Registrar log
            registrar_log(
                acao="Cadastrar Multa Veículo",
                descricao=f"Nova multa cadastrada para o veículo {veiculo.nome} - {form.tipo_infracao.data}"
            )
            
            # Se for requisição AJAX, retornar JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': True,
                    'message': 'Multa cadastrada com sucesso!'
                })
            
            flash('Multa cadastrada com sucesso!', 'success')
            return redirect(url_for('listar_multas_veiculo', id_veiculo=id_veiculo))
            
        except Exception as e:
            db.session.rollback()
            
            # Se for requisição AJAX, retornar JSON com erro
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': False,
                    'message': f'Erro ao cadastrar multa: {str(e)}'
                })
            
            flash(f'Erro ao cadastrar multa: {str(e)}', 'danger')
    else:
        # Se houver erros de validação e for AJAX, retornar os erros
        if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': 'Dados inválidos.',
                'errors': form.errors
            })
    
    # Se for requisição para modal, retornar apenas o formulário
    if request.args.get('modal') == '1':
        return render_template('modal_form_servico_veiculo.html', 
                             form=form, 
                             veiculo=veiculo, 
                             tipo_servico='multa',
                             acao='nova')
    
    return render_template('modal_servico_veiculo.html', 
                         form=form, 
                         veiculo=veiculo, 
                         tipo_servico='multa',
                         acao='nova')

@app.route('/cadastros/veiculos/<int:id_veiculo>/multas/<int:id_multa>/editar', methods=['GET', 'POST'])
def editar_multa_veiculo(id_veiculo, id_multa):
    """Edita uma multa existente"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    multa = MultaVeiculo.query.filter_by(id_multa=id_multa, id_veiculo=id_veiculo).first_or_404()
    form = MultaVeiculoForm(obj=multa)
    
    if form.validate_on_submit():
        try:
            form.populate_obj(multa)
            multa.valor_original = float(form.valor_original.data)
            multa.valor_pago = float(form.valor_pago.data) if form.valor_pago.data else None
            
            db.session.commit()
            
            # Registrar log
            registrar_log(
                acao="Editar Multa Veículo",
                descricao=f"Multa editada para o veículo {veiculo.nome} - {form.tipo_infracao.data}"
            )
            
            flash('Multa atualizada com sucesso!', 'success')
            return redirect(url_for('listar_multas_veiculo', id_veiculo=id_veiculo))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar multa: {str(e)}', 'danger')
    
    # Se for requisição para modal, retornar apenas o formulário
    if request.args.get('modal') == '1':
        return render_template('modal_form_servico_veiculo.html', 
                             form=form, 
                             veiculo=veiculo, 
                             servico=multa,
                             tipo_servico='multa',
                             acao='editar')
    
    return render_template('modal_servico_veiculo.html', 
                         form=form, 
                         veiculo=veiculo, 
                         servico=multa,
                         tipo_servico='multa',
                         acao='editar')

@app.route('/cadastros/veiculos/<int:id_veiculo>/multas/<int:id_multa>/excluir', methods=['POST'])
def excluir_multa_veiculo(id_veiculo, id_multa):
    """Exclui uma multa"""
    if 'user_id' not in session:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Sessão expirada'})
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    multa = MultaVeiculo.query.filter_by(id_multa=id_multa, id_veiculo=id_veiculo).first_or_404()
    
    try:
        tipo_infracao = multa.tipo_infracao
        db.session.delete(multa)
        db.session.commit()
        
        # Registrar log
        registrar_log(
            acao="Excluir Multa Veículo",
            descricao=f"Multa excluída do veículo {veiculo.nome} - {tipo_infracao}"
        )
        
        # Se for requisição AJAX, retornar JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': 'Multa excluída com sucesso!'
            })
        
        flash('Multa excluída com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        
        # Se for requisição AJAX, retornar JSON com erro
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': f'Erro ao excluir multa: {str(e)}'
            })
        
        flash(f'Erro ao excluir multa: {str(e)}', 'danger')
    
    return redirect(url_for('listar_multas_veiculo', id_veiculo=id_veiculo))

# ==================== IPVA ====================

@app.route('/cadastros/veiculos/<int:id_veiculo>/ipva')
def listar_ipva_veiculo(id_veiculo):
    """Lista todos os IPVAs de um veículo específico"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    ipvas = IpvaVeiculo.query.filter_by(id_veiculo=id_veiculo).order_by(IpvaVeiculo.ano_exercicio.desc()).all()
    
    # Se for requisição para modal, retornar apenas o conteúdo
    if request.args.get('modal') == '1':
        return render_template('modal_content_servicos.html', 
                             veiculo=veiculo, 
                             servicos=ipvas, 
                             tipo_servico='ipva',
                             titulo='IPVA')
    
    return render_template('servicos_veiculo.html', 
                         veiculo=veiculo, 
                         servicos=ipvas, 
                         tipo_servico='ipva',
                         titulo='IPVA')

@app.route('/cadastros/veiculos/<int:id_veiculo>/ipva/novo', methods=['GET', 'POST'])
def novo_ipva_veiculo(id_veiculo):
    """Cadastra novo IPVA para o veículo"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    form = IpvaVeiculoForm()
    
    if form.validate_on_submit():
        try:
            novo_ipva = IpvaVeiculo(
                id_veiculo=id_veiculo,
                ano_exercicio=form.ano_exercicio.data,
                data_vencimento=form.data_vencimento.data,
                data_pagamento=form.data_pagamento.data,
                valor_ipva=float(form.valor_ipva.data),
                valor_taxa_detran=float(form.valor_taxa_detran.data) if form.valor_taxa_detran.data else None,
                valor_multa_juros=float(form.valor_multa_juros.data) if form.valor_multa_juros.data else None,
                valor_total=float(form.valor_total.data),
                valor_pago=float(form.valor_pago.data) if form.valor_pago.data else None,
                numero_documento=form.numero_documento.data,
                status=form.status.data,
                observacoes=form.observacoes.data
            )
            
            db.session.add(novo_ipva)
            db.session.commit()
            
            # Registrar log
            registrar_log(
                acao="Cadastrar IPVA Veículo",
                descricao=f"Novo IPVA cadastrado para o veículo {veiculo.nome} - Ano {form.ano_exercicio.data}"
            )
            
            # Se for requisição AJAX, retornar JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': True,
                    'message': 'IPVA cadastrado com sucesso!'
                })
            
            flash('IPVA cadastrado com sucesso!', 'success')
            return redirect(url_for('listar_ipva_veiculo', id_veiculo=id_veiculo))
            
        except Exception as e:
            db.session.rollback()
            
            # Se for requisição AJAX, retornar JSON com erro
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': False,
                    'message': f'Erro ao cadastrar IPVA: {str(e)}'
                })
            
            flash(f'Erro ao cadastrar IPVA: {str(e)}', 'danger')
    else:
        # Se houver erros de validação e for AJAX, retornar os erros
        if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': 'Dados inválidos.',
                'errors': form.errors
            })
    
    # Se for requisição para modal, retornar apenas o formulário
    if request.args.get('modal') == '1':
        return render_template('modal_form_servico_veiculo.html', 
                             form=form, 
                             veiculo=veiculo, 
                             tipo_servico='ipva',
                             acao='novo')
    
    return render_template('modal_servico_veiculo.html', 
                         form=form, 
                         veiculo=veiculo, 
                         tipo_servico='ipva',
                         acao='novo')

@app.route('/cadastros/veiculos/<int:id_veiculo>/ipva/<int:id_ipva>/editar', methods=['GET', 'POST'])
def editar_ipva_veiculo(id_veiculo, id_ipva):
    """Edita um IPVA existente"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    ipva = IpvaVeiculo.query.filter_by(id_ipva=id_ipva, id_veiculo=id_veiculo).first_or_404()
    form = IpvaVeiculoForm(obj=ipva)
    
    if form.validate_on_submit():
        try:
            form.populate_obj(ipva)
            ipva.valor_ipva = float(form.valor_ipva.data)
            ipva.valor_taxa_detran = float(form.valor_taxa_detran.data) if form.valor_taxa_detran.data else None
            ipva.valor_multa_juros = float(form.valor_multa_juros.data) if form.valor_multa_juros.data else None
            ipva.valor_total = float(form.valor_total.data)
            ipva.valor_pago = float(form.valor_pago.data) if form.valor_pago.data else None
            
            db.session.commit()
            
            # Registrar log
            registrar_log(
                acao="Editar IPVA Veículo",
                descricao=f"IPVA editado para o veículo {veiculo.nome} - Ano {form.ano_exercicio.data}"
            )
            
            flash('IPVA atualizado com sucesso!', 'success')
            return redirect(url_for('listar_ipva_veiculo', id_veiculo=id_veiculo))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar IPVA: {str(e)}', 'danger')
    
    # Se for requisição para modal, retornar apenas o formulário
    if request.args.get('modal') == '1':
        return render_template('modal_form_servico_veiculo.html', 
                             form=form, 
                             veiculo=veiculo, 
                             servico=ipva,
                             tipo_servico='ipva',
                             acao='editar')
    
    return render_template('modal_servico_veiculo.html', 
                         form=form, 
                         veiculo=veiculo, 
                         servico=ipva,
                         tipo_servico='ipva',
                         acao='editar')

@app.route('/cadastros/veiculos/<int:id_veiculo>/ipva/<int:id_ipva>/excluir')
def excluir_ipva_veiculo(id_veiculo, id_ipva):
    """Exclui um IPVA"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    ipva = IpvaVeiculo.query.filter_by(id_ipva=id_ipva, id_veiculo=id_veiculo).first_or_404()
    
    try:
        ano_exercicio = ipva.ano_exercicio
        db.session.delete(ipva)
        db.session.commit()
        
        # Registrar log
        registrar_log(
            acao="Excluir IPVA Veículo",
            descricao=f"IPVA excluído do veículo {veiculo.nome} - Ano {ano_exercicio}"
        )
        
        flash('IPVA excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir IPVA: {str(e)}', 'danger')
    
    return redirect(url_for('listar_ipva_veiculo', id_veiculo=id_veiculo))

# ==================== LICENCIAMENTO ====================

@app.route('/cadastros/veiculos/<int:id_veiculo>/licenciamento')
def listar_licenciamento_veiculo(id_veiculo):
    """Lista todos os licenciamentos de um veículo específico"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    licenciamentos = LicenciamentoVeiculo.query.filter_by(id_veiculo=id_veiculo).order_by(LicenciamentoVeiculo.ano_exercicio.desc()).all()
    
    # Se for requisição para modal, retornar apenas o conteúdo
    if request.args.get('modal') == '1':
        return render_template('modal_content_servicos.html', 
                             veiculo=veiculo, 
                             servicos=licenciamentos, 
                             tipo_servico='licenciamento',
                             titulo='Licenciamento')
    
    return render_template('servicos_veiculo.html', 
                         veiculo=veiculo, 
                         servicos=licenciamentos, 
                         tipo_servico='licenciamento',
                         titulo='Licenciamento')

@app.route('/cadastros/veiculos/<int:id_veiculo>/licenciamento/novo', methods=['GET', 'POST'])
def novo_licenciamento_veiculo(id_veiculo):
    """Cadastra novo licenciamento para o veículo"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    form = LicenciamentoVeiculoForm()
    
    if form.validate_on_submit():
        try:
            novo_licenciamento = LicenciamentoVeiculo(
                id_veiculo=id_veiculo,
                ano_exercicio=form.ano_exercicio.data,
                data_vencimento=form.data_vencimento.data,
                data_pagamento=form.data_pagamento.data,
                valor_licenciamento=float(form.valor_licenciamento.data),
                valor_taxa_detran=float(form.valor_taxa_detran.data) if form.valor_taxa_detran.data else None,
                valor_multa_juros=float(form.valor_multa_juros.data) if form.valor_multa_juros.data else None,
                valor_total=float(form.valor_total.data),
                valor_pago=float(form.valor_pago.data) if form.valor_pago.data else None,
                numero_documento=form.numero_documento.data,
                status=form.status.data,
                observacoes=form.observacoes.data
            )
            
            db.session.add(novo_licenciamento)
            db.session.commit()
            
            # Registrar log
            registrar_log(
                acao="Cadastrar Licenciamento Veículo",
                descricao=f"Novo licenciamento cadastrado para o veículo {veiculo.nome} - Ano {form.ano_exercicio.data}"
            )
            
            flash('Licenciamento cadastrado com sucesso!', 'success')
            return redirect(url_for('listar_licenciamento_veiculo', id_veiculo=id_veiculo))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar licenciamento: {str(e)}', 'danger')
    
    # Se for requisição para modal, retornar apenas o formulário
    if request.args.get('modal') == '1':
        return render_template('modal_form_servico_veiculo.html', 
                             form=form, 
                             veiculo=veiculo, 
                             tipo_servico='licenciamento',
                             acao='novo')
    
    return render_template('modal_servico_veiculo.html', 
                         form=form, 
                         veiculo=veiculo, 
                         tipo_servico='licenciamento',
                         acao='novo')

@app.route('/cadastros/veiculos/<int:id_veiculo>/licenciamento/<int:id_licenciamento>/editar', methods=['GET', 'POST'])
def editar_licenciamento_veiculo(id_veiculo, id_licenciamento):
    """Edita um licenciamento existente"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    licenciamento = LicenciamentoVeiculo.query.filter_by(id_licenciamento=id_licenciamento, id_veiculo=id_veiculo).first_or_404()
    form = LicenciamentoVeiculoForm(obj=licenciamento)
    
    if form.validate_on_submit():
        try:
            form.populate_obj(licenciamento)
            licenciamento.valor_licenciamento = float(form.valor_licenciamento.data)
            licenciamento.valor_taxa_detran = float(form.valor_taxa_detran.data) if form.valor_taxa_detran.data else None
            licenciamento.valor_multa_juros = float(form.valor_multa_juros.data) if form.valor_multa_juros.data else None
            licenciamento.valor_total = float(form.valor_total.data)
            licenciamento.valor_pago = float(form.valor_pago.data) if form.valor_pago.data else None
            
            db.session.commit()
            
            # Registrar log
            registrar_log(
                acao="Editar Licenciamento Veículo",
                descricao=f"Licenciamento editado para o veículo {veiculo.nome} - Ano {form.ano_exercicio.data}"
            )
            
            flash('Licenciamento atualizado com sucesso!', 'success')
            return redirect(url_for('listar_licenciamento_veiculo', id_veiculo=id_veiculo))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar licenciamento: {str(e)}', 'danger')
    
    # Se for requisição para modal, retornar apenas o formulário
    if request.args.get('modal') == '1':
        return render_template('modal_form_servico_veiculo.html', 
                             form=form, 
                             veiculo=veiculo, 
                             servico=licenciamento,
                             tipo_servico='licenciamento',
                             acao='editar')
    
    return render_template('modal_servico_veiculo.html', 
                         form=form, 
                         veiculo=veiculo, 
                         servico=licenciamento,
                         tipo_servico='licenciamento',
                         acao='editar')

@app.route('/cadastros/veiculos/<int:id_veiculo>/licenciamento/<int:id_licenciamento>/excluir')
def excluir_licenciamento_veiculo(id_veiculo, id_licenciamento):
    """Exclui um licenciamento"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    licenciamento = LicenciamentoVeiculo.query.filter_by(id_licenciamento=id_licenciamento, id_veiculo=id_veiculo).first_or_404()
    
    try:
        ano_exercicio = licenciamento.ano_exercicio
        db.session.delete(licenciamento)
        db.session.commit()
        
        # Registrar log
        registrar_log(
            acao="Excluir Licenciamento Veículo",
            descricao=f"Licenciamento excluído do veículo {veiculo.nome} - Ano {ano_exercicio}"
        )
        
        flash('Licenciamento excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir licenciamento: {str(e)}', 'danger')
    
    return redirect(url_for('listar_licenciamento_veiculo', id_veiculo=id_veiculo))

# ==================== MANUTENÇÃO ====================

@app.route('/cadastros/veiculos/<int:id_veiculo>/manutencao')
def listar_manutencao_veiculo(id_veiculo):
    """Lista todas as manutenções de um veículo específico"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    manutencoes = ManutencaoVeiculo.query.filter_by(id_veiculo=id_veiculo).order_by(ManutencaoVeiculo.data_servico.desc()).all()
    
    # Se for requisição para modal, retornar apenas o conteúdo
    if request.args.get('modal') == '1':
        return render_template('modal_content_servicos.html', 
                             veiculo=veiculo, 
                             servicos=manutencoes, 
                             tipo_servico='manutencao',
                             titulo='Manutenção')
    
    return render_template('servicos_veiculo.html', 
                         veiculo=veiculo, 
                         servicos=manutencoes, 
                         tipo_servico='manutencao',
                         titulo='Manutenção')

@app.route('/cadastros/veiculos/<int:id_veiculo>/manutencao/nova', methods=['GET', 'POST'])
def nova_manutencao_veiculo(id_veiculo):
    """Cadastra nova manutenção para o veículo"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    form = ManutencaoVeiculoForm()
    
    if form.validate_on_submit():
        try:
            nova_manutencao = ManutencaoVeiculo(
                id_veiculo=id_veiculo,
                data_servico=form.data_servico.data,
                tipo_manutencao=form.tipo_manutencao.data,
                descricao=form.descricao.data,
                fornecedor_servico=form.fornecedor_servico.data,
                km_veiculo=form.km_veiculo.data,
                valor_servico=float(form.valor_servico.data),
                valor_pecas=float(form.valor_pecas.data) if form.valor_pecas.data else None,
                valor_total=float(form.valor_total.data),
                data_proxima_revisao=form.data_proxima_revisao.data,
                km_proxima_revisao=form.km_proxima_revisao.data,
                garantia_dias=form.garantia_dias.data,
                observacoes=form.observacoes.data
            )
            
            db.session.add(nova_manutencao)
            db.session.commit()
            
            # Registrar log
            registrar_log(
                acao="Cadastrar Manutenção Veículo",
                descricao=f"Nova manutenção cadastrada para o veículo {veiculo.nome} - {form.tipo_manutencao.data}"
            )
            
            flash('Manutenção cadastrada com sucesso!', 'success')
            return redirect(url_for('listar_manutencao_veiculo', id_veiculo=id_veiculo))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar manutenção: {str(e)}', 'danger')
    
    # Se for requisição para modal, retornar apenas o formulário
    if request.args.get('modal') == '1':
        return render_template('modal_form_servico_veiculo.html', 
                             form=form, 
                             veiculo=veiculo, 
                             tipo_servico='manutencao',
                             acao='nova')
    
    return render_template('modal_servico_veiculo.html', 
                         form=form, 
                         veiculo=veiculo, 
                         tipo_servico='manutencao',
                         acao='nova')

@app.route('/cadastros/veiculos/<int:id_veiculo>/manutencao/<int:id_manutencao>/editar', methods=['GET', 'POST'])
def editar_manutencao_veiculo(id_veiculo, id_manutencao):
    """Edita uma manutenção existente"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    manutencao = ManutencaoVeiculo.query.filter_by(id_manutencao=id_manutencao, id_veiculo=id_veiculo).first_or_404()
    form = ManutencaoVeiculoForm(obj=manutencao)
    
    if form.validate_on_submit():
        try:
            form.populate_obj(manutencao)
            manutencao.valor_servico = float(form.valor_servico.data)
            manutencao.valor_pecas = float(form.valor_pecas.data) if form.valor_pecas.data else None
            manutencao.valor_total = float(form.valor_total.data)
            
            db.session.commit()
            
            # Registrar log
            registrar_log(
                acao="Editar Manutenção Veículo",
                descricao=f"Manutenção editada para o veículo {veiculo.nome} - {form.tipo_manutencao.data}"
            )
            
            flash('Manutenção atualizada com sucesso!', 'success')
            return redirect(url_for('listar_manutencao_veiculo', id_veiculo=id_veiculo))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar manutenção: {str(e)}', 'danger')
    
    # Se for requisição para modal, retornar apenas o formulário
    if request.args.get('modal') == '1':
        return render_template('modal_form_servico_veiculo.html', 
                             form=form, 
                             veiculo=veiculo, 
                             servico=manutencao,
                             tipo_servico='manutencao',
                             acao='editar')
    
    return render_template('modal_servico_veiculo.html', 
                         form=form, 
                         veiculo=veiculo, 
                         servico=manutencao,
                         tipo_servico='manutencao',
                         acao='editar')

@app.route('/cadastros/veiculos/<int:id_veiculo>/manutencao/<int:id_manutencao>/excluir')
def excluir_manutencao_veiculo(id_veiculo, id_manutencao):
    """Exclui uma manutenção"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    veiculo = Veiculo.query.get_or_404(id_veiculo)
    manutencao = ManutencaoVeiculo.query.filter_by(id_manutencao=id_manutencao, id_veiculo=id_veiculo).first_or_404()
    
    try:
        tipo_manutencao = manutencao.tipo_manutencao
        db.session.delete(manutencao)
        db.session.commit()
        
        # Registrar log
        registrar_log(
            acao="Excluir Manutenção Veículo",
            descricao=f"Manutenção excluída do veículo {veiculo.nome} - {tipo_manutencao}"
        )
        
        flash('Manutenção excluída com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir manutenção: {str(e)}', 'danger')
    
    return redirect(url_for('listar_manutencao_veiculo', id_veiculo=id_veiculo))

@app.route('/cadastros/despesas', methods=['GET', 'POST'])
def cadastrar_despesa():
    # Verificar se existem categorias
    categorias_existentes = CategoriaDespesa.query.all()
    if not categorias_existentes:
        flash('É necessário cadastrar pelo menos uma categoria de despesa antes de cadastrar despesas.', 'warning')
        return redirect(url_for('cadastrar_categoria_despesa'))
    
    form = DespesaForm()
    form.id_categoria_despesa.choices = [(c.id_categoria_despesa, c.nome) for c in categorias_existentes]
    
    if form.validate_on_submit():
        # Converter valor médio para float
        valor_medio = None
        if form.valor_medio_despesa.data:
            try:
                valor_medio = float(form.valor_medio_despesa.data.replace(',', '.'))
            except ValueError:
                flash('Valor médio deve ser um número válido.', 'danger')
                despesas = Despesa.query.all()
                return render_template('despesas.html', form=form, despesas=despesas)
        
        nova = Despesa(
            nome=form.nome.data, 
            id_categoria_despesa=form.id_categoria_despesa.data,
            id_tipo_despesa=form.id_tipo_despesa.data,
            valor_medio_despesa=valor_medio,
            flag_alimentacao=form.flag_alimentacao.data,
            flag_combustivel=form.flag_combustivel.data
        )
        db.session.add(nova)
        db.session.commit()
        flash('Despesa cadastrada com sucesso!', 'success')
        return redirect(url_for('cadastrar_despesa'))
    despesas = Despesa.query.order_by(Despesa.nome).all()
    return render_template('despesas.html', form=form, despesas=despesas)

@app.route('/cadastros/despesas/editar/<int:id>', methods=['GET', 'POST'])
def editar_despesa(id):
    # Verificar se existem categorias
    categorias_existentes = CategoriaDespesa.query.all()
    if not categorias_existentes:
        flash('É necessário cadastrar pelo menos uma categoria de despesa antes de editar despesas.', 'warning')
        return redirect(url_for('cadastrar_categoria_despesa'))
    
    despesa = Despesa.query.get_or_404(id)
    form = DespesaForm(obj=despesa)
    form.id_categoria_despesa.choices = [(c.id_categoria_despesa, c.nome) for c in categorias_existentes]
    
    # Preencher o valor médio no formato brasileiro
    if request.method == 'GET' and despesa.valor_medio_despesa:
        form.valor_medio_despesa.data = str(despesa.valor_medio_despesa).replace('.', ',')
    
    if form.validate_on_submit():
        # Converter valor médio para float
        valor_medio = None
        if form.valor_medio_despesa.data:
            try:
                valor_medio = float(form.valor_medio_despesa.data.replace(',', '.'))
            except ValueError:
                flash('Valor médio deve ser um número válido.', 'danger')
                despesas = Despesa.query.all()
                return render_template('despesas.html', form=form, despesas=despesas)
        
        despesa.nome = form.nome.data
        despesa.id_categoria_despesa = form.id_categoria_despesa.data
        despesa.id_tipo_despesa = form.id_tipo_despesa.data
        despesa.valor_medio_despesa = valor_medio
        despesa.flag_alimentacao = form.flag_alimentacao.data
        despesa.flag_combustivel = form.flag_combustivel.data
        db.session.commit()
        flash('Despesa atualizada com sucesso!', 'success')
        return redirect(url_for('cadastrar_despesa'))
    despesas = Despesa.query.all()
    return render_template('despesas.html', form=form, despesas=despesas)

@app.route('/cadastros/despesas/excluir/<int:id>')
def excluir_despesa(id):
    despesa = Despesa.query.get_or_404(id)
    
    # Verificar se existem eventos usando esta despesa
    eventos_usando = DespesaEvento.query.filter_by(id_despesa=id).count()
    if eventos_usando > 0:
        flash(f'Não é possível excluir esta despesa pois existem {eventos_usando} registro(s) de despesa em evento(s) associado(s) a ela.', 'danger')
        return redirect(url_for('cadastrar_despesa'))
    
    db.session.delete(despesa)
    db.session.commit()
    flash('Despesa excluída com sucesso!', 'success')
    return redirect(url_for('cadastrar_despesa'))

@app.route('/eventos/<int:id_evento>/despesas', methods=['GET', 'POST'])
def cadastrar_despesa_evento(id_evento):
    evento = Evento.query.get_or_404(id_evento)
    
    # Verificar se existem categorias
    categorias_existentes = CategoriaDespesa.query.all()
    if not categorias_existentes:
        flash('É necessário cadastrar pelo menos uma categoria de despesa antes de cadastrar despesas.', 'warning')
        return redirect(url_for('cadastrar_categoria_despesa'))
    
    form = DespesaForm()
    form.id_categoria_despesa.choices = [(c.id_categoria_despesa, c.nome) for c in categorias_existentes]
    # Filtrar apenas tipos de evento (1 e 2)
    form.id_tipo_despesa.choices = [
        (1, 'Fixas - Evento'),
        (2, 'Variáveis - Evento')
    ]
    
    if form.validate_on_submit():
        # Converter valor médio para float
        valor_medio = None
        if form.valor_medio_despesa.data:
            try:
                valor_medio = float(form.valor_medio_despesa.data.replace(',', '.'))
            except ValueError:
                flash('Valor médio deve ser um número válido.', 'danger')
                # Buscar apenas despesas de evento
                despesas = Despesa.query.filter(Despesa.id_tipo_despesa.in_([1, 2])).all()
                return render_template('despesas_evento.html', form=form, despesas=despesas, evento=evento)
        
        nova = Despesa(
            nome=form.nome.data, 
            id_categoria_despesa=form.id_categoria_despesa.data,
            id_tipo_despesa=form.id_tipo_despesa.data,
            valor_medio_despesa=valor_medio,
            flag_alimentacao=form.flag_alimentacao.data,
            flag_combustivel=form.flag_combustivel.data
        )
        db.session.add(nova)
        db.session.commit()
        flash('Despesa cadastrada com sucesso!', 'success')
        return redirect(url_for('cadastrar_despesa_evento', id_evento=id_evento))
    
    # Buscar apenas despesas de evento
    despesas = Despesa.query.filter(Despesa.id_tipo_despesa.in_([1, 2])).all()
    return render_template('despesas_evento.html', form=form, despesas=despesas, evento=evento)

@app.route('/eventos/<int:id_evento>/salvar-despesa', methods=['POST'])
def salvar_despesa_individual(id_evento):
    """Salva ou atualiza uma despesa individual via AJAX com suporte a upload"""
    
    evento = Evento.query.get_or_404(id_evento)
    
    try:
        # Verificar se é JSON ou FormData
        if request.content_type and 'application/json' in request.content_type:
            # Processar dados JSON (vindo do modal rápido)
            data = request.get_json()
            despesa_id = data.get('despesa_id')
            valor_str = data.get('valor', '')
            data_despesa = data.get('data_vencimento')
            status_pagamento = data.get('status_pagamento', 'pendente')
            forma_pagamento = data.get('forma_pagamento', 'débito')
            pago_por = data.get('pago_por', '')
            observacoes = data.get('observacoes', '')
            id_fornecedor = data.get('id_fornecedor', None)
            despesa_cabeca = data.get('despesa_cabeca', False)
            qtd_dias = data.get('qtd_dias', None)
            qtd_pessoas = data.get('qtd_pessoas', None)
            valor_pago_socrates_str = data.get('valor_pago_socrates', '')
        else:
            # Processar dados do FormData (upload com arquivo)
            despesa_id = request.form.get('despesa_id')
            valor_str = request.form.get('valor', '')
            data_despesa = request.form.get('data_vencimento')
            status_pagamento = request.form.get('status_pagamento', 'pendente')
            forma_pagamento = request.form.get('forma_pagamento', 'débito')
            pago_por = request.form.get('pago_por', '')
            observacoes = request.form.get('observacoes', '')
            id_fornecedor = request.form.get('id_fornecedor', None)
            despesa_cabeca = request.form.get('despesa_cabeca') == '1'
            qtd_dias = request.form.get('qtd_dias', None)
            qtd_pessoas = request.form.get('qtd_pessoas', None)
            valor_pago_socrates_str = request.form.get('valor_pago_socrates', '')
        
        print(f"=== SALVAMENTO INDIVIDUAL ===")
        print(f"Tipo de dados: {'JSON' if request.content_type and 'application/json' in request.content_type else 'FormData'}")
        print(f"Despesa ID: {despesa_id}")
        print(f"Valor recebido: '{valor_str}'")
        print(f"Fornecedor ID: {id_fornecedor}")
        print(f"Despesa cabeça: {despesa_cabeca}")
        
        if not despesa_id or not valor_str:
            return jsonify({'success': False, 'message': 'Despesa e valor são obrigatórios'})
        
        # Converter valor - tratar formato brasileiro
        try:
            valor_str = str(valor_str).strip()
            
            # Se contém ponto e vírgula, é formato brasileiro (ex: 1.000,50)
            if '.' in valor_str and ',' in valor_str:
                # Remover pontos de milhares e trocar vírgula por ponto
                valor_str = valor_str.replace('.', '').replace(',', '.')
            # Se contém apenas vírgula, trocar por ponto
            elif ',' in valor_str and '.' not in valor_str:
                valor_str = valor_str.replace(',', '.')
            
            valor_float = float(valor_str)
            
            if valor_float <= 0:
                return jsonify({'success': False, 'message': 'Valor deve ser maior que zero'})
                
        except (ValueError, TypeError) as e:
            print(f"Erro ao converter valor '{valor_str}': {e}")
            return jsonify({'success': False, 'message': 'Valor inválido'})
        
        # Converter valor pago Sócrates Online
        valor_pago_socrates_float = None
        if valor_pago_socrates_str and valor_pago_socrates_str.strip():
            try:
                valor_pago_socrates_str = str(valor_pago_socrates_str).strip()
                
                # Aplicar mesma lógica de conversão do valor principal
                if '.' in valor_pago_socrates_str and ',' in valor_pago_socrates_str:
                    valor_pago_socrates_str = valor_pago_socrates_str.replace('.', '').replace(',', '.')
                elif ',' in valor_pago_socrates_str and '.' not in valor_pago_socrates_str:
                    valor_pago_socrates_str = valor_pago_socrates_str.replace(',', '.')
                
                valor_pago_socrates_float = float(valor_pago_socrates_str)
                
                if valor_pago_socrates_float < 0:
                    return jsonify({'success': False, 'message': 'Valor pago Sócrates Online não pode ser negativo'})
                    
            except (ValueError, TypeError) as e:
                print(f"Erro ao converter valor pago Sócrates Online '{valor_pago_socrates_str}': {e}")
                return jsonify({'success': False, 'message': 'Valor pago Sócrates Online inválido'})
        else:
            print(f"⚠️  Valor Pago Sócrates vazio ou nulo")
        
        # Converter data
        data_obj = datetime.strptime(data_despesa, '%Y-%m-%d').date() if data_despesa else date.today()
        fornecedor_id = int(id_fornecedor) if id_fornecedor and id_fornecedor != '0' else None
        
        # Converter qtd_dias e qtd_pessoas para inteiros quando válidos
        qtd_dias_int = None
        qtd_pessoas_int = None
        try:
            if qtd_dias and qtd_dias.strip():
                qtd_dias_int = int(qtd_dias)
        except (ValueError, AttributeError):
            pass
        
        try:
            if qtd_pessoas and qtd_pessoas.strip():
                qtd_pessoas_int = int(qtd_pessoas)
        except (ValueError, AttributeError):
            pass
        
        # Processar upload do comprovante
        comprovante_filename = None
        if 'comprovante' in request.files:
            file = request.files['comprovante']
            if file and file.filename != '':
                filename = secure_filename(file.filename)
                # Gerar nome único para o arquivo (mais simples)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                name_part, ext_part = os.path.splitext(filename)
                # Usar apenas os primeiros 8 caracteres do nome + timestamp + extensão
                unique_filename = f"{name_part[:8]}_{timestamp}{ext_part}"
                
                # Criar pasta se não existir
                upload_folder = os.path.join(app.config['UPLOAD_FOLDER'])
                os.makedirs(upload_folder, exist_ok=True)
                
                # Salvar arquivo
                file_path = os.path.join(upload_folder, unique_filename)
                file.save(file_path)
                comprovante_filename = unique_filename
                print(f"📎 Comprovante salvo: {comprovante_filename}")
        
        # Criar nova despesa no evento
        print(f"🆕 Criando nova despesa no evento")
        
        nova_despesa = DespesaEvento(
            id_evento=id_evento,
            id_despesa=int(despesa_id),
            data_vencimento=data_obj,
            valor=valor_float,
            valor_pago_socrates=valor_pago_socrates_float,
            status_pagamento=status_pagamento,
            forma_pagamento=forma_pagamento,
            pago_por=pago_por,
            observacoes=observacoes,
            id_fornecedor=fornecedor_id,
            comprovante=comprovante_filename,
            despesa_cabeca=despesa_cabeca,
            qtd_dias=qtd_dias_int,
            qtd_pessoas=qtd_pessoas_int
        )
        
        db.session.add(nova_despesa)
        
        # Se foi informado um fornecedor, adicionar automaticamente na tabela fornecedor_evento
        if fornecedor_id:
            fornecedor_evento_existente = FornecedorEvento.query.filter_by(
                id_evento=id_evento,
                id_fornecedor=fornecedor_id
            ).first()
            
            if not fornecedor_evento_existente:
                novo_fornecedor_evento = FornecedorEvento(
                    id_evento=id_evento,
                    id_fornecedor=fornecedor_id,
                    observacoes=observacoes if observacoes else ''
                )
                db.session.add(novo_fornecedor_evento)
                print(f"✅ Fornecedor {fornecedor_id} adicionado automaticamente ao evento")
        
        db.session.commit()
        
        print(f"✅ Nova despesa criada com sucesso! ID: {nova_despesa.id_despesa_evento}")
        
        # Buscar dados da despesa para retornar
        despesa = Despesa.query.get(despesa_id)
        fornecedor = Fornecedor.query.get(fornecedor_id) if fornecedor_id else None
        
        return jsonify({
            'success': True, 
            'message': 'Despesa salva com sucesso!',
            'despesa_evento_id': nova_despesa.id_despesa_evento,
            'despesa_nome': despesa.nome,
            'fornecedor_nome': fornecedor.nome if fornecedor else None,
            'valor_salvo': valor_float,
            'comprovante': comprovante_filename,
            'action': 'created'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ ERRO ao salvar despesa: {e}")
        return jsonify({'success': False, 'message': f'Erro ao salvar despesa: {str(e)}'})

@app.template_filter('date_br')
def date_br(value):
    if isinstance(value, str) and '-' in value:
        y, m, d = value.split('-')
        return f'{d}/{m}/{y}'
    if isinstance(value, (date, datetime)):
        return value.strftime('%d/%m/%Y')
    return value

# =============== ROTAS PARA EQUIPE DO EVENTO ===============
@app.route('/eventos/<int:id_evento>/equipe', methods=['GET', 'POST'])
def equipe_evento(id_evento):
    evento = Evento.query.get_or_404(id_evento)
    
    # Verificar se existem colaboradores
    colaboradores_existentes = Colaborador.query.all()
    if not colaboradores_existentes:
        flash('É necessário cadastrar pelo menos um colaborador antes de adicionar equipe ao evento.', 'warning')
        return redirect(url_for('cadastrar_colaborador'))
    
    form = EquipeEventoForm()
    form.id_colaborador.choices = [(0, 'Selecione um colaborador')] + [(c.id_colaborador, c.nome) for c in colaboradores_existentes]
    
    if form.validate_on_submit():
        # Verificar se o colaborador já está na equipe do evento
        equipe_existente = EquipeEvento.query.filter_by(
            id_evento=id_evento,
            id_colaborador=form.id_colaborador.data
        ).first()
        
        if equipe_existente:
            flash('Este colaborador já faz parte da equipe deste evento.', 'warning')
        else:
            nova_equipe = EquipeEvento(
                id_evento=id_evento,
                id_colaborador=form.id_colaborador.data,
                funcao=form.funcao.data,
                observacoes=form.observacoes.data
            )
            db.session.add(nova_equipe)
            db.session.commit()
            flash('Colaborador adicionado à equipe com sucesso!', 'success')
        return redirect(url_for('equipe_evento', id_evento=id_evento))
    
    equipe_evento = EquipeEvento.query.filter_by(id_evento=id_evento).all()
    return render_template('equipe_evento.html', form=form, equipe_evento=equipe_evento, evento=evento)

@app.route('/eventos/<int:id_evento>/equipe/editar/<int:id>', methods=['GET', 'POST'])
def editar_equipe_evento(id_evento, id):
    evento = Evento.query.get_or_404(id_evento)
    equipe = EquipeEvento.query.get_or_404(id)
    
    colaboradores_existentes = Colaborador.query.all()
    form = EquipeEventoForm(obj=equipe)
    form.id_colaborador.choices = [(c.id_colaborador, c.nome) for c in colaboradores_existentes]
    
    if form.validate_on_submit():
        # Verificar se outro colaborador já está na equipe (exceto o atual)
        equipe_existente = EquipeEvento.query.filter_by(
            id_evento=id_evento,
            id_colaborador=form.id_colaborador.data
        ).filter(EquipeEvento.id_equipe_evento != id).first()
        
        if equipe_existente:
            flash('Este colaborador já faz parte da equipe deste evento.', 'warning')
        else:
            equipe.id_colaborador = form.id_colaborador.data
            equipe.funcao = form.funcao.data
            equipe.observacoes = form.observacoes.data
            db.session.commit()
            flash('Equipe atualizada com sucesso!', 'success')
        return redirect(url_for('equipe_evento', id_evento=id_evento))
    
    equipe_evento = EquipeEvento.query.filter_by(id_evento=id_evento).all()
    return render_template('equipe_evento.html', form=form, equipe_evento=equipe_evento, evento=evento)

@app.route('/eventos/<int:id_evento>/equipe/excluir/<int:id>')
def excluir_equipe_evento(id_evento, id):
    equipe = EquipeEvento.query.get_or_404(id)
    db.session.delete(equipe)
    db.session.commit()
    flash('Colaborador removido da equipe com sucesso!', 'success')
    return redirect(url_for('equipe_evento', id_evento=id_evento))

# =============== FUNÇÕES DE VALIDAÇÃO PARA VEÍCULOS ===============

def verificar_conflito_veiculo(id_veiculo, data_inicio, hora_inicio, data_devolucao, hora_fim, id_exclusao=None):
    """
    Verifica se há conflito de uso do veículo no período informado
    """
    from datetime import datetime, time
    
    def obter_nome_motorista(uso):
        """Função auxiliar para obter nome do motorista de forma segura"""
        return uso.motorista.nome if uso.motorista else 'Motorista não identificado'
    
    # Verificar se dados obrigatórios estão presentes
    if not id_veiculo or not data_inicio:
        return None
    
    # Buscar usos existentes do veículo (excluindo o registro atual se for edição)
    query = VeiculoEvento.query.filter(VeiculoEvento.id_veiculo == id_veiculo)
    if id_exclusao:
        query = query.filter(VeiculoEvento.id_veiculo_evento != id_exclusao)
    
    usos_existentes = query.all()
    
    for uso in usos_existentes:
        # Se há uso em aberto (sem data_devolucao), há conflito
        if not uso.data_devolucao:
            return f"O veículo já está em uso por {obter_nome_motorista(uso)} desde {uso.data_inicio.strftime('%d/%m/%Y')} (uso em aberto)"
        
        # Se estamos criando um uso sem fim, verificar se há conflito com períodos fechados
        if not data_devolucao:
            # Novo uso sem fim: verificar se conflita com uso anterior
            if data_inicio < uso.data_devolucao:
                return f"O veículo foi usado por {obter_nome_motorista(uso)} até {uso.data_devolucao.strftime('%d/%m/%Y')}. Informe data posterior."
            elif data_inicio == uso.data_devolucao:
                # Mesmo dia: verificar horários se disponíveis
                if hora_inicio and uso.hora_fim:
                    if hora_inicio <= uso.hora_fim:
                        return f"O veículo foi usado por {obter_nome_motorista(uso)} até {uso.hora_fim.strftime('%H:%M')} em {uso.data_devolucao.strftime('%d/%m/%Y')}. Informe horário posterior."
                elif not uso.hora_fim:
                    # Se uso anterior não tem hora fim, não podemos usar no mesmo dia
                    return f"O veículo foi usado por {obter_nome_motorista(uso)} em {uso.data_devolucao.strftime('%d/%m/%Y')} sem horário de fim definido. Use data posterior."
        else:
            # Verificar sobreposição de períodos definidos com horários
            conflito_periodo = False
            
            # Se temos horários completos, fazer comparação precisa
            if hora_inicio and hora_fim and uso.hora_inicio and uso.hora_fim:
                from datetime import datetime, time
                
                # Criar datetime completos para comparação
                inicio_novo = datetime.combine(data_inicio, hora_inicio)
                fim_novo = datetime.combine(data_devolucao, hora_fim)
                inicio_existente = datetime.combine(uso.data_inicio, uso.hora_inicio)
                fim_existente = datetime.combine(uso.data_devolucao, uso.hora_fim)
                
                # Verificar sobreposição temporal exata
                conflito_periodo = not (fim_novo <= inicio_existente or inicio_novo >= fim_existente)
            else:
                # Sem horários completos, usar apenas datas (mais conservador)
                conflito_periodo = not (data_devolucao < uso.data_inicio or data_inicio > uso.data_devolucao)
            
            if conflito_periodo:
                if hora_inicio and uso.hora_inicio and data_inicio == uso.data_inicio and hora_inicio == uso.hora_inicio:
                    return f"Conflito: mesmo horário de início ({hora_inicio.strftime('%H:%M')}) que {obter_nome_motorista(uso)}"
                return f"Período conflita com uso por {obter_nome_motorista(uso)} de {uso.data_inicio.strftime('%d/%m/%Y')} a {uso.data_devolucao.strftime('%d/%m/%Y')}"
    
    return None

def verificar_conflito_motorista(id_motorista, data_inicio, hora_inicio, data_devolucao, hora_fim, id_exclusao=None):
    """
    Verifica se o motorista já está usando outro veículo no período
    """
    # Verificar se dados obrigatórios estão presentes
    if not id_motorista or not data_inicio:
        return None
    
    # Buscar usos existentes do motorista (excluindo o registro atual se for edição)
    query = VeiculoEvento.query.filter(VeiculoEvento.id_motorista == id_motorista)
    if id_exclusao:
        query = query.filter(VeiculoEvento.id_veiculo_evento != id_exclusao)
    
    usos_existentes = query.all()
    
    for uso in usos_existentes:
        # Se há uso em aberto, há conflito
        if not uso.data_devolucao:
            return f"Motorista já está usando {uso.veiculo.nome} desde {uso.data_inicio.strftime('%d/%m/%Y')} (uso em aberto)"
        
        # Se estamos criando um uso sem fim, verificar conflito
        if not data_devolucao:
            if data_inicio < uso.data_devolucao:
                return f"Motorista usou {uso.veiculo.nome} até {uso.data_devolucao.strftime('%d/%m/%Y')}. Informe data posterior."
            elif data_inicio == uso.data_devolucao:
                # Mesmo dia: verificar horários se disponíveis
                if hora_inicio and uso.hora_fim:
                    if hora_inicio <= uso.hora_fim:
                        return f"Motorista usou {uso.veiculo.nome} até {uso.hora_fim.strftime('%H:%M')} em {uso.data_devolucao.strftime('%d/%m/%Y')}. Informe horário posterior."
                elif not uso.hora_fim:
                    # Se uso anterior não tem hora fim, não podemos usar no mesmo dia
                    return f"Motorista usou {uso.veiculo.nome} em {uso.data_devolucao.strftime('%d/%m/%Y')} sem horário de fim definido. Use data posterior."
        else:
            # Verificar sobreposição de períodos definidos com horários
            conflito_periodo = False
            
            # Se temos horários completos, fazer comparação precisa
            if hora_inicio and hora_fim and uso.hora_inicio and uso.hora_fim:
                from datetime import datetime, time
                
                # Criar datetime completos para comparação
                inicio_novo = datetime.combine(data_inicio, hora_inicio)
                fim_novo = datetime.combine(data_devolucao, hora_fim)
                inicio_existente = datetime.combine(uso.data_inicio, uso.hora_inicio)
                fim_existente = datetime.combine(uso.data_devolucao, uso.hora_fim)
                
                # Verificar sobreposição temporal exata
                conflito_periodo = not (fim_novo <= inicio_existente or inicio_novo >= fim_existente)
            else:
                # Sem horários completos, usar apenas datas (mais conservador)
                conflito_periodo = not (data_devolucao < uso.data_inicio or data_inicio > uso.data_devolucao)
            
            if conflito_periodo:
                if hora_inicio and uso.hora_inicio and data_inicio == uso.data_inicio and hora_inicio == uso.hora_inicio:
                    return f"Conflito: motorista já usou {uso.veiculo.nome} no mesmo horário ({hora_inicio.strftime('%H:%M')})"
                return f"Motorista já usou {uso.veiculo.nome} de {uso.data_inicio.strftime('%d/%m/%Y')} a {uso.data_devolucao.strftime('%d/%m/%Y')}"
    
    return None

def verificar_km_sequencial(id_veiculo, km_inicio, id_exclusao=None):
    """
    Verifica se a quilometragem inicial é válida baseada no último uso do veículo
    """
    if not id_veiculo or not km_inicio:
        return None
    
    # Buscar último uso do veículo (excluindo o atual se for edição)
    query = VeiculoEvento.query.filter(
        VeiculoEvento.id_veiculo == id_veiculo,
        VeiculoEvento.km_fim.isnot(None)
    )
    if id_exclusao:
        query = query.filter(VeiculoEvento.id_veiculo_evento != id_exclusao)
    
    ultimo_uso = query.order_by(VeiculoEvento.data_inicio.desc()).first()
    
    if ultimo_uso and ultimo_uso.km_fim and km_inicio < ultimo_uso.km_fim:
        km_inicio_fmt = f"{km_inicio:,}".replace(',', '.')
        km_fim_fmt = f"{ultimo_uso.km_fim:,}".replace(',', '.')
        return f"KM inicial ({km_inicio_fmt}) deve ser maior ou igual ao KM final do último uso ({km_fim_fmt})"
    
    return None

def verificar_ultimo_uso_completo(id_veiculo, id_exclusao=None):
    """
    Verifica se o último uso do veículo está completo (data_devolucao, hora_fim, km_fim)
    """
    def obter_nome_motorista(uso):
        """Função auxiliar para obter nome do motorista de forma segura"""
        return uso.motorista.nome if uso.motorista else 'Motorista não identificado'
    
    # Verificar se dados obrigatórios estão presentes
    if not id_veiculo:
        return None
    
    # Buscar último uso do veículo (excluindo o atual se for edição)
    query = VeiculoEvento.query.filter(VeiculoEvento.id_veiculo == id_veiculo)
    if id_exclusao:
        query = query.filter(VeiculoEvento.id_veiculo_evento != id_exclusao)
    
    ultimo_uso = query.order_by(VeiculoEvento.data_inicio.desc()).first()
    
    if ultimo_uso:
        # Verificar se algum campo obrigatório de finalização está faltando
        if not ultimo_uso.data_devolucao:
            return f"O último uso do veículo por {obter_nome_motorista(ultimo_uso)} está em aberto (sem data de devolução)"
        if not ultimo_uso.hora_fim:
            return f"O último uso do veículo por {obter_nome_motorista(ultimo_uso)} está sem horário de fim"
        if not ultimo_uso.km_fim:
            return f"O último uso do veículo por {obter_nome_motorista(ultimo_uso)} está sem quilometragem final"
    
    return None

def buscar_usos_em_aberto():
    """
    Busca usos de veículos que estão em aberto (sem data de devolução)
    """
    return VeiculoEvento.query.filter(VeiculoEvento.data_devolucao.is_(None)).all()

def buscar_usos_em_aberto_para_evento(id_evento):
    """
    Busca apenas usos em aberto que realmente impedem novos cadastros.
    Múltiplos veículos podem estar em uso simultaneamente, desde que sejam veículos diferentes.
    """
    # Buscar todos os usos em aberto
    usos_em_aberto = VeiculoEvento.query.filter(VeiculoEvento.data_devolucao.is_(None)).all()
    
    # Retorna todos os usos em aberto para exibição informativa
    # A validação de conflitos será feita individualmente por veículo no momento do cadastro
    return usos_em_aberto

# =============== ROTAS PARA VEÍCULOS DO EVENTO ===============
@app.route('/eventos/<int:id_evento>/veiculos', methods=['GET', 'POST'])
def veiculos_evento(id_evento):
    evento = Evento.query.get_or_404(id_evento)
    
    # Verificar se existem veículos
    veiculos_existentes = Veiculo.query.all()
    if not veiculos_existentes:
        flash('É necessário cadastrar pelo menos um veículo antes de adicionar ao evento.', 'warning')
        return redirect(url_for('cadastrar_veiculo'))
    
    # Verificar se existem colaboradores
    colaboradores_existentes = Colaborador.query.all()
    if not colaboradores_existentes:
        flash('É necessário cadastrar pelo menos um colaborador antes de adicionar veículos ao evento.', 'warning')
        return redirect(url_for('cadastrar_colaborador'))
    
    form = VeiculoEventoForm()
    # Repovoar choices sempre
    form.id_veiculo.choices = [(0, 'Selecione um veículo')] + [(v.id_veiculo, f"{v.nome} - {v.placa or 'Sem placa'}") for v in veiculos_existentes]
    
    # Buscar apenas colaboradores que estão na equipe do evento
    equipe_colaboradores = db.session.query(Colaborador).join(EquipeEvento).filter(EquipeEvento.id_evento == id_evento).all()
    
    if not equipe_colaboradores:
        flash('É necessário adicionar colaboradores à equipe do evento antes de cadastrar veículos.', 'warning')
        return redirect(url_for('equipe_evento', id_evento=id_evento))
    
    form.id_motorista.choices = [(0, 'Selecione um motorista')] + [(c.id_colaborador, c.nome) for c in equipe_colaboradores]
    
    if form.validate_on_submit():
        # Validações obrigatórias primeiro
        if form.id_veiculo.data == 0:
            flash('Por favor, selecione um veículo.', 'danger')
        elif form.id_motorista.data == 0:
            flash('Por favor, selecione um motorista.', 'danger')
        else:
            # Aplicar validações robustas
            erro_veiculo = verificar_conflito_veiculo(
                form.id_veiculo.data, 
                form.data_inicio.data, 
                form.hora_inicio.data,
                form.data_devolucao.data, 
                form.hora_fim.data
            )
            
            erro_motorista = verificar_conflito_motorista(
                form.id_motorista.data,
                form.data_inicio.data,
                form.hora_inicio.data, 
                form.data_devolucao.data,
                form.hora_fim.data
            )
            
            erro_km = verificar_km_sequencial(
                form.id_veiculo.data,
                form.km_inicio.data
            )
            
            erro_uso_anterior = verificar_ultimo_uso_completo(
                form.id_veiculo.data
            )
            
            if erro_veiculo:
                flash(erro_veiculo, 'warning')
            elif erro_motorista:
                flash(erro_motorista, 'warning')
            elif erro_km:
                flash(erro_km, 'warning')
            elif erro_uso_anterior:
                flash(erro_uso_anterior, 'warning')
            else:
                    novo_veiculo = VeiculoEvento(
                        id_evento=id_evento,
                        id_veiculo=form.id_veiculo.data,
                        id_motorista=form.id_motorista.data,
                        data_inicio=form.data_inicio.data,
                        data_devolucao=form.data_devolucao.data,
                        hora_inicio=form.hora_inicio.data,
                        hora_fim=form.hora_fim.data,
                        km_inicio=form.km_inicio.data,
                        km_fim=form.km_fim.data,
                        observacoes=form.observacoes.data
                    )
                    db.session.add(novo_veiculo)
                    db.session.commit()
                    flash('Veículo adicionado ao evento com sucesso!', 'success')
                    return redirect(url_for('veiculos_evento', id_evento=id_evento))
    
    veiculos_evento = VeiculoEvento.query.filter_by(id_evento=id_evento).all()
    usos_em_aberto = buscar_usos_em_aberto_para_evento(id_evento)
    
    # Filtrar usos em aberto apenas deste evento
    usos_em_aberto_evento = [uso for uso in usos_em_aberto if uso.id_evento == id_evento]
    
    return render_template('veiculos_evento.html', form=form, veiculos_evento=veiculos_evento, evento=evento, usos_em_aberto=usos_em_aberto, usos_em_aberto_evento=usos_em_aberto_evento)

@app.route('/eventos/<int:id_evento>/veiculos/editar/<int:id>', methods=['GET', 'POST'])
def editar_veiculo_evento(id_evento, id):
    evento = Evento.query.get_or_404(id_evento)
    veiculo_evento = VeiculoEvento.query.get_or_404(id)
    
    veiculos_existentes = Veiculo.query.all()
    
    # Buscar apenas colaboradores que estão na equipe do evento
    equipe_colaboradores = db.session.query(Colaborador).join(EquipeEvento).filter(EquipeEvento.id_evento == id_evento).all()
    
    form = VeiculoEventoForm(obj=veiculo_evento)
    form.id_veiculo.choices = [(v.id_veiculo, f"{v.nome} - {v.placa or 'Sem placa'}") for v in veiculos_existentes]
    form.id_motorista.choices = [(c.id_colaborador, c.nome) for c in equipe_colaboradores]
    
    if form.validate_on_submit():
        # Validações obrigatórias primeiro  
        if form.id_veiculo.data == 0:
            flash('Por favor, selecione um veículo.', 'danger')
        elif form.id_motorista.data == 0:
            flash('Por favor, selecione um motorista.', 'danger')
        else:
            # Aplicar validações robustas (excluindo o registro atual)
            erro_veiculo = verificar_conflito_veiculo(
                form.id_veiculo.data, 
                form.data_inicio.data, 
                form.hora_inicio.data,
                form.data_devolucao.data, 
                form.hora_fim.data,
                id_exclusao=id
            )
            
            erro_motorista = verificar_conflito_motorista(
                form.id_motorista.data,
                form.data_inicio.data,
                form.hora_inicio.data, 
                form.data_devolucao.data,
                form.hora_fim.data,
                id_exclusao=id
            )
            
            erro_km = verificar_km_sequencial(
                form.id_veiculo.data,
                form.km_inicio.data,
                id_exclusao=id
            )
            
            erro_uso_anterior = verificar_ultimo_uso_completo(
                form.id_veiculo.data,
                id_exclusao=id
            )
            
            if erro_veiculo:
                flash(erro_veiculo, 'warning')
            elif erro_motorista:
                flash(erro_motorista, 'warning')
            elif erro_km:
                flash(erro_km, 'warning')
            elif erro_uso_anterior:
                flash(erro_uso_anterior, 'warning')
            else:
                veiculo_evento.id_veiculo = form.id_veiculo.data
                veiculo_evento.id_motorista = form.id_motorista.data
                veiculo_evento.data_inicio = form.data_inicio.data
                veiculo_evento.data_devolucao = form.data_devolucao.data
                veiculo_evento.hora_inicio = form.hora_inicio.data
                veiculo_evento.hora_fim = form.hora_fim.data
                veiculo_evento.km_inicio = form.km_inicio.data
                veiculo_evento.km_fim = form.km_fim.data
                veiculo_evento.observacoes = form.observacoes.data
                db.session.commit()
                flash('Veículo do evento atualizado com sucesso!', 'success')
                return redirect(url_for('veiculos_evento', id_evento=id_evento))
    
    # Sempre repovoar choices (mesmo em caso de erro de validação)
    form.id_veiculo.choices = [(v.id_veiculo, f"{v.nome} - {v.placa or 'Sem placa'}") for v in veiculos_existentes]
    form.id_motorista.choices = [(c.id_colaborador, c.nome) for c in equipe_colaboradores]
    
    veiculos_evento = VeiculoEvento.query.filter_by(id_evento=id_evento).all()
    usos_em_aberto = buscar_usos_em_aberto_para_evento(id_evento)
    
    # Filtrar usos em aberto apenas deste evento
    usos_em_aberto_evento = [uso for uso in usos_em_aberto if uso.id_evento == id_evento]
    
    return render_template('veiculos_evento.html', form=form, veiculos_evento=veiculos_evento, evento=evento, editando=True, veiculo_editando=veiculo_evento, usos_em_aberto=usos_em_aberto, usos_em_aberto_evento=usos_em_aberto_evento)

@app.route('/eventos/<int:id_evento>/veiculos/excluir/<int:id>')
def excluir_veiculo_evento(id_evento, id):
    veiculo_evento = VeiculoEvento.query.get_or_404(id)
    db.session.delete(veiculo_evento)
    db.session.commit()
    flash('Veículo removido do evento com sucesso!', 'success')
    return redirect(url_for('veiculos_evento', id_evento=id_evento))

# =============== ROTAS PARA ELENCO DO EVENTO ===============
@app.route('/eventos/<int:id_evento>/elenco', methods=['GET', 'POST'])
def elenco_evento(id_evento):
    evento = Evento.query.get_or_404(id_evento)
    
    # Verificar se existem membros do elenco
    elencos_existentes = Elenco.query.all()
    if not elencos_existentes:
        flash('É necessário cadastrar pelo menos um membro do elenco antes de adicionar ao evento.', 'warning')
        return redirect(url_for('cadastrar_elenco'))
    
    form = ElencoEventoForm()
    form.id_elenco.choices = [(0, 'Selecione um membro do elenco')] + [(e.id_elenco, e.nome) for e in elencos_existentes]
    
    if form.validate_on_submit():
        # Verificar se o elenco já está no evento
        elenco_existente = ElencoEvento.query.filter_by(
            id_evento=id_evento,
            id_elenco=form.id_elenco.data
        ).first()
        
        if elenco_existente:
            flash('Este membro do elenco já faz parte deste evento.', 'warning')
        else:
            novo_elenco = ElencoEvento(
                id_evento=id_evento,
                id_elenco=form.id_elenco.data,
                observacoes=form.observacoes.data
            )
            db.session.add(novo_elenco)
            db.session.commit()
            flash('Membro do elenco adicionado ao evento com sucesso!', 'success')
        return redirect(url_for('elenco_evento', id_evento=id_evento))
    
    elenco_evento = ElencoEvento.query.filter_by(id_evento=id_evento).all()
    return render_template('elenco_evento.html', form=form, elenco_evento=elenco_evento, evento=evento)

@app.route('/eventos/<int:id_evento>/elenco/editar/<int:id>', methods=['GET', 'POST'])
def editar_elenco_evento(id_evento, id):
    evento = Evento.query.get_or_404(id_evento)
    elenco_evt = ElencoEvento.query.get_or_404(id)
    
    elencos_existentes = Elenco.query.all()
    form = ElencoEventoForm(obj=elenco_evt)
    form.id_elenco.choices = [(e.id_elenco, e.nome) for e in elencos_existentes]
    
    if form.validate_on_submit():
        # Verificar se outro elenco já está no evento (exceto o atual)
        elenco_existente = ElencoEvento.query.filter_by(
            id_evento=id_evento,
            id_elenco=form.id_elenco.data
        ).filter(ElencoEvento.id_elenco_evento != id).first()
        
        if elenco_existente:
            flash('Este membro do elenco já faz parte deste evento.', 'warning')
        else:
            elenco_evt.id_elenco = form.id_elenco.data
            elenco_evt.observacoes = form.observacoes.data
            db.session.commit()
            flash('Elenco do evento atualizado com sucesso!', 'success')
        return redirect(url_for('elenco_evento', id_evento=id_evento))
    
    elenco_evento = ElencoEvento.query.filter_by(id_evento=id_evento).all()
    return render_template('elenco_evento.html', form=form, elenco_evento=elenco_evento, evento=evento)

@app.route('/eventos/<int:id_evento>/elenco/excluir/<int:id>')
def excluir_elenco_evento(id_evento, id):
    elenco_evt = ElencoEvento.query.get_or_404(id)
    db.session.delete(elenco_evt)
    db.session.commit()
    flash('Membro do elenco removido do evento com sucesso!', 'success')
    return redirect(url_for('elenco_evento', id_evento=id_evento))

# =============== ROTAS PARA FORNECEDORES DO EVENTO ===============
@app.route('/eventos/<int:id_evento>/fornecedores', methods=['GET', 'POST'])
def fornecedor_evento(id_evento):
    evento = Evento.query.get_or_404(id_evento)
    
    # Verificar se existem fornecedores
    fornecedores_existentes = Fornecedor.query.all()
    if not fornecedores_existentes:
        flash('É necessário cadastrar pelo menos um fornecedor antes de adicionar ao evento.', 'warning')
        return redirect(url_for('cadastrar_fornecedor'))
    
    form = FornecedorEventoForm()
    form.id_fornecedor.choices = [(0, 'Selecione um fornecedor')] + [(f.id_fornecedor, f.nome) for f in fornecedores_existentes]
    
    if form.validate_on_submit():
        # Verificar se o fornecedor já está no evento
        fornecedor_existente = FornecedorEvento.query.filter_by(
            id_evento=id_evento,
            id_fornecedor=form.id_fornecedor.data
        ).first()
        
        if fornecedor_existente:
            flash('Este fornecedor já faz parte deste evento.', 'warning')
        else:
            novo_fornecedor = FornecedorEvento(
                id_evento=id_evento,
                id_fornecedor=form.id_fornecedor.data,
                observacoes=form.observacoes.data
            )
            db.session.add(novo_fornecedor)
            db.session.commit()
            flash('Fornecedor adicionado ao evento com sucesso!', 'success')
        return redirect(url_for('fornecedor_evento', id_evento=id_evento))
    
    fornecedor_evento = FornecedorEvento.query.filter_by(id_evento=id_evento).all()
    return render_template('fornecedor_evento.html', form=form, fornecedor_evento=fornecedor_evento, evento=evento)

@app.route('/eventos/<int:id_evento>/fornecedores/editar/<int:id>', methods=['GET', 'POST'])
def editar_fornecedor_evento(id_evento, id):
    evento = Evento.query.get_or_404(id_evento)
    fornecedor_evt = FornecedorEvento.query.get_or_404(id)
    
    fornecedores_existentes = Fornecedor.query.all()
    form = FornecedorEventoForm(obj=fornecedor_evt)
    form.id_fornecedor.choices = [(f.id_fornecedor, f.nome) for f in fornecedores_existentes]
    
    if form.validate_on_submit():
        # Verificar se outro fornecedor já está no evento (exceto o atual)
        fornecedor_existente = FornecedorEvento.query.filter_by(
            id_evento=id_evento,
            id_fornecedor=form.id_fornecedor.data
        ).filter(FornecedorEvento.id_fornecedor_evento != id).first()
        
        if fornecedor_existente:
            flash('Este fornecedor já faz parte deste evento.', 'warning')
        else:
            fornecedor_evt.id_fornecedor = form.id_fornecedor.data
            fornecedor_evt.observacoes = form.observacoes.data
            db.session.commit()
            flash('Fornecedor do evento atualizado com sucesso!', 'success')
        return redirect(url_for('fornecedor_evento', id_evento=id_evento))
    
    fornecedor_evento = FornecedorEvento.query.filter_by(id_evento=id_evento).all()
    return render_template('fornecedor_evento.html', form=form, fornecedor_evento=fornecedor_evento, evento=evento)

@app.route('/eventos/<int:id_evento>/fornecedores/excluir/<int:id>')
def excluir_fornecedor_evento(id_evento, id):
    fornecedor_evt = FornecedorEvento.query.get_or_404(id)
    db.session.delete(fornecedor_evt)
    db.session.commit()
    flash('Fornecedor removido do evento com sucesso!', 'success')
    return redirect(url_for('fornecedor_evento', id_evento=id_evento))

@app.route('/eventos/<int:id_evento>/excluir-receita/<int:receita_evento_id>', methods=['DELETE'])
def excluir_receita_evento(id_evento, receita_evento_id):
    try:
        # Verificar se o evento existe
        evento = Evento.query.get(id_evento)
        if not evento:
            return jsonify({'success': False, 'message': 'Evento não encontrado'})
        
        # Buscar a receita do evento
        receita_evento = ReceitaEvento.query.filter_by(
            id_receita_evento=receita_evento_id,
            id_evento=id_evento
        ).first()
        
        if not receita_evento:
            return jsonify({'success': False, 'message': 'Receita não encontrada neste evento'})
        
        # Excluir a receita do evento
        db.session.delete(receita_evento)
        db.session.commit()
        
        print(f"âœ… Receita excluída do evento: ID {receita_evento_id}")
        
        return jsonify({
            'success': True, 
            'message': 'Receita excluída com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"âŒ Erro ao excluir receita: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'})

@app.route('/eventos/<int:id_evento>/atualizar-receita/<int:receita_evento_id>', methods=['PUT'])
def atualizar_receita_evento(id_evento, receita_evento_id):
    try:
        # Verificar se o evento existe
        evento = Evento.query.get(id_evento)
        if not evento:
            return jsonify({'success': False, 'message': 'Evento não encontrado'})
        
        # Buscar a receita do evento
        receita_evento = ReceitaEvento.query.filter_by(
            id_receita_evento=receita_evento_id,
            id_evento=id_evento
        ).first()
        
        if not receita_evento:
            return jsonify({'success': False, 'message': 'Receita não encontrada neste evento'})
        
        data = request.get_json()
        
        # Validar dados obrigatórios
        if not data.get('valor') or not data.get('data'):
            return jsonify({'success': False, 'message': 'Data e valor são obrigatórios'})
        
        # Converter valor - tratar tanto formato brasileiro quanto americano
        try:
            valor_str = str(data['valor']).strip()
            
            # Se contém ponto e vírgula, é formato brasileiro (ex: 1.000,50)
            if '.' in valor_str and ',' in valor_str:
                # Remover pontos de milhares e trocar vírgula por ponto
                valor_str = valor_str.replace('.', '').replace(',', '.')
            # Se contém apenas vírgula, trocar por ponto
            elif ',' in valor_str and '.' not in valor_str:
                valor_str = valor_str.replace(',', '.')
            
            valor = float(valor_str)
            
            if valor <= 0:
                return jsonify({'success': False, 'message': 'Valor deve ser maior que zero'})
                
        except (ValueError, TypeError) as e:
            print(f"Erro ao converter valor '{data['valor']}': {e}")
            return jsonify({'success': False, 'message': 'Valor inválido'})
        
        # Converter data
        try:
            data_receita = datetime.strptime(data['data'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': 'Data inválida'})
        
        # Atualizar receita do evento
        receita_evento.valor = valor
        receita_evento.data = data_receita
        receita_evento.observacoes = data.get('observacoes', '')
        
        db.session.commit()
        
        print(f"âœ… Receita atualizada no evento: ID {receita_evento_id}")
        
        return jsonify({
            'success': True, 
            'message': 'Receita atualizada com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"âŒ Erro ao atualizar receita: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'})

@app.route('/eventos/<int:id_evento>/excluir-despesa/<int:despesa_evento_id>', methods=['DELETE'])
def excluir_despesa_evento(id_evento, despesa_evento_id):
    try:
        # Verificar se o evento existe
        evento = Evento.query.get(id_evento)
        if not evento:
            return jsonify({'success': False, 'message': 'Evento não encontrado'})
        
        # Buscar a despesa do evento
        despesa_evento = DespesaEvento.query.filter_by(
            id_despesa_evento=despesa_evento_id,
            id_evento=id_evento
        ).first()
        
        if not despesa_evento:
            return jsonify({'success': False, 'message': 'Despesa não encontrada neste evento'})
        
        # Excluir a despesa do evento
        db.session.delete(despesa_evento)
        db.session.commit()
        
        print(f"âœ… Despesa excluída do evento: ID {despesa_evento_id}")
        
        return jsonify({
            'success': True, 
            'message': 'Despesa excluída com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"âŒ Erro ao excluir despesa: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'})

@app.route('/eventos/<int:id_evento>/editar-despesa/<int:despesa_evento_id>', methods=['PUT'])
def editar_despesa_evento(id_evento, despesa_evento_id):
    """Edita uma despesa específica do evento"""
    try:
        # Verificar se o evento existe
        evento = Evento.query.get(id_evento)
        if not evento:
            return jsonify({'success': False, 'message': 'Evento não encontrado'})
        
        # Buscar a despesa do evento
        despesa_evento = DespesaEvento.query.filter_by(
            id_despesa_evento=despesa_evento_id,
            id_evento=id_evento
        ).first()
        
        if not despesa_evento:
            return jsonify({'success': False, 'message': 'Despesa não encontrada neste evento'})
        
        # Obter dados do FormData
        despesa_id = request.form.get('despesa_id')
        valor_str = request.form.get('valor', '')
        valor_pago_socrates_str = request.form.get('valor_pago_socrates', '')
        data_despesa = request.form.get('data_vencimento')
        
        # Validar dados obrigatórios
        if not despesa_id or not valor_str:
            return jsonify({'success': False, 'message': 'Despesa e valor são obrigatórios'})
        
        # Converter valor
        try:
            valor_str = str(valor_str).strip()
            
            # Se contém ponto e vírgula, é formato brasileiro (ex: 1.000,50)
            if '.' in valor_str and ',' in valor_str:
                valor_str = valor_str.replace('.', '').replace(',', '.')
            # Se contém apenas vírgula, trocar por ponto
            elif ',' in valor_str and '.' not in valor_str:
                valor_str = valor_str.replace(',', '.')
            
            valor = float(valor_str)
            
            if valor <= 0:
                return jsonify({'success': False, 'message': 'Valor deve ser maior que zero'})
                
        except (ValueError, TypeError) as e:
            print(f"Erro ao converter valor '{valor_str}': {e}")
            return jsonify({'success': False, 'message': 'Valor inválido'})
        
        # Converter valor pago Sócrates Online
        valor_pago_socrates_float = 0.0
        if valor_pago_socrates_str and valor_pago_socrates_str.strip():
            try:
                valor_pago_socrates_str = str(valor_pago_socrates_str).strip()
                
                # Se contém ponto e vírgula, é formato brasileiro (ex: 1.000,50)
                if '.' in valor_pago_socrates_str and ',' in valor_pago_socrates_str:
                    valor_pago_socrates_str = valor_pago_socrates_str.replace('.', '').replace(',', '.')
                # Se contém apenas vírgula, trocar por ponto
                elif ',' in valor_pago_socrates_str and '.' not in valor_pago_socrates_str:
                    valor_pago_socrates_str = valor_pago_socrates_str.replace(',', '.')
                
                valor_pago_socrates_float = float(valor_pago_socrates_str)
                
                if valor_pago_socrates_float < 0:
                    return jsonify({'success': False, 'message': 'Valor pago Sócrates Online não pode ser negativo'})
                    
            except (ValueError, TypeError) as e:
                print(f"Erro ao converter valor pago Sócrates Online '{valor_pago_socrates_str}': {e}")
                return jsonify({'success': False, 'message': 'Valor pago Sócrates Online inválido'})
        
        # Converter data
        try:
            data_obj = datetime.strptime(data_despesa, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': 'Data inválida'})
        
        # Atualizar campos da despesa do evento
        despesa_evento.id_despesa = int(despesa_id)
        despesa_evento.data_vencimento = data_obj
        despesa_evento.valor = valor
        despesa_evento.valor_pago_socrates = valor_pago_socrates_float
        despesa_evento.status_pagamento = request.form.get('status_pagamento', 'pendente')
        despesa_evento.forma_pagamento = request.form.get('forma_pagamento', 'débito')
        despesa_evento.pago_por = request.form.get('pago_por', '')
        despesa_evento.observacoes = request.form.get('observacoes', '')
        
        # Atualizar campos de quantidade se presentes
        qtd_dias_form = request.form.get('qtd_dias', None)
        qtd_pessoas_form = request.form.get('qtd_pessoas', None)
        
        try:
            if qtd_dias_form and qtd_dias_form.strip():
                despesa_evento.qtd_dias = int(qtd_dias_form)
            else:
                despesa_evento.qtd_dias = None
        except (ValueError, AttributeError):
            despesa_evento.qtd_dias = None
        
        try:
            if qtd_pessoas_form and qtd_pessoas_form.strip():
                despesa_evento.qtd_pessoas = int(qtd_pessoas_form)
            else:
                despesa_evento.qtd_pessoas = None
        except (ValueError, AttributeError):
            despesa_evento.qtd_pessoas = None
        
        # Processar upload de comprovante se fornecido
        comprovante = request.files.get('comprovante')
        if comprovante and comprovante.filename:
            if allowed_file(comprovante.filename):
                # Remover arquivo anterior se existir
                if despesa_evento.comprovante:
                    try:
                        old_file_path = os.path.join(app.config['UPLOAD_FOLDER'], despesa_evento.comprovante)
                        if os.path.exists(old_file_path):
                            os.remove(old_file_path)
                            print(f"✅ Arquivo antigo substituído: {despesa_evento.comprovante}")
                    except Exception as e:
                        print(f"⚠️ Erro ao remover arquivo antigo: {e}")
                
                # Gerar nome único para o arquivo (mais simples)
                filename = secure_filename(comprovante.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                name_part, ext_part = os.path.splitext(filename)
                # Usar apenas os primeiros 8 caracteres do nome + timestamp + extensão
                unique_filename = f"{name_part[:8]}_{timestamp}{ext_part}"
                
                # Criar diretório se não existir
                upload_path = os.path.join(app.config['UPLOAD_FOLDER'])
                os.makedirs(upload_path, exist_ok=True)
                
                # Salvar arquivo
                file_path = os.path.join(upload_path, unique_filename)
                comprovante.save(file_path)
                
                # Atualizar campo no banco
                despesa_evento.comprovante = unique_filename
                
                print(f"✅ Comprovante editado e salvo: {unique_filename}")
        
        # Atualizar fornecedor se informado
        id_fornecedor = request.form.get('id_fornecedor')
        if id_fornecedor and id_fornecedor != '0':
            despesa_evento.id_fornecedor = int(id_fornecedor)
            
            # Adicionar fornecedor ao evento se não existir
            fornecedor_evento_existente = FornecedorEvento.query.filter_by(
                id_evento=id_evento,
                id_fornecedor=int(id_fornecedor)
            ).first()
            
            if not fornecedor_evento_existente:
                novo_fornecedor_evento = FornecedorEvento(
                    id_evento=id_evento,
                    id_fornecedor=int(id_fornecedor),
                    observacoes=request.form.get('observacoes', '')
                )
                db.session.add(novo_fornecedor_evento)
        else:
            despesa_evento.id_fornecedor = None
        
        # Processar campo despesa_cabeca
        despesa_cabeca = request.form.get('despesa_cabeca') == '1'
        despesa_evento.despesa_cabeca = despesa_cabeca
        print(f"✅ Campo despesa_cabeca atualizado: {despesa_cabeca}")
        
        db.session.commit()
        
        print(f"✅ Despesa editada com sucesso: ID {despesa_evento_id}")
        print(f"💰 Valores salvos: valor={valor}, valor_pago_socrates={valor_pago_socrates_float}")
        
        return jsonify({
            'success': True, 
            'message': 'Despesa atualizada com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao editar despesa: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'})

@app.route('/api/cidades/<string:estado>')
def api_cidades_por_estado(estado):
    """API para buscar cidades por estado"""
    try:
        # Dicionário de cidades por estado (pode ser movido para um arquivo separado)
        cidades_por_estado = {
            'AC': ['Rio Branco', 'Cruzeiro do Sul', 'Sena Madureira', 'Tarauacá', 'Feijó', 'Senador Guiomard', 'Plácido de Castro', 'Brasiléia', 'Xapuri', 'Epitaciolândia'],
            'AL': ['Maceió', 'Arapiraca', 'Palmeira dos Índios', 'Rio Largo', 'Penedo', 'União dos Palmares', 'São Miguel dos Campos', 'Santana do Ipanema', 'Delmiro Gouveia', 'Coruripe'],
            'AP': ['Macapá', 'Santana', 'Laranjal do Jari', 'Oiapoque', 'Mazagão', 'Porto Grande', 'Tartarugalzinho', 'Vitória do Jari', 'Ferreira Gomes', 'Pedra Branca do Amapari'],
            'AM': ['Manaus', 'Parintins', 'Itacoatiara', 'Manacapuru', 'Coari', 'Tefé', 'Tabatinga', 'Maués', 'São Gabriel da Cachoeira', 'Humaitá'],
            'BA': ['Salvador', 'Feira de Santana', 'Vitória da Conquista', 'Camaçari', 'Juazeiro', 'Ilhéus', 'Itabuna', 'Lauro de Freitas', 'Jequié', 'Teixeira de Freitas', 'Alagoinhas', 'Barreiras', 'Simões Filho', 'Paulo Afonso', 'Eunápolis', 'Porto Seguro'],
            'CE': ['Fortaleza', 'Caucaia', 'Juazeiro do Norte', 'Maracanaú', 'Sobral', 'Crato', 'Itapipoca', 'Maranguape', 'Iguatu', 'Quixadá', 'Canindé', 'Aquiraz', 'Pacatuba', 'Crateús', 'Russas'],
            'DF': ['Brasília', 'Gama', 'Taguatinga', 'Ceilândia', 'Sobradinho', 'Planaltina', 'Samambaia', 'Santa Maria', 'São Sebastião', 'Recanto das Emas'],
            'ES': ['Vitória', 'Cariacica', 'Serra', 'Vila Velha', 'Linhares', 'Colatina', 'Guarapari', 'São Mateus', 'Cachoeiro de Itapemirim', 'Aracruz', 'Viana', 'Nova Venécia', 'Barra de São Francisco'],
            'GO': ['Goiânia', 'Aparecida de Goiânia', 'Anápolis', 'Rio Verde', 'Luziânia', 'Águas Lindas de Goiás', 'Valparaíso de Goiás', 'Trindade', 'Formosa', 'Novo Gama', 'Itumbiara', 'Senador Canedo', 'Catalão', 'Jataí', 'Planaltina'],
            'MA': ['São Luís', 'Imperatriz', 'São José de Ribamar', 'Timon', 'Caxias', 'Codó', 'Paço do Lumiar', 'Açailândia', 'Bacabal', 'Balsas', 'Barra do Corda', 'Santa Inês', 'Pinheiro', 'Pedreiras'],
            'MT': ['Cuiabá', 'Várzea Grande', 'Rondonópolis', 'Sinop', 'Tangará da Serra', 'Cáceres', 'Sorriso', 'Lucas do Rio Verde', 'Barra do Garças', 'Primavera do Leste', 'Alta Floresta', 'Ponta Porã'],
            'MS': ['Campo Grande', 'Dourados', 'Três Lagoas', 'Corumbá', 'Ponta Porã', 'Naviraí', 'Nova Andradina', 'Sidrolândia', 'Maracaju', 'São Gabriel do Oeste', 'Coxim', 'Aquidauana'],
            'MG': ['Belo Horizonte', 'Uberlândia', 'Contagem', 'Juiz de Fora', 'Betim', 'Montes Claros', 'Ribeirão das Neves', 'Uberaba', 'Governador Valadares', 'Ipatinga', 'Sete Lagoas', 'Divinópolis', 'Santa Luzia', 'Ibirité', 'Poços de Caldas', 'Patos de Minas', 'Pouso Alegre', 'Teófilo Otoni', 'Barbacena', 'Sabará', 'Vespasiano', 'Conselheiro Lafaiete', 'Varginha', 'Itabira', 'Passos'],
            'PA': ['Belém', 'Ananindeua', 'Santarém', 'Marabá', 'Parauapebas', 'Castanhal', 'Abaetetuba', 'Cametá', 'Marituba', 'Bragança', 'Altamira', 'Itaituba', 'Tucuruí', 'Benevides'],
            'PB': ['João Pessoa', 'Campina Grande', 'Santa Rita', 'Patos', 'Bayeux', 'Sousa', 'Cajazeiras', 'Cabedelo', 'Guarabira', 'Mamanguape', 'Sapé', 'Itabaiana', 'Monteiro', 'Pombal'],
            'PR': ['Curitiba', 'Londrina', 'Maringá', 'Ponta Grossa', 'Cascavel', 'São José dos Pinhais', 'Foz do Iguaçu', 'Colombo', 'Guarapuava', 'Paranaguá', 'Araucária', 'Toledo', 'Apucarana', 'Pinhais', 'Campo Largo', 'Arapongas', 'Almirante Tamandaré', 'Umuarama', 'Paranavaí', 'Sarandi'],
            'PE': ['Recife', 'Jaboatão dos Guararapes', 'Olinda', 'Caruaru', 'Petrolina', 'Paulista', 'Cabo de Santo Agostinho', 'Camaragibe', 'Garanhuns', 'Vitória de Santo Antão', 'Igarassu', 'São Lourenço da Mata', 'Santa Cruz do Capibaribe', 'Abreu e Lima', 'Ipojuca', 'Serra Talhada', 'Araripina', 'Gravatá', 'Carpina'],
            'PI': ['Teresina', 'Parnaíba', 'Picos', 'Piripiri', 'Floriano', 'Campo Maior', 'Barras', 'União', 'Altos', 'Pedro II', 'Valença do Piauí', 'Esperantina', 'São Raimundo Nonato', 'Cocal'],
            'RJ': ['Rio de Janeiro', 'São Gonçalo', 'Duque de Caxias', 'Nova Iguaçu', 'Niterói', 'Belford Roxo', 'São João de Meriti', 'Campos dos Goytacazes', 'Petrópolis', 'Volta Redonda', 'Magé', 'Macaé', 'Itaboraí', 'Cabo Frio', 'Angra dos Reis', 'Nova Friburgo', 'Barra Mansa', 'Teresópolis', 'Mesquita', 'Nilópolis'],
            'RN': ['Natal', 'Mossoró', 'Parnamirim', 'São Gonçalo do Amarante', 'Macaíba', 'Ceará-Mirim', 'Caicó', 'Assu', 'Currais Novos', 'Nova Cruz', 'São José de Mipibu', 'Apodi', 'João Câmara', 'Pau dos Ferros'],
            'RS': ['Porto Alegre', 'Caxias do Sul', 'Pelotas', 'Canoas', 'Santa Maria', 'Gravataí', 'Viamão', 'Novo Hamburgo', 'São Leopoldo', 'Rio Grande', 'Alvorada', 'Passo Fundo', 'Sapucaia do Sul', 'Uruguaiana', 'Santa Cruz do Sul', 'Cachoeirinha', 'Bagé', 'Bento Gonçalves', 'Erechim', 'Guaíba', 'Cachoeira do Sul'],
            'RO': ['Porto Velho', 'Ji-Paraná', 'Ariquemes', 'Vilhena', 'Cacoal', 'Rolim de Moura', 'Guajará-Mirim', 'Jaru', 'Ouro Preto do Oeste', 'Machadinho do Oeste'],
            'RR': ['Boa Vista', 'Rorainópolis', 'Caracaraí', 'Alto Alegre', 'Mucajaí', 'São Luiz', 'São João da Baliza', 'Pacaraima', 'Iracema', 'Amajari'],
            'SC': ['Florianópolis', 'Joinville', 'Blumenau', 'São José', 'Criciúma', 'Chapecó', 'Itajaí', 'Lages', 'Jaraguá do Sul', 'Palhoça', 'Balneário Camboriú', 'Brusque', 'Tubarão', 'São Bento do Sul', 'Caçador', 'Camboriú', 'Navegantes', 'Concórdia', 'Rio do Sul', 'Araranguá'],
            'SP': ['São Paulo', 'Guarulhos', 'Campinas', 'São Bernardo do Campo', 'Santo André', 'Osasco', 'Ribeirão Preto', 'Sorocaba', 'Mauá', 'São José dos Campos', 'Mogi das Cruzes', 'Diadema', 'Jundiaí', 'Carapicuíba', 'Piracicaba', 'Bauru', 'Itaquaquecetuba', 'São Vicente', 'Franca', 'Guarujá', 'Taubaté', 'Praia Grande', 'Limeira', 'Suzano', 'Taboão da Serra', 'Sumaré', 'Barueri', 'Embu das Artes', 'São Carlos', 'Marília'],
            'SE': ['Aracaju', 'Nossa Senhora do Socorro', 'Lagarto', 'Itabaiana', 'São Cristóvão', 'Estância', 'Tobias Barreto', 'Simão Dias', 'Propriá', 'Capela'],
            'TO': ['Palmas', 'Araguaína', 'Gurupi', 'Porto Nacional', 'Paraíso do Tocantins', 'Colinas do Tocantins', 'Guaraí', 'Tocantinópolis', 'Miracema do Tocantins', 'Dianópolis']
        }
        
        estado_upper = estado.upper()
        
        if estado_upper not in cidades_por_estado:
            return jsonify({
                'success': False,
                'message': 'Estado não encontrado'
            }), 404
        
        cidades = cidades_por_estado[estado_upper]
        
        return jsonify({
            'success': True,
            'estado': estado_upper,
            'cidades': cidades,
            'total': len(cidades)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao buscar cidades: {str(e)}'
        }), 500

@app.route('/api/fornecedores-busca')
def api_fornecedores_busca():
    """API para buscar fornecedores com priorizaÃ§ão por localizaÃ§ão"""
    try:
        # ParÃ¢metros de busca
        termo_busca = request.args.get('q', '').strip()
        cidade_evento = request.args.get('cidade', '').strip()
        estado_evento = request.args.get('estado', '').strip()
        
        # Query base
        query = Fornecedor.query
        
        # Se há termo de busca, filtrar por nome
        if termo_busca:
            query = query.filter(Fornecedor.nome.ilike(f'%{termo_busca}%'))
        
        # Buscar todos os fornecedores que atendem ao critério
        fornecedores = query.all()
        
        # Separar fornecedores por prioridade
        fornecedores_locais = []
        fornecedores_outros = []
        
        for fornecedor in fornecedores:
            # Priorizar fornecedores da mesma cidade/estado
            if (fornecedor.cidade and fornecedor.estado and 
                cidade_evento and estado_evento and
                fornecedor.cidade.lower() == cidade_evento.lower() and 
                fornecedor.estado.lower() == estado_evento.lower()):
                fornecedores_locais.append(fornecedor)
            else:
                fornecedores_outros.append(fornecedor)
        
        # Ordenar alfabeticamente dentro de cada grupo
        fornecedores_locais.sort(key=lambda x: x.nome.lower())
        fornecedores_outros.sort(key=lambda x: x.nome.lower())
        
        # Combinar listas (locais primeiro)
        fornecedores_ordenados = fornecedores_locais + fornecedores_outros
        
        # Preparar resposta JSON
        resultado = []
        for fornecedor in fornecedores_ordenados:
            resultado.append({
                'id': fornecedor.id_fornecedor,
                'nome': fornecedor.nome,
                'telefone': fornecedor.telefone or '',
                'cidade': fornecedor.cidade or '',
                'estado': fornecedor.estado or '',
                'categoria': fornecedor.categoria.nome if fornecedor.categoria else '',
                'is_local': fornecedor in fornecedores_locais
            })
        
        return jsonify({
            'success': True,
            'fornecedores': resultado,
            'total': len(resultado),
            'locais': len(fornecedores_locais),
            'outros': len(fornecedores_outros)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao buscar fornecedores: {str(e)}'
        }), 500

@app.route('/eventos/<int:id_evento>/atualizar-despesa-cabeca/<int:despesa_evento_id>', methods=['PUT'])
def atualizar_despesa_cabeca(id_evento, despesa_evento_id):
    """Atualiza apenas o campo despesa_cabeca de uma despesa específica do evento"""
    try:
        data = request.get_json()
        despesa_cabeca = data.get('despesa_cabeca', False)
        
        print(f"=== ATUALIZANDO DESPESA CABEÇA ===")
        print(f"Evento ID: {id_evento}")
        print(f"Despesa Evento ID: {despesa_evento_id}")
        print(f"Nova flag despesa_cabeca: {despesa_cabeca}")
        
        # Buscar a despesa do evento
        despesa_evento = DespesaEvento.query.filter_by(
            id_despesa_evento=despesa_evento_id,
            id_evento=id_evento
        ).first()
        
        if not despesa_evento:
            return jsonify({
                'success': False, 
                'message': 'Despesa do evento não encontrada'
            }), 404
        
        # Atualizar apenas o campo despesa_cabeca
        despesa_evento.despesa_cabeca = despesa_cabeca
        
        db.session.commit()
        
        print(f"✅ Despesa cabeça atualizada: {despesa_evento.despesa.nome} - Cabeça: {despesa_cabeca}")
        
        return jsonify({
            'success': True,
            'message': f'Despesa {"marcada" if despesa_cabeca else "desmarcada"} como despesa da cabeça',
            'despesa_cabeca': despesa_cabeca,
            'despesa_nome': despesa_evento.despesa.nome
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao atualizar despesa cabeça: {e}")
        return jsonify({
            'success': False, 
            'message': f'Erro ao atualizar: {str(e)}'
        }), 500

# Rota para excluir comprovante de despesa
@app.route('/eventos/<int:id_evento>/excluir-comprovante/<int:despesa_evento_id>', methods=['DELETE'])
def excluir_comprovante_despesa(id_evento, despesa_evento_id):
    """Exclui o comprovante de uma despesa específica do evento"""
    try:
        # Buscar a despesa do evento
        despesa_evento = DespesaEvento.query.filter_by(
            id_despesa_evento=despesa_evento_id,
            id_evento=id_evento
        ).first()
        
        if not despesa_evento:
            return jsonify({
                'success': False, 
                'message': 'Despesa não encontrada'
            }), 404
        
        # Verificar se há comprovante para excluir
        if not despesa_evento.comprovante:
            return jsonify({
                'success': False, 
                'message': 'Nenhum comprovante encontrado para esta despesa'
            }), 400
        
        # Tentar remover o arquivo físico
        try:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], despesa_evento.comprovante)
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"✅ Arquivo removido: {file_path}")
            else:
                print(f"⚠️ Arquivo não encontrado: {file_path}")
        except Exception as e:
            print(f"⚠️ Erro ao remover arquivo: {e}")
            # Continua mesmo se não conseguir remover o arquivo físico
        
        # Limpar o campo comprovante no banco
        despesa_evento.comprovante = None
        db.session.commit()
        
        print(f"✅ Comprovante excluído da despesa ID: {despesa_evento_id}")
        
        return jsonify({
            'success': True,
            'message': 'Comprovante excluído com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao excluir comprovante: {e}")
        return jsonify({
            'success': False, 
            'message': f'Erro ao excluir comprovante: {str(e)}'
        }), 500

# Rota para servir arquivos de comprovante
@app.route('/uploads/comprovantes/<filename>')
def uploaded_file(filename):
    # Verificar se o usuário está logado
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    from flask import send_from_directory
    import os
    return send_from_directory(os.path.join(app.config['UPLOAD_FOLDER']), filename)

@app.route('/api/evento/<int:id_evento>/detalhes-completos')
def api_evento_detalhes_completos(id_evento):
    """API para buscar detalhes completos do evento para o modal"""
    # Verificar se o usuário está logado
    if 'user_id' not in session:
        return jsonify({
            'success': False,
            'message': 'Usuário não autenticado'
        }), 401
        
    try:
        app.logger.info(f"Buscando detalhes completos para evento ID: {id_evento}")
        
        # Verificar se o evento existe
        evento = Evento.query.get(id_evento)
        if not evento:
            app.logger.error(f"Evento não encontrado: {id_evento}")
            return jsonify({
                'success': False,
                'message': 'Evento não encontrado'
            }), 404
        
        app.logger.info(f"Evento encontrado: {evento.nome}")
        
        # Usar a função unificada para obter dados
        dados = obter_dados_completos_evento(id_evento)
        app.logger.info("Dados obtidos com sucesso")
        
        # Preparar receitas formatadas para JSON
        receitas_formatadas = []
        for categoria in dados['receitas_agrupadas']:
            categoria_formatada = {
                'categoria_nome': categoria['categoria_nome'],
                'total_categoria': categoria['total_categoria'],
                'itens': []
            }
            for item in categoria['itens']:
                categoria_formatada['itens'].append({
                    'receita_nome': item['receita_nome'],
                    'valor': item['valor'],
                    'data': item['data'].strftime('%d/%m/%Y') if item['data'] else '',
                    'observacoes': item['observacoes']
                })
            receitas_formatadas.append(categoria_formatada)
        
        # Preparar despesas formatadas para JSON
        despesas_formatadas = []
        for categoria in dados['despesas_agrupadas']:
            categoria_formatada = {
                'categoria_nome': categoria['categoria_nome'],
                'total_categoria': categoria['total_categoria'],
                'itens': []
            }
            for item in categoria['itens']:
                categoria_formatada['itens'].append({
                    'despesa_nome': item['despesa_nome'],
                    'valor': item['valor'],
                    'valor_pago_socrates': item['valor_pago_socrates'],
                    'data': item['data'].strftime('%d/%m/%Y') if item['data'] else '',
                    'status_pagamento': item['status_pagamento'],
                    'despesa_cabeca': item['despesa_cabeca'],
                    'fornecedor_nome': item['fornecedor_nome'],
                    'observacoes': item['observacoes']
                })
            despesas_formatadas.append(categoria_formatada)
        
        return jsonify({
            'success': True,
            'evento': {
                'nome': dados['evento'].nome,
                'data_inicio': dados['evento'].data_inicio.strftime('%d/%m/%Y'),
                'data_fim': dados['evento'].data_fim.strftime('%d/%m/%Y'),
                'cidade': dados['evento'].cidade,
                'estado': dados['evento'].estado,
                'endereco': dados['evento'].endereco or '',
                'status': dados['evento'].status,
                'circo_nome': dados['evento'].circo.nome if dados['evento'].circo else '',
                'produtor_nome': dados['evento'].produtor.nome if dados['evento'].produtor else '',
                'observacoes': dados['evento'].observacoes or ''
            },
            'receitas': {
                'categorias': receitas_formatadas,
                'total': dados['totais_calculados']['total_receitas']
            },
            'despesas': {
                'categorias': despesas_formatadas,
                'total': dados['totais_calculados']['total_despesas'],
                'despesas_cabeca_total': dados['totais_calculados']['despesas_cabeca_total']
            },
            'financeiro': {
                'total_receitas': dados['calculo_financeiro']['total_receitas'],
                'total_despesas': dados['calculo_financeiro']['total_despesas'],
                'despesas_cabeca': dados['calculo_financeiro']['despesas_cabeca'],
                'total_liquido': dados['calculo_financeiro']['total_liquido'],
                'cinquenta_porcento_show': dados['calculo_financeiro']['cinquenta_porcento_show'],
                'reembolso_midias': dados['calculo_financeiro']['reembolso_midias'],
                'repasse_total': dados['calculo_financeiro']['repasse_total'],
                'total_despesas_socrates': dados['calculo_financeiro']['total_despesas_socrates'],
                'resultado_show': dados['calculo_financeiro']['resultado_show']
            }
        })
        
    except Exception as e:
        app.logger.error(f"Erro ao buscar detalhes do evento {id_evento}: {str(e)}")
        app.logger.error(f"Traceback: {e.__class__.__name__}: {str(e)}")
        import traceback
        app.logger.error(f"Stack trace: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': f'Erro ao buscar detalhes do evento: {str(e)}'
        }), 500

def obter_dados_completos_evento(id_evento):
    """
    Função unificada para obter dados completos de um evento
    Retorna: dicionário com evento, receitas agrupadas, despesas agrupadas e cálculos financeiros
    """
    evento = Evento.query.get_or_404(id_evento)
    
    # Buscar receitas do evento agrupadas por categoria
    receitas_query = db.session.query(
        CategoriaReceita.nome.label('categoria_nome'),
        Receita.nome.label('receita_nome'),
        ReceitaEvento.valor,
        ReceitaEvento.data,
        ReceitaEvento.observacoes
    ).join(
        ReceitaEvento, ReceitaEvento.id_receita == Receita.id_receita
    ).join(
        CategoriaReceita, Receita.id_categoria_receita == CategoriaReceita.id_categoria_receita
    ).filter(
        ReceitaEvento.id_evento == id_evento
    ).order_by(CategoriaReceita.nome, Receita.nome).all()
    
    # Agrupar receitas por categoria
    receitas_agrupadas = {}
    for item in receitas_query:
        if item.categoria_nome not in receitas_agrupadas:
            receitas_agrupadas[item.categoria_nome] = {
                'categoria_nome': item.categoria_nome,
                'total_categoria': 0,
                'itens': []
            }
        
        receitas_agrupadas[item.categoria_nome]['total_categoria'] += float(item.valor)
        receitas_agrupadas[item.categoria_nome]['itens'].append({
            'receita_nome': item.receita_nome,
            'valor': float(item.valor),
            'data': item.data,
            'observacoes': item.observacoes or ''
        })
    
    # Buscar despesas do evento agrupadas por categoria
    despesas_query = db.session.query(
        CategoriaDespesa.nome.label('categoria_nome'),
        Despesa.nome.label('despesa_nome'),
        DespesaEvento.valor,
        DespesaEvento.valor_pago_socrates,
        DespesaEvento.data_vencimento,
        DespesaEvento.status_pagamento,
        DespesaEvento.despesa_cabeca,
        Fornecedor.nome.label('fornecedor_nome'),
        DespesaEvento.observacoes
    ).join(
        DespesaEvento, DespesaEvento.id_despesa == Despesa.id_despesa
    ).join(
        CategoriaDespesa, Despesa.id_categoria_despesa == CategoriaDespesa.id_categoria_despesa
    ).outerjoin(
        Fornecedor, DespesaEvento.id_fornecedor == Fornecedor.id_fornecedor
    ).filter(
        DespesaEvento.id_evento == id_evento
    ).order_by(CategoriaDespesa.nome, Despesa.nome).all()
    
    # Agrupar despesas por categoria
    despesas_agrupadas = {}
    for item in despesas_query:
        if item.categoria_nome not in despesas_agrupadas:
            despesas_agrupadas[item.categoria_nome] = {
                'categoria_nome': item.categoria_nome,
                'total_categoria': 0,
                'itens': []
            }
        
        # Usar valor_pago_socrates se disponível, senão usar valor total
        valor_para_soma = float(item.valor_pago_socrates) if item.valor_pago_socrates else float(item.valor)
        despesas_agrupadas[item.categoria_nome]['total_categoria'] += valor_para_soma
        
        despesas_agrupadas[item.categoria_nome]['itens'].append({
            'despesa_nome': item.despesa_nome,
            'valor': float(item.valor),
            'valor_pago_socrates': float(item.valor_pago_socrates) if item.valor_pago_socrates else None,
            'valor_exibicao': valor_para_soma,  # Valor que será exibido (prioriza valor_pago_socrates)
            'data': item.data_vencimento,
            'status_pagamento': item.status_pagamento or 'pendente',
            'despesa_cabeca': bool(item.despesa_cabeca),
            'fornecedor_nome': item.fornecedor_nome or '',
            'observacoes': item.observacoes or ''
        })
    
    # Calcular valores usando a função unificada (única fonte de verdade)
    calculo = calcular_lucro_evento(id_evento)
    
    # Filtrar despesas agrupadas excluindo categorias "PAGAS PELO CIRCO" e "PAGO PELO CIRCO"
    despesas_agrupadas_filtradas = []
    for categoria in despesas_agrupadas.values():
        categoria_nome_upper = categoria['categoria_nome'].upper()
        if 'PAGAS PELO CIRCO' not in categoria_nome_upper and 'PAGO PELO CIRCO' not in categoria_nome_upper:
            despesas_agrupadas_filtradas.append(categoria)
    
    # Filtrar despesas de cabeça - apenas itens com despesa_cabeca=True
    # EXCETO para categoria "PAGAS PELO CIRCO" que deve aparecer completa
    despesas_cabeca_agrupadas = []
    for categoria in despesas_agrupadas.values():
        categoria_nome_upper = categoria['categoria_nome'].upper()
        
        # Se é categoria "PAGAS PELO CIRCO", incluir todos os itens
        if 'PAGAS PELO CIRCO' in categoria_nome_upper or 'PAGO PELO CIRCO' in categoria_nome_upper:
            # Garantir que todos os itens tenham valor_exibicao (que já foi calculado anteriormente)
            despesas_cabeca_agrupadas.append({
                'categoria_nome': categoria['categoria_nome'],
                'total_categoria': categoria['total_categoria'],
                'itens': categoria['itens']  # Estes já têm valor_exibicao
            })
        else:
            # Para outras categorias, filtrar apenas os itens que são despesas de cabeça
            itens_cabeca = [item for item in categoria['itens'] if item['despesa_cabeca']]
            
            if itens_cabeca:  # Se há itens de cabeça nesta categoria
                # Recalcular o total da categoria apenas com os itens de cabeça
                total_categoria_cabeca = sum(item['valor_exibicao'] for item in itens_cabeca)
                
                despesas_cabeca_agrupadas.append({
                    'categoria_nome': categoria['categoria_nome'],
                    'total_categoria': total_categoria_cabeca,
                    'itens': itens_cabeca
                })
    
    # Calcular o total das despesas de cabeça baseado no que está sendo exibido
    # (despesas com flag despesa_cabeca=True + todas as despesas PAGAS PELO CIRCO)
    total_despesas_cabeca_exibidas = sum(categoria['total_categoria'] for categoria in despesas_cabeca_agrupadas)
    
    # Calcular o total das despesas Sócrates baseado nos valores exibidos (valor_pago_socrates)
    total_despesas_socrates_exibidas = sum(categoria['total_categoria'] for categoria in despesas_agrupadas_filtradas)
    
    return {
        'evento': evento,
        'receitas_agrupadas': list(receitas_agrupadas.values()),
        'despesas_agrupadas': despesas_agrupadas_filtradas,
        'despesas_cabeca_agrupadas': despesas_cabeca_agrupadas,
        'totais_calculados': {
            'total_receitas': calculo['total_receitas'],      # Usar valores da função unificada
            'total_despesas': calculo['total_despesas'],      # para garantir consistência
            'despesas_cabeca_total': total_despesas_cabeca_exibidas,  # Total baseado no que é exibido
            'reembolso_midias': calculo['reembolso_midias'],
            'total_despesas_socrates_exibidas': total_despesas_socrates_exibidas  # Total baseado nos valores exibidos
        },
        'calculo_financeiro': calculo
    }

# Funções utilitárias para exportação
def criar_excel_response(headers, data, filename):
    """Cria um arquivo Excel e retorna como response Flask"""
    wb = Workbook()
    ws = wb.active
    
    # Configurar cabeçalho
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True)
    
    # Adicionar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Adicionar dados
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    
    return response

def criar_pdf_response(headers, data, title, filename):
    """Cria um arquivo PDF e retorna como response Flask"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Título
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))
    
    # Preparar dados da tabela
    table_data = [headers] + data
    
    # Criar tabela
    table = Table(table_data)
    
    # Estilo da tabela
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}.pdf'
    
    return response

# Rotas de exportação
# Rotas API para listar eventos por entidade
@app.route('/api/colaborador/<int:colaborador_id>/eventos')
def api_colaborador_eventos(colaborador_id):
    """API para listar eventos onde um colaborador participou da equipe"""
    if 'user_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        colaborador = Colaborador.query.get_or_404(colaborador_id)
        
        # Buscar eventos onde o colaborador participou da equipe
        equipes_evento = EquipeEvento.query.filter_by(id_colaborador=colaborador_id).all()
        
        eventos = []
        for equipe in equipes_evento:
            evento = Evento.query.get(equipe.id_evento)
            if evento:
                eventos.append({
                    'id_evento': evento.id_evento,
                    'nome': evento.nome,
                    'circo': evento.circo.nome if evento.circo else 'Sem circo',
                    'data_inicio': evento.data_inicio.strftime('%d/%m/%Y') if evento.data_inicio else '',
                    'data_fim': evento.data_fim.strftime('%d/%m/%Y') if evento.data_fim else '',
                    'cidade': evento.cidade or '',
                    'estado': evento.estado or '',
                    'funcao': equipe.funcao or 'Não informado',
                    'observacoes': equipe.observacoes or ''
                })
        
        # Ordenar por data mais recente primeiro
        eventos.sort(key=lambda x: x['data_inicio'], reverse=True)
        
        return jsonify({
            'success': True,
            'colaborador': colaborador.nome,
            'eventos': eventos,
            'total': len(eventos)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/fornecedor/<int:fornecedor_id>/eventos')
def api_fornecedor_eventos(fornecedor_id):
    """API para listar eventos onde um fornecedor prestou serviços"""
    if 'user_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        fornecedor = Fornecedor.query.get_or_404(fornecedor_id)
        
        # Buscar eventos onde o fornecedor foi usado (via FornecedorEvento ou DespesaEvento)
        eventos_ids = set()
        
        # Via tabela fornecedor_evento
        fornecedores_evento = FornecedorEvento.query.filter_by(id_fornecedor=fornecedor_id).all()
        for fe in fornecedores_evento:
            eventos_ids.add(fe.id_evento)
        
        # Via despesas que usaram esse fornecedor
        despesas_evento = DespesaEvento.query.filter_by(id_fornecedor=fornecedor_id).all()
        for de in despesas_evento:
            eventos_ids.add(de.id_evento)
        
        # Buscar dados dos eventos
        eventos = []
        for evento_id in eventos_ids:
            evento = Evento.query.get(evento_id)
            if evento:
                # Calcular valor total gasto com este fornecedor no evento
                valor_total = db.session.query(func.sum(DespesaEvento.valor)).filter_by(
                    id_evento=evento_id,
                    id_fornecedor=fornecedor_id
                ).scalar() or 0
                
                eventos.append({
                    'id_evento': evento.id_evento,
                    'nome': evento.nome,
                    'circo': evento.circo.nome if evento.circo else 'Sem circo',
                    'data_inicio': evento.data_inicio.strftime('%d/%m/%Y') if evento.data_inicio else '',
                    'data_fim': evento.data_fim.strftime('%d/%m/%Y') if evento.data_fim else '',
                    'cidade': evento.cidade or '',
                    'estado': evento.estado or '',
                    'valor_total': float(valor_total)
                })
        
        # Ordenar por data mais recente primeiro
        eventos.sort(key=lambda x: x['data_inicio'], reverse=True)
        
        return jsonify({
            'success': True,
            'fornecedor': fornecedor.nome,
            'eventos': eventos,
            'total': len(eventos)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/elenco/<int:elenco_id>/eventos')
def api_elenco_eventos(elenco_id):
    """API para listar eventos onde um membro do elenco participou"""
    if 'user_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        membro_elenco = Elenco.query.get_or_404(elenco_id)
        
        # Buscar eventos onde o membro do elenco participou
        elencos_evento = ElencoEvento.query.filter_by(id_elenco=elenco_id).all()
        
        eventos = []
        for elenco_evt in elencos_evento:
            evento = Evento.query.get(elenco_evt.id_evento)
            if evento:
                eventos.append({
                    'id_evento': evento.id_evento,
                    'nome': evento.nome,
                    'circo': evento.circo.nome if evento.circo else 'Sem circo',
                    'data_inicio': evento.data_inicio.strftime('%d/%m/%Y') if evento.data_inicio else '',
                    'data_fim': evento.data_fim.strftime('%d/%m/%Y') if evento.data_fim else '',
                    'cidade': evento.cidade or '',
                    'estado': evento.estado or '',
                    'observacoes': elenco_evt.observacoes or ''
                })
        
        # Ordenar por data mais recente primeiro
        eventos.sort(key=lambda x: x['data_inicio'], reverse=True)
        
        return jsonify({
            'success': True,
            'membro_elenco': membro_elenco.nome,
            'eventos': eventos,
            'total': len(eventos)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/exportar/<string:table_name>/<string:format>', methods=['POST'])
def exportar_dados(table_name, format):
    """Rota genérica para exportação de dados"""
    if 'user_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        # Receber dados do frontend
        data = request.get_json()
        headers = data.get('headers', [])
        rows = data.get('data', [])
        
        # Gerar nome do arquivo
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f"{table_name}_{today}"
        
        # Mapear nomes de tabelas para títulos amigáveis
        table_titles = {
            'colaboradores': 'Colaboradores',
            'fornecedores': 'Fornecedores',
            'veiculos': 'Veículos',
            'despesas': 'Despesas',
            'receitas': 'Receitas',
            'elenco': 'Elenco',
            'categorias_colaborador': 'Categorias de Colaborador',
            'categorias_fornecedor': 'Categorias de Fornecedor',
            'categorias_receita': 'Categorias de Receita',
            'categorias_despesa': 'Categorias de Despesa',
            'categorias_veiculo': 'Categorias de Veículo',
            'circos': 'Circos'
        }
        
        title = table_titles.get(table_name, table_name.title())
        
        if format == 'excel':
            return criar_excel_response(headers, rows, filename)
        elif format == 'pdf':
            return criar_pdf_response(headers, rows, title, filename)
        else:
            return jsonify({'error': 'Formato não suportado'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/relatorios/despesas-fixas')
def relatorio_despesas_fixas():
    """Relatório de todas as despesas fixas da empresa"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        # Buscar todas as despesas fixas (tipos 1 e 3)
        # Tipo 1: Fixas - Evento
        # Tipo 3: Fixas - SócratesOnline
        despesas_fixas = db.session.query(Despesa).filter(
            Despesa.id_tipo_despesa.in_([1, 3])
        ).join(CategoriaDespesa).order_by(
            Despesa.id_tipo_despesa,
            CategoriaDespesa.nome,
            Despesa.nome
        ).all()
        
        # Organizar despesas por tipo
        despesas_por_tipo = {
            1: [],  # Fixas - Evento
            3: []   # Fixas - SócratesOnline
        }
        
        total_valor_medio = 0
        
        for despesa in despesas_fixas:
            despesa_data = {
                'id_despesa': despesa.id_despesa,
                'nome': despesa.nome,
                'categoria': despesa.categoria.nome if despesa.categoria else 'Sem categoria',
                'valor_medio': float(despesa.valor_medio_despesa) if despesa.valor_medio_despesa else 0,
                'tipo_nome': despesa.tipo_nome
            }
            
            despesas_por_tipo[despesa.id_tipo_despesa].append(despesa_data)
            total_valor_medio += despesa_data['valor_medio']
        
        # Calcular totais por tipo
        total_fixas_evento = sum(d['valor_medio'] for d in despesas_por_tipo[1])
        total_fixas_socrates = sum(d['valor_medio'] for d in despesas_por_tipo[3])
        
        estatisticas = {
            'total_despesas': len(despesas_fixas),
            'total_fixas_evento': len(despesas_por_tipo[1]),
            'total_fixas_socrates': len(despesas_por_tipo[3]),
            'valor_total_fixas_evento': total_fixas_evento,
            'valor_total_fixas_socrates': total_fixas_socrates,
            'valor_total_geral': total_valor_medio
        }
        
        return render_template('relatorio_despesas_fixas.html',
                             despesas_por_tipo=despesas_por_tipo,
                             estatisticas=estatisticas,
                             tipos_despesa=TIPOS_DESPESA)
        
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/relatorios/despesas-fixas/exportar/<string:formato>')
def exportar_despesas_fixas(formato):
    """Exportar relatório de despesas fixas"""
    if 'user_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        # Buscar dados das despesas fixas
        despesas_fixas = db.session.query(Despesa).filter(
            Despesa.id_tipo_despesa.in_([1, 3])
        ).join(CategoriaDespesa).order_by(
            Despesa.id_tipo_despesa,
            CategoriaDespesa.nome,
            Despesa.nome
        ).all()
        
        # Preparar dados para exportação
        headers = ['Tipo', 'Nome', 'Categoria', 'Valor Médio']
        data = []
        
        for despesa in despesas_fixas:
            data.append([
                despesa.tipo_nome,
                despesa.nome,
                despesa.categoria.nome if despesa.categoria else 'Sem categoria',
                f"R$ {float(despesa.valor_medio_despesa):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if despesa.valor_medio_despesa else 'R$ 0,00'
            ])
        
        # Adicionar linha de total
        total_valor = sum(float(d.valor_medio_despesa) if d.valor_medio_despesa else 0 for d in despesas_fixas)
        data.append(['', '', 'TOTAL GERAL:', f"R$ {total_valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')])
        
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f"despesas_fixas_{today}"
        
        if formato == 'excel':
            return criar_excel_response(headers, data, filename)
        elif formato == 'pdf':
            return criar_pdf_response(headers, data, 'Relatório de Despesas Fixas', filename)
        else:
            return jsonify({'error': 'Formato não suportado'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/relatorios/veiculos')
def relatorio_veiculos():
    """Relatório de veículos e eventos associados"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verificar se é ROOT ou administrador
    if not (is_root_user() or session.get('categoria', '').lower() == 'administrativo'):
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        # Obter filtros
        categoria_filtro = request.args.get('categoria', '')
        data_inicio = request.args.get('data_inicio', '')
        data_fim = request.args.get('data_fim', '')
        
        # Query base de veículos
        veiculos_query = db.session.query(Veiculo).join(CategoriaVeiculo)
        
        # Aplicar filtro de categoria se fornecido
        if categoria_filtro:
            veiculos_query = veiculos_query.filter(Veiculo.id_categoria_veiculo == categoria_filtro)
        
        # Ordenar por categoria e nome
        veiculos = veiculos_query.order_by(CategoriaVeiculo.nome, Veiculo.nome).all()
        
        # Buscar eventos de veículos com filtro de data
        eventos_query = db.session.query(VeiculoEvento).join(
            Evento, VeiculoEvento.id_evento == Evento.id_evento
        ).join(
            Veiculo, VeiculoEvento.id_veiculo == Veiculo.id_veiculo
        ).join(
            Colaborador, VeiculoEvento.id_motorista == Colaborador.id_colaborador
        )
        
        # Aplicar filtros de data se fornecidos
        if data_inicio:
            eventos_query = eventos_query.filter(VeiculoEvento.data_inicio >= datetime.strptime(data_inicio, '%Y-%m-%d').date())
        if data_fim:
            eventos_query = eventos_query.filter(VeiculoEvento.data_devolucao <= datetime.strptime(data_fim, '%Y-%m-%d').date())
        
        # Aplicar filtro de categoria se fornecido
        if categoria_filtro:
            eventos_query = eventos_query.filter(Veiculo.id_categoria_veiculo == categoria_filtro)
        
        # Ordenar por data de início
        eventos_veiculos = eventos_query.order_by(VeiculoEvento.data_inicio.desc()).all()
        
        # Organizar dados por veículo
        dados_veiculos = {}
        for veiculo in veiculos:
            dados_veiculos[veiculo.id_veiculo] = {
                'veiculo': veiculo,
                'eventos': [],
                'total_eventos': 0,
                'dias_uso': 0
            }
        
        # Adicionar eventos aos veículos
        for evento_veiculo in eventos_veiculos:
            veiculo_id = evento_veiculo.id_veiculo
            if veiculo_id in dados_veiculos:
                dados_veiculos[veiculo_id]['eventos'].append(evento_veiculo)
                dados_veiculos[veiculo_id]['total_eventos'] += 1
                
                # Calcular dias de uso (apenas se ambas as datas existirem)
                if evento_veiculo.data_devolucao and evento_veiculo.data_inicio:
                    delta = evento_veiculo.data_devolucao - evento_veiculo.data_inicio
                    dados_veiculos[veiculo_id]['dias_uso'] += delta.days + 1
                elif evento_veiculo.data_inicio:
                    # Se só tem data de início, considera 1 dia
                    dados_veiculos[veiculo_id]['dias_uso'] += 1
        
        # Buscar categorias para o filtro
        categorias = CategoriaVeiculo.query.order_by(CategoriaVeiculo.nome).all()
        
        # Estatísticas gerais
        total_veiculos = len(veiculos)
        veiculos_com_eventos = sum(1 for dados in dados_veiculos.values() if dados['total_eventos'] > 0)
        total_eventos_veiculos = sum(dados['total_eventos'] for dados in dados_veiculos.values())
        total_dias_uso = sum(dados['dias_uso'] for dados in dados_veiculos.values())
        
        estatisticas = {
            'total_veiculos': total_veiculos,
            'veiculos_com_eventos': veiculos_com_eventos,
            'veiculos_sem_eventos': total_veiculos - veiculos_com_eventos,
            'total_eventos_veiculos': total_eventos_veiculos,
            'total_dias_uso': total_dias_uso,
            'media_dias_por_veiculo': round(total_dias_uso / total_veiculos, 1) if total_veiculos > 0 else 0
        }
        
        return render_template('relatorio_veiculos.html',
                             dados_veiculos=dados_veiculos,
                             categorias=categorias,
                             categoria_filtro=categoria_filtro,
                             data_inicio=data_inicio,
                             data_fim=data_fim,
                             estatisticas=estatisticas)
        
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/relatorios/veiculos/exportar/<string:formato>')
def exportar_relatorio_veiculos(formato):
    """Exportar relatório de veículos"""
    if 'user_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    # Verificar se é ROOT ou administrador
    if not (is_root_user() or session.get('categoria', '').lower() == 'administrativo'):
        return jsonify({'error': 'Acesso restrito a administradores'}), 403
    
    try:
        # Obter filtros
        categoria_filtro = request.args.get('categoria', '')
        data_inicio = request.args.get('data_inicio', '')
        data_fim = request.args.get('data_fim', '')
        
        # Query base de veículos
        veiculos_query = db.session.query(Veiculo).join(CategoriaVeiculo)
        
        # Aplicar filtro de categoria se fornecido
        if categoria_filtro:
            veiculos_query = veiculos_query.filter(Veiculo.id_categoria_veiculo == categoria_filtro)
        
        veiculos = veiculos_query.order_by(CategoriaVeiculo.nome, Veiculo.nome).all()
        
        # Buscar eventos de veículos com filtro de data
        eventos_query = db.session.query(VeiculoEvento).join(
            Evento, VeiculoEvento.id_evento == Evento.id_evento
        ).join(
            Veiculo, VeiculoEvento.id_veiculo == Veiculo.id_veiculo
        ).join(
            Colaborador, VeiculoEvento.id_motorista == Colaborador.id_colaborador
        )
        
        # Aplicar filtros de data se fornecidos
        if data_inicio:
            eventos_query = eventos_query.filter(VeiculoEvento.data_inicio >= datetime.strptime(data_inicio, '%Y-%m-%d').date())
        if data_fim:
            eventos_query = eventos_query.filter(VeiculoEvento.data_devolucao <= datetime.strptime(data_fim, '%Y-%m-%d').date())
        
        # Aplicar filtro de categoria se fornecido
        if categoria_filtro:
            eventos_query = eventos_query.filter(Veiculo.id_categoria_veiculo == categoria_filtro)
        
        eventos_veiculos = eventos_query.order_by(VeiculoEvento.data_inicio.desc()).all()
        
        # Preparar dados para exportação
        headers = ['Veículo', 'Categoria', 'Placa', 'Evento', 'Motorista', 'Data Início', 'Data Devolução', 'Dias de Uso', 'Observações']
        data = []
        
        # Organizar dados por veículo
        for veiculo in veiculos:
            eventos_veiculo = [ev for ev in eventos_veiculos if ev.id_veiculo == veiculo.id_veiculo]
            
            if eventos_veiculo:
                for evento_veiculo in eventos_veiculo:
                    dias_uso = (evento_veiculo.data_devolucao - evento_veiculo.data_inicio).days + 1
                    data.append([
                        veiculo.nome,
                        veiculo.categoria.nome if veiculo.categoria else 'Sem categoria',
                        veiculo.placa or 'Sem placa',
                        evento_veiculo.evento.nome,
                        evento_veiculo.motorista.nome,
                        evento_veiculo.data_inicio.strftime('%d/%m/%Y'),
                        evento_veiculo.data_devolucao.strftime('%d/%m/%Y'),
                        str(dias_uso),
                        evento_veiculo.observacoes or ''
                    ])
            else:
                # Veículo sem eventos
                data.append([
                    veiculo.nome,
                    veiculo.categoria.nome if veiculo.categoria else 'Sem categoria',
                    veiculo.placa or 'Sem placa',
                    'Nenhum evento',
                    '',
                    '',
                    '',
                    '',
                    ''
                ])
        
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f"relatorio_veiculos_{today}"
        
        if formato == 'excel':
            return criar_excel_response(headers, data, filename)
        elif formato == 'pdf':
            return criar_pdf_response(headers, data, 'Relatório de Veículos e Eventos', filename)
        else:
            return jsonify({'error': 'Formato não suportado'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== GESTÃO FINANCEIRA EMPRESA ====================

@app.route('/empresa/despesas', methods=['GET', 'POST'])
def despesas_empresa():
    """Gestão de despesas da empresa"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verificar se usuário é administrativo
    if session.get('categoria') != 'administrativo':
        flash('Acesso negado. Apenas usuários administrativos podem acessar esta funcionalidade.', 'danger')
        return redirect(url_for('dashboard'))
    
    form = DespesaEmpresaForm()
    
    # Carregar opções dos selects - apenas categorias que possuem despesas da empresa (tipos 3 e 4)
    categorias = db.session.query(CategoriaDespesa).join(Despesa).filter(
        Despesa.id_tipo_despesa.in_([3, 4])
    ).distinct().all()
    form.categoria_despesa.choices = [(0, 'Selecione uma categoria')] + [(c.id_categoria_despesa, c.nome) for c in categorias]
    
    # Filtrar apenas despesas da empresa (tipos 3 e 4)
    despesas_opcoes = Despesa.query.filter(Despesa.id_tipo_despesa.in_([3, 4])).all()
    form.despesa_id.choices = [(0, 'Selecione uma categoria primeiro')] + [(d.id_despesa, d.nome) for d in despesas_opcoes]
    
    fornecedores = Fornecedor.query.all()
    form.fornecedor_id.choices = [(0, 'Selecione um fornecedor (opcional)')] + [(f.id_fornecedor, f.nome) for f in fornecedores]
    
    if form.validate_on_submit():
        try:
            # Upload do comprovante se houver
            filename = None
            if form.comprovante.data:
                file = form.comprovante.data
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    # Adicionar timestamp para evitar conflitos
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
                    filename = timestamp + filename
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'comprovantes', filename))
            
            despesa_empresa = DespesaEmpresa(
                id_despesa=form.despesa_id.data,
                data_vencimento=form.data_vencimento.data,
                data_pagamento=form.data_pagamento.data,
                valor=form.valor.data,
                valor_pago_socrates=form.valor_pago_socrates.data if form.valor_pago_socrates.data else None,
                id_fornecedor=form.fornecedor_id.data if form.fornecedor_id.data != 0 else None,
                status_pagamento=form.status_pagamento.data,
                forma_pagamento=form.forma_pagamento.data,
                pago_por=form.pago_por.data,
                observacoes=form.observacoes.data,
                comprovante=filename,
                qtd_dias=form.qtd_dias.data,
                qtd_pessoas=form.qtd_pessoas.data
            )
            
            db.session.add(despesa_empresa)
            db.session.commit()
            
            # Registrar log da operação
            despesa_nome = Despesa.query.get(form.despesa_id.data).nome if form.despesa_id.data else "Despesa"
            registrar_log('Criar Despesa Empresa', f'Despesa "{despesa_nome}" da empresa cadastrada - R$ {form.valor.data:.2f}')
            
            flash('Despesa da empresa cadastrada com sucesso!', 'success')
            return redirect(url_for('despesas_empresa'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar despesa: {str(e)}', 'danger')
    
    # Listar despesas cadastradas
    despesas = db.session.query(DespesaEmpresa).join(Despesa).join(CategoriaDespesa).order_by(DespesaEmpresa.data_vencimento.desc()).all()
    
    return render_template('despesas_empresa.html', form=form, despesas=despesas)

@app.route('/empresa/receitas', methods=['GET', 'POST'])
def receitas_empresa():
    """Gestão de receitas da empresa"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verificar se usuário é administrativo
    if session.get('categoria') != 'administrativo':
        flash('Acesso negado. Apenas usuários administrativos podem acessar esta funcionalidade.', 'danger')
        return redirect(url_for('dashboard'))
    
    form = ReceitaEmpresaForm()
    
    # Carregar opções dos selects
    categorias = CategoriaReceita.query.all()
    form.categoria_receita.choices = [(0, 'Selecione uma categoria')] + [(c.id_categoria_receita, c.nome) for c in categorias]
    
    receitas_opcoes = Receita.query.all()
    form.receita_id.choices = [(0, 'Selecione uma categoria primeiro')] + [(r.id_receita, r.nome) for r in receitas_opcoes]
    
    if form.validate_on_submit():
        try:
            receita_empresa = ReceitaEmpresa(
                id_receita=form.receita_id.data,
                data=form.data.data,
                valor=form.valor.data,
                observacoes=form.observacoes.data
            )
            
            db.session.add(receita_empresa)
            db.session.commit()
            
            # Registrar log da operação
            receita_nome = Receita.query.get(form.receita_id.data).nome if form.receita_id.data else "Receita"
            registrar_log('Criar Receita Empresa', f'Receita "{receita_nome}" da empresa cadastrada - R$ {form.valor.data:.2f}')
            
            flash('Receita da empresa cadastrada com sucesso!', 'success')
            return redirect(url_for('receitas_empresa'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar receita: {str(e)}', 'danger')
    
    # Listar receitas cadastradas
    receitas = db.session.query(ReceitaEmpresa).join(Receita).join(CategoriaReceita).order_by(ReceitaEmpresa.data.desc()).all()
    
    return render_template('receitas_empresa.html', form=form, receitas=receitas)

@app.route('/empresa/despesas/editar/<int:id>', methods=['GET', 'POST'])
def editar_despesa_empresa(id):
    """Editar despesa da empresa"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if session.get('categoria') != 'administrativo':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('dashboard'))
    
    despesa_empresa = DespesaEmpresa.query.get_or_404(id)
    form = DespesaEmpresaForm(obj=despesa_empresa)
    
    # Carregar opções dos selects - apenas categorias que possuem despesas da empresa (tipos 3 e 4)
    categorias = db.session.query(CategoriaDespesa).join(Despesa).filter(
        Despesa.id_tipo_despesa.in_([3, 4])
    ).distinct().all()
    form.categoria_despesa.choices = [(0, 'Selecione uma categoria')] + [(c.id_categoria_despesa, c.nome) for c in categorias]
    form.categoria_despesa.data = despesa_empresa.despesa.id_categoria_despesa
    
    despesas_opcoes = Despesa.query.filter(Despesa.id_tipo_despesa.in_([3, 4])).all()
    form.despesa_id.choices = [(d.id_despesa, d.nome) for d in despesas_opcoes]
    
    fornecedores = Fornecedor.query.all()
    form.fornecedor_id.choices = [(0, 'Selecione um fornecedor (opcional)')] + [(f.id_fornecedor, f.nome) for f in fornecedores]
    
    # Carregar explicitamente os valores dos campos de quantidade no GET request
    if request.method == 'GET':
        form.despesa_id.data = despesa_empresa.id_despesa  # Pré-selecionar a despesa correta
        form.qtd_dias.data = despesa_empresa.qtd_dias
        form.qtd_pessoas.data = despesa_empresa.qtd_pessoas
        print(f"DEBUG EDIÇÃO: despesa_id={despesa_empresa.id_despesa}, qtd_dias={despesa_empresa.qtd_dias}, qtd_pessoas={despesa_empresa.qtd_pessoas}")
    
    if form.validate_on_submit():
        try:
            # Upload do comprovante se houver
            if form.comprovante.data:
                file = form.comprovante.data
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
                    filename = timestamp + filename
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'comprovantes', filename))
                    despesa_empresa.comprovante = filename
            
            despesa_empresa.id_despesa = form.despesa_id.data
            despesa_empresa.data_vencimento = form.data_vencimento.data
            despesa_empresa.data_pagamento = form.data_pagamento.data
            despesa_empresa.valor = form.valor.data
            despesa_empresa.id_fornecedor = form.fornecedor_id.data if form.fornecedor_id.data != 0 else None
            despesa_empresa.status_pagamento = form.status_pagamento.data
            despesa_empresa.forma_pagamento = form.forma_pagamento.data
            despesa_empresa.pago_por = form.pago_por.data
            despesa_empresa.observacoes = form.observacoes.data
            despesa_empresa.qtd_dias = form.qtd_dias.data
            despesa_empresa.qtd_pessoas = form.qtd_pessoas.data
            
            db.session.commit()
            
            # Registrar log da operação
            despesa_nome = Despesa.query.get(form.despesa_id.data).nome if form.despesa_id.data else "Despesa"
            registrar_log('Editar Despesa Empresa', f'Despesa "{despesa_nome}" da empresa editada - R$ {form.valor.data:.2f}')
            
            flash('Despesa da empresa atualizada com sucesso!', 'success')
            return redirect(url_for('despesas_empresa'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar despesa: {str(e)}', 'danger')
    
    return render_template('editar_despesa_empresa.html', form=form, despesa_empresa=despesa_empresa)

@app.route('/empresa/receitas/editar/<int:id>', methods=['GET', 'POST'])
def editar_receita_empresa(id):
    """Editar receita da empresa"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if session.get('categoria') != 'administrativo':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('dashboard'))
    
    receita_empresa = ReceitaEmpresa.query.get_or_404(id)
    form = ReceitaEmpresaForm(obj=receita_empresa)
    
    # Carregar opções dos selects
    categorias = CategoriaReceita.query.all()
    form.categoria_receita.choices = [(c.id_categoria_receita, c.nome) for c in categorias]
    form.categoria_receita.data = receita_empresa.receita.id_categoria_receita
    
    receitas_opcoes = Receita.query.all()
    form.receita_id.choices = [(r.id_receita, r.nome) for r in receitas_opcoes]
    
    # Preencher o valor no formato brasileiro se for GET
    if request.method == 'GET' and receita_empresa.valor:
        form.valor.data = str(receita_empresa.valor).replace('.', ',')
    
    if form.validate_on_submit():
        try:
            receita_empresa.id_receita = form.receita_id.data
            receita_empresa.data = form.data.data
            receita_empresa.valor = form.valor.data
            receita_empresa.observacoes = form.observacoes.data
            
            db.session.commit()
            
            # Registrar log da operação
            receita_nome = Receita.query.get(form.receita_id.data).nome if form.receita_id.data else "Receita"
            registrar_log('Editar Receita Empresa', f'Receita "{receita_nome}" da empresa editada - R$ {form.valor.data:.2f}')
            
            flash('Receita da empresa atualizada com sucesso!', 'success')
            return redirect(url_for('receitas_empresa'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar receita: {str(e)}', 'danger')
    
    return render_template('editar_receita_empresa.html', form=form, receita_empresa=receita_empresa)

@app.route('/empresa/despesas/excluir/<int:id>')
def excluir_despesa_empresa(id):
    """Excluir despesa da empresa"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if session.get('categoria') != 'administrativo':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        despesa_empresa = DespesaEmpresa.query.get_or_404(id)
        
        # Registrar log da operação antes da exclusão
        despesa_nome = despesa_empresa.despesa.nome if despesa_empresa.despesa else "Despesa"
        registrar_log('Excluir Despesa Empresa', f'Despesa "{despesa_nome}" da empresa excluída - R$ {despesa_empresa.valor:.2f}')
        
        # Excluir arquivo de comprovante se existir
        if despesa_empresa.comprovante:
            arquivo_path = os.path.join(app.config['UPLOAD_FOLDER'], 'comprovantes', despesa_empresa.comprovante)
            if os.path.exists(arquivo_path):
                os.remove(arquivo_path)
        
        db.session.delete(despesa_empresa)
        db.session.commit()
        flash('Despesa da empresa excluída com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir despesa: {str(e)}', 'danger')
    
    return redirect(url_for('despesas_empresa'))

@app.route('/empresa/receitas/excluir/<int:id>')
def excluir_receita_empresa(id):
    """Excluir receita da empresa"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if session.get('categoria') != 'administrativo':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        receita_empresa = ReceitaEmpresa.query.get_or_404(id)
        
        # Registrar log da operação antes da exclusão
        receita_nome = receita_empresa.receita.nome if receita_empresa.receita else "Receita"
        registrar_log('Excluir Receita Empresa', f'Receita "{receita_nome}" da empresa excluída - R$ {receita_empresa.valor:.2f}')
        
        db.session.delete(receita_empresa)
        db.session.commit()
        flash('Receita da empresa excluída com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir receita: {str(e)}', 'danger')
    
    return redirect(url_for('receitas_empresa'))

# ==================== FINANCEIRO MÃS ====================

@app.route('/empresa/financeiro-mes', methods=['GET', 'POST'])
def financeiro_mes():
    """Visualizar despesas e receitas da empresa para um mês específico"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if session.get('categoria') != 'administrativo':
        flash('Acesso negado. Apenas usuários administrativos podem acessar esta funcionalidade.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Obter mês e ano atual como padrão
    hoje = date.today()
    mes_atual = hoje.month
    ano_atual = hoje.year
    
    if request.method == 'POST':
        mes = request.form.get('mes')
        ano = request.form.get('ano')
        
        if not mes or not ano:
            flash('Por favor, selecione o mês e o ano.', 'danger')
            return render_template('financeiro_mes.html', 
                                 mes_padrao=mes_atual, 
                                 ano_padrao=ano_atual)
        
        try:
            mes_int = int(mes)
            ano_int = int(ano)
        except ValueError:
            flash('Mês e ano devem ser números válidos.', 'danger')
            return render_template('financeiro_mes.html', 
                                 mes_padrao=mes_atual, 
                                 ano_padrao=ano_atual)
    else:
        # Para requisições GET, verificar se tem parâmetros na URL
        mes_get = request.args.get('mes')
        ano_get = request.args.get('ano')
        
        if mes_get and ano_get:
            try:
                mes_int = int(mes_get)
                ano_int = int(ano_get)
            except ValueError:
                mes_int = mes_atual
                ano_int = ano_atual
        else:
            mes_int = mes_atual
            ano_int = ano_atual
    
    # Buscar despesas e receitas da empresa sempre
    despesas_mes = listar_despesas_empresa_mes(mes_int, ano_int)
    receitas_mes = listar_receitas_empresa_mes(mes_int, ano_int)
    
    # Buscar despesas e receitas de eventos para o mês
    despesas_evento_mes = listar_despesas_evento_mes(mes_int, ano_int)
    receitas_evento_mes = listar_receitas_evento_mes(mes_int, ano_int)
    
    # Calcular subtotais
    subtotal_despesas_empresa = sum(d.valor for d in despesas_mes) if despesas_mes else 0
    subtotal_receitas_empresa = sum(r.valor for r in receitas_mes) if receitas_mes else 0
    subtotal_despesas_evento = sum(d.valor for d in despesas_evento_mes) if despesas_evento_mes else 0
    subtotal_receitas_evento = sum(r.valor for r in receitas_evento_mes) if receitas_evento_mes else 0
    
    # Verificar se existem despesas fixas da empresa que ainda não foram criadas
    despesas_fixas_pendentes = verificar_despesas_fixas_pendentes(mes_int, ano_int)
    
    if request.method == 'GET':
        total_registros = len(despesas_mes) + len(receitas_mes) + len(despesas_evento_mes) + len(receitas_evento_mes)
        if total_registros > 0:
            flash(f'Exibindo dados financeiros de {mes_int:02d}/{ano_int}.', 'info')
        else:
            flash(f'Nenhuma despesa ou receita encontrada para {mes_int:02d}/{ano_int}.', 'info')
    
    return render_template('financeiro_mes.html', 
                         despesas_mes=despesas_mes, 
                         receitas_mes=receitas_mes, 
                         despesas_evento_mes=despesas_evento_mes,
                         receitas_evento_mes=receitas_evento_mes,
                         subtotal_despesas_empresa=subtotal_despesas_empresa,
                         subtotal_receitas_empresa=subtotal_receitas_empresa,
                         subtotal_despesas_evento=subtotal_despesas_evento,
                         subtotal_receitas_evento=subtotal_receitas_evento,
                         mes_selecionado=mes_int, 
                         ano_selecionado=ano_int,
                         mes_padrao=mes_atual,
                         ano_padrao=ano_atual,
                         mostrar_listagem=True,
                         despesas_fixas_pendentes=despesas_fixas_pendentes)

def verificar_despesas_empresa_mes(mes, ano):
    """Verificar se já existem despesas da empresa para o mês especificado"""
    from datetime import date
    
    # Criar data de início e fim do mês
    data_inicio = date(ano, mes, 1)
    
    # Calcular último dia do mês
    if mes == 12:
        data_fim = date(ano + 1, 1, 1) - timedelta(days=1)
    else:
        data_fim = date(ano, mes + 1, 1) - timedelta(days=1)
    
    # Verificar se existem despesas da empresa no período
    despesas_existentes = DespesaEmpresa.query.filter(
        DespesaEmpresa.data_vencimento >= data_inicio,
        DespesaEmpresa.data_vencimento <= data_fim
    ).count()
    
    return despesas_existentes > 0

def verificar_despesas_fixas_pendentes(mes, ano):
    """Verificar quais despesas fixas da empresa ainda não foram criadas para o mês especificado"""
    from datetime import date
    
    # Criar data de início e fim do mês
    data_inicio = date(ano, mes, 1)
    
    # Calcular último dia do mês
    if mes == 12:
        data_fim = date(ano + 1, 1, 1) - timedelta(days=1)
    else:
        data_fim = date(ano, mes + 1, 1) - timedelta(days=1)
    
    # Buscar todas as despesas fixas da empresa (tipo 3)
    despesas_fixas = Despesa.query.filter_by(id_tipo_despesa=3).all()
    
    # Verificar quais ainda não foram criadas para este mês
    despesas_pendentes = []
    
    for despesa_fixa in despesas_fixas:
        # Verificar se já existe uma despesa desta categoria neste mês
        despesa_existente = DespesaEmpresa.query.filter_by(
            id_despesa=despesa_fixa.id_despesa
        ).filter(
            DespesaEmpresa.data_vencimento >= data_inicio,
            DespesaEmpresa.data_vencimento <= data_fim
        ).first()
        
        if not despesa_existente:
            despesas_pendentes.append(despesa_fixa)
    
    return despesas_pendentes

def criar_despesas_fixas_mes(mes, ano):
    """Criar despesas fixas da empresa para o mês especificado"""
    from datetime import date
    
    # Buscar todas as despesas fixas da empresa (tipo 3)
    despesas_fixas = Despesa.query.filter_by(id_tipo_despesa=3).all()
    
    despesas_criadas = 0
    
    # Data padrão: primeiro dia do mês
    data_padrao = date(ano, mes, 1)
    
    for despesa_fixa in despesas_fixas:
        try:
            # Verificar se já existe uma despesa desta categoria neste mês
            despesa_existente = DespesaEmpresa.query.join(Despesa).filter(
                DespesaEmpresa.id_despesa == despesa_fixa.id_despesa,
                DespesaEmpresa.data_vencimento >= data_padrao,
                DespesaEmpresa.data_vencimento <= date(ano, mes, 28)  # Verificar até o dia 28 para evitar problemas com meses diferentes
            ).first()
            
            if not despesa_existente:
                # Criar nova despesa da empresa
                nova_despesa = DespesaEmpresa(
                    id_despesa=despesa_fixa.id_despesa,
                    data_vencimento=data_padrao,
                    valor=float(despesa_fixa.valor_medio_despesa) if despesa_fixa.valor_medio_despesa else 0.0,
                    status_pagamento='pendente',
                    forma_pagamento='débito',
                    observacoes=f'Despesa fixa criada automaticamente para {mes:02d}/{ano}'
                )
                
                db.session.add(nova_despesa)
                despesas_criadas += 1
        
        except Exception as e:
            print(f"Erro ao criar despesa fixa {despesa_fixa.nome}: {str(e)}")
            continue
    
    try:
        db.session.commit()
        return despesas_criadas
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao salvar despesas fixas: {str(e)}")
        return 0

def listar_despesas_empresa_mes(mes, ano):
    """Listar despesas da empresa para o mês especificado"""
    from datetime import date
    
    # Criar data de início e fim do mês
    data_inicio = date(ano, mes, 1)
    
    # Calcular último dia do mês
    if mes == 12:
        data_fim = date(ano + 1, 1, 1) - timedelta(days=1)
    else:
        data_fim = date(ano, mes + 1, 1) - timedelta(days=1)
    
    # Buscar despesas da empresa no período
    despesas_mes = db.session.query(DespesaEmpresa).join(Despesa).join(CategoriaDespesa).filter(
        DespesaEmpresa.data_vencimento >= data_inicio,
        DespesaEmpresa.data_vencimento <= data_fim
    ).order_by(DespesaEmpresa.data_vencimento.desc()).all()
    
    return despesas_mes

def listar_receitas_empresa_mes(mes, ano):
    """Listar receitas da empresa para o mês especificado"""
    from datetime import date
    
    # Criar data de início e fim do mês
    data_inicio = date(ano, mes, 1)
    
    # Calcular último dia do mês
    if mes == 12:
        data_fim = date(ano + 1, 1, 1) - timedelta(days=1)
    else:
        data_fim = date(ano, mes + 1, 1) - timedelta(days=1)
    
    # Buscar receitas da empresa no período
    receitas_mes = db.session.query(ReceitaEmpresa).join(Receita).join(CategoriaReceita).filter(
        ReceitaEmpresa.data >= data_inicio,
        ReceitaEmpresa.data <= data_fim
    ).order_by(ReceitaEmpresa.data.desc()).all()
    
    return receitas_mes

def listar_despesas_evento_mes(mes, ano):
    """Listar despesas de eventos para o mês especificado"""
    from datetime import date
    
    # Criar data de início e fim do mês
    data_inicio = date(ano, mes, 1)
    
    # Calcular último dia do mês
    if mes == 12:
        data_fim = date(ano + 1, 1, 1) - timedelta(days=1)
    else:
        data_fim = date(ano, mes + 1, 1) - timedelta(days=1)
    
    # Buscar despesas de eventos no período (por data_vencimento)
    despesas_evento_mes = db.session.query(DespesaEvento).join(Despesa).join(CategoriaDespesa).join(Evento).filter(
        DespesaEvento.data_vencimento >= data_inicio,
        DespesaEvento.data_vencimento <= data_fim
    ).order_by(DespesaEvento.data_vencimento.desc()).all()
    
    return despesas_evento_mes

def listar_receitas_evento_mes(mes, ano):
    """Listar receitas de eventos para o mês especificado"""
    from datetime import date
    
    # Criar data de início e fim do mês
    data_inicio = date(ano, mes, 1)
    
    # Calcular último dia do mês
    if mes == 12:
        data_fim = date(ano + 1, 1, 1) - timedelta(days=1)
    else:
        data_fim = date(ano, mes + 1, 1) - timedelta(days=1)
    
    # Buscar receitas de eventos no período (por data)
    receitas_evento_mes = db.session.query(ReceitaEvento).join(Receita).join(CategoriaReceita).join(Evento).filter(
        ReceitaEvento.data >= data_inicio,
        ReceitaEvento.data <= data_fim
    ).order_by(ReceitaEvento.data.desc()).all()
    
    return receitas_evento_mes

@app.route('/empresa/adicionar-despesas-fixas', methods=['POST'])
def adicionar_despesas_fixas():
    """Adicionar despesas fixas da empresa para um mês específico"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if session.get('categoria') != 'administrativo':
        flash('Acesso negado. Apenas usuários administrativos podem acessar esta funcionalidade.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Obter mês e ano do formulário
    mes = request.form.get('mes')
    ano = request.form.get('ano')
    
    if not mes or not ano:
        flash('Por favor, selecione o mês e o ano.', 'danger')
        return redirect(url_for('financeiro_mes'))
    
    try:
        mes_int = int(mes)
        ano_int = int(ano)
    except ValueError:
        flash('Mês e ano devem ser números válidos.', 'danger')
        return redirect(url_for('financeiro_mes'))
    
    # Criar despesas fixas
    despesas_criadas = criar_despesas_fixas_mes(mes_int, ano_int)
    
    if despesas_criadas > 0:
        flash(f'{despesas_criadas} despesas fixas foram adicionadas para {mes_int:02d}/{ano_int}.', 'success')
    else:
        flash('Nenhuma despesa fixa foi encontrada para adicionar ou todas já foram criadas.', 'info')
    
    # Retornar para a página do financeiro com os parâmetros corretos
    return redirect(url_for('financeiro_mes') + f'?mes={mes_int}&ano={ano_int}')

# ==================== RELATÓRIO CUSTO DA FROTA ====================

@app.route('/relatorios/custo-frota')
def relatorio_custo_frota():
    """Relatório de custo da frota por evento"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verificar se é administrador
    if session.get('categoria', '').lower() != 'administrativo':
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        # Obter evento selecionado
        evento_id = request.args.get('evento_id', '')
        
        # Buscar todos os eventos para o seletor
        eventos = Evento.query.order_by(Evento.data_inicio.desc()).all()
        
        dados_relatorio = None
        erro_parametro = False
        
        if evento_id:
            # Verificar se o preço da gasolina está cadastrado
            parametro_gasolina = Parametro.query.filter_by(parametro='gasolina').first()
            
            if not parametro_gasolina or not parametro_gasolina.valor:
                erro_parametro = True
            else:
                try:
                    preco_gasolina = float(parametro_gasolina.valor.replace(',', '.'))
                    
                    # Buscar veículos usados no evento
                    veiculos_evento = db.session.query(VeiculoEvento).join(
                        Veiculo, VeiculoEvento.id_veiculo == Veiculo.id_veiculo
                    ).join(
                        Colaborador, VeiculoEvento.id_motorista == Colaborador.id_colaborador
                    ).filter(
                        VeiculoEvento.id_evento == evento_id,
                        VeiculoEvento.km_inicio.isnot(None),
                        VeiculoEvento.km_fim.isnot(None)
                    ).all()
                    
                    custos_veiculos = []
                    custo_total = 0
                    
                    for veiculo_evento in veiculos_evento:
                        if veiculo_evento.veiculo.media_km_litro and veiculo_evento.veiculo.media_km_litro > 0:
                            km_rodados = veiculo_evento.km_fim - veiculo_evento.km_inicio
                            
                            if km_rodados > 0:
                                litros_gastos = km_rodados / veiculo_evento.veiculo.media_km_litro
                                custo_veiculo = litros_gastos * preco_gasolina
                                
                                custos_veiculos.append({
                                    'veiculo': veiculo_evento.veiculo,
                                    'motorista': veiculo_evento.motorista,
                                    'km_inicial': veiculo_evento.km_inicio,
                                    'km_final': veiculo_evento.km_fim,
                                    'km_rodados': km_rodados,
                                    'media_km_litro': veiculo_evento.veiculo.media_km_litro,
                                    'litros_gastos': round(litros_gastos, 2),
                                    'custo': round(custo_veiculo, 2),
                                    'data_inicio': veiculo_evento.data_inicio,
                                    'data_devolucao': veiculo_evento.data_devolucao
                                })
                                
                                custo_total += custo_veiculo
                    
                    # Buscar despesas de combustível do evento
                    despesas_combustivel = db.session.query(DespesaEvento).join(
                        Despesa, DespesaEvento.id_despesa == Despesa.id_despesa
                    ).filter(
                        DespesaEvento.id_evento == evento_id,
                        Despesa.flag_combustivel == True
                    ).all()
                    
                    # Calcular total gasto em combustível
                    total_gasto_combustivel = sum(despesa.valor for despesa in despesas_combustivel)
                    
                    # Calcular diferença entre real e esperado (real - esperado)
                    # Positivo: gastou mais que o esperado
                    # Negativo: gastou menos que o esperado (economia)
                    diferenca = total_gasto_combustivel - custo_total
                    percentual_diferenca = 0
                    if custo_total > 0:
                        percentual_diferenca = ((total_gasto_combustivel - custo_total) / custo_total) * 100
                    
                    evento_selecionado = Evento.query.get(evento_id)
                    dados_relatorio = {
                        'evento': evento_selecionado,
                        'custos_veiculos': custos_veiculos,
                        'custo_total': round(custo_total, 2),
                        'preco_gasolina': preco_gasolina,
                        'total_veiculos': len(custos_veiculos),
                        'despesas_combustivel': despesas_combustivel,
                        'total_gasto_combustivel': round(total_gasto_combustivel, 2),
                        'diferenca': round(diferenca, 2),
                        'percentual_diferenca': round(percentual_diferenca, 1)
                    }
                    
                except ValueError:
                    flash('Preço da gasolina inválido. Verifique o cadastro em Parâmetros.', 'danger')
        
        return render_template('relatorio_custo_frota.html',
                             eventos=eventos,
                             evento_id=evento_id,
                             dados_relatorio=dados_relatorio,
                             erro_parametro=erro_parametro)
        
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/relatorios/custo-frota/exportar/<string:formato>')
def exportar_custo_frota(formato):
    """Exportar relatório de custo da frota"""
    if 'user_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    # Verificar se é administrador
    if session.get('categoria', '').lower() != 'administrativo':
        return jsonify({'error': 'Acesso restrito a administradores'}), 403
    
    try:
        evento_id = request.args.get('evento_id', '')
        
        if not evento_id:
            return jsonify({'error': 'Evento não selecionado'}), 400
        
        # Verificar preço da gasolina
        parametro_gasolina = Parametro.query.filter_by(parametro='gasolina').first()
        if not parametro_gasolina or not parametro_gasolina.valor:
            return jsonify({'error': 'Preço da gasolina não cadastrado'}), 400
        
        preco_gasolina = float(parametro_gasolina.valor.replace(',', '.'))
        
        # Buscar dados do evento e veículos
        evento = Evento.query.get_or_404(evento_id)
        veiculos_evento = db.session.query(VeiculoEvento).join(
            Veiculo, VeiculoEvento.id_veiculo == Veiculo.id_veiculo
        ).join(
            Colaborador, VeiculoEvento.id_motorista == Colaborador.id_colaborador
        ).filter(
            VeiculoEvento.id_evento == evento_id,
            VeiculoEvento.km_inicio.isnot(None),
            VeiculoEvento.km_fim.isnot(None)
        ).all()
        
        # Buscar despesas de combustível do evento
        despesas_combustivel = db.session.query(DespesaEvento).join(
            Despesa, DespesaEvento.id_despesa == Despesa.id_despesa
        ).filter(
            DespesaEvento.id_evento == evento_id,
            Despesa.flag_combustivel == True
        ).all()
        
        # Preparar dados para exportação
        headers = ['Veículo', 'Placa', 'Motorista', 'KM Inicial', 'KM Final', 'KM Rodados', 'Média KM/L', 'Litros Gastos', 'Custo esperado (R$)']
        data = []
        custo_total = 0
        
        for veiculo_evento in veiculos_evento:
            if veiculo_evento.veiculo.media_km_litro and veiculo_evento.veiculo.media_km_litro > 0:
                km_rodados = veiculo_evento.km_fim - veiculo_evento.km_inicio
                
                if km_rodados > 0:
                    litros_gastos = km_rodados / veiculo_evento.veiculo.media_km_litro
                    custo_veiculo = litros_gastos * preco_gasolina
                    custo_total += custo_veiculo
                    
                    data.append([
                        veiculo_evento.veiculo.nome,
                        veiculo_evento.veiculo.placa or 'Sem placa',
                        veiculo_evento.motorista.nome,
                        f"{veiculo_evento.km_inicio:,}".replace(',', '.'),
                        f"{veiculo_evento.km_fim:,}".replace(',', '.'),
                        f"{km_rodados:,}".replace(',', '.'),
                        f"{veiculo_evento.veiculo.media_km_litro:,.1f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                        f"{litros_gastos:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                        f"R$ {custo_veiculo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    ])
        
        # Adicionar linha de total esperado
        data.append(['', '', '', '', '', '', '', 'CUSTO TOTAL ESPERADO:', 
                    f"R$ {custo_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')])
        
        # Adicionar seção de despesas reais
        if despesas_combustivel:
            data.append(['', '', '', '', '', '', '', '', ''])  # Linha em branco
            data.append(['', '', '', '', '', '', '', 'DESPESAS REAIS DE COMBUSTÍVEL:', ''])
            
            total_gasto_real = 0
            for despesa in despesas_combustivel:
                fornecedor = despesa.fornecedor.nome if despesa.fornecedor else 'Não informado'
                status = despesa.status_pagamento.title()
                data_pagamento = despesa.data_pagamento.strftime('%d/%m/%Y') if despesa.data_pagamento else 'Não pago'
                
                data.append([
                    despesa.despesa.nome,
                    fornecedor,
                    '',
                    '',
                    '',
                    '',
                    status,
                    data_pagamento,
                    f"R$ {despesa.valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                ])
                total_gasto_real += despesa.valor
            
            # Total real gasto
            data.append(['', '', '', '', '', '', '', 'TOTAL REAL GASTO:', 
                        f"R$ {total_gasto_real:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')])
            
            # Diferença
            diferenca = total_gasto_real - custo_total
            data.append(['', '', '', '', '', '', '', 'DIFERENÇA (Real - Esperado):', 
                        f"R$ {diferenca:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')])
        
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f"custo_frota_{evento.nome.replace(' ', '_')}_{today}"
        
        # Título do relatório
        titulo = f"Custo da Frota - {evento.nome}"
        
        if formato == 'excel':
            return criar_excel_response(headers, data, filename)
        elif formato == 'pdf':
            return criar_pdf_response(headers, data, titulo, filename)
        else:
            return jsonify({'error': 'Formato não suportado'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/relatorios/veiculos-servicos')
def relatorio_veiculos_servicos():
    """Relatório simplificado de veículos e seus serviços"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verificar se é ROOT ou administrador
    if not (is_root_user() or session.get('categoria', '').lower() == 'administrativo'):
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        # Obter filtros
        veiculo_filtro = request.args.get('veiculo', '')
        servico_filtro = request.args.get('servico', '')
        
        # Buscar todos os veículos para o filtro
        veiculos = db.session.query(Veiculo).join(CategoriaVeiculo).order_by(
            CategoriaVeiculo.nome, Veiculo.nome
        ).all()
        
        # Query base - depende do filtro de veículo
        if veiculo_filtro:
            veiculos_selecionados = [v for v in veiculos if str(v.id_veiculo) == veiculo_filtro]
        else:
            veiculos_selecionados = veiculos
        
        # Dados dos serviços por veículo
        dados_veiculos = []
        
        # Estatísticas gerais
        total_multas = 0
        total_ipva = 0
        total_licenciamento = 0
        total_manutencao = 0
        valor_total_geral = 0
        
        for veiculo in veiculos_selecionados:
            # Buscar serviços do veículo
            multas = MultaVeiculo.query.filter_by(id_veiculo=veiculo.id_veiculo).all()
            ipvas = IpvaVeiculo.query.filter_by(id_veiculo=veiculo.id_veiculo).all()
            licenciamentos = LicenciamentoVeiculo.query.filter_by(id_veiculo=veiculo.id_veiculo).all()
            manutencoes = ManutencaoVeiculo.query.filter_by(id_veiculo=veiculo.id_veiculo).all()
            
            # Calcular totais por tipo de serviço
            valor_multas = sum(m.valor_pago or m.valor_original for m in multas)
            valor_ipva = sum(i.valor_total for i in ipvas)
            valor_licenciamento = sum(l.valor_total for l in licenciamentos)
            valor_manutencao = sum(m.valor_total for m in manutencoes)
            
            # Contar serviços
            qtd_multas = len(multas)
            qtd_ipva = len(ipvas)
            qtd_licenciamento = len(licenciamentos)
            qtd_manutencao = len(manutencoes)
            
            # Filtrar por tipo de serviço se especificado
            if servico_filtro:
                if servico_filtro == 'multas' and qtd_multas == 0:
                    continue
                elif servico_filtro == 'ipva' and qtd_ipva == 0:
                    continue
                elif servico_filtro == 'licenciamento' and qtd_licenciamento == 0:
                    continue
                elif servico_filtro == 'manutencao' and qtd_manutencao == 0:
                    continue
            
            # Adicionar aos dados
            dados_veiculo = {
                'veiculo': veiculo,
                'multas': {
                    'quantidade': qtd_multas,
                    'valor': valor_multas,
                    'servicos': multas
                },
                'ipva': {
                    'quantidade': qtd_ipva,
                    'valor': valor_ipva,
                    'servicos': ipvas
                },
                'licenciamento': {
                    'quantidade': qtd_licenciamento,
                    'valor': valor_licenciamento,
                    'servicos': licenciamentos
                },
                'manutencao': {
                    'quantidade': qtd_manutencao,
                    'valor': valor_manutencao,
                    'servicos': manutencoes
                },
                'total_valor': valor_multas + valor_ipva + valor_licenciamento + valor_manutencao
            }
            
            dados_veiculos.append(dados_veiculo)
            
            # Somar aos totais gerais
            total_multas += qtd_multas
            total_ipva += qtd_ipva
            total_licenciamento += qtd_licenciamento
            total_manutencao += qtd_manutencao
            valor_total_geral += dados_veiculo['total_valor']
        
        # Preparar dados para gráficos baseados no filtro de serviço
        graficos_data = {}
        
        if servico_filtro:
            # Gráficos específicos para o tipo de serviço selecionado
            if servico_filtro == 'multas':
                # Buscar todas as multas dos veículos filtrados
                todas_multas = []
                for dados in dados_veiculos:
                    todas_multas.extend(dados['multas']['servicos'])
                
                # Top 10 veículos por multas
                veiculos_multas = sorted(dados_veiculos, key=lambda x: x['multas']['valor'], reverse=True)[:10]
                graficos_data['veiculos_top'] = {
                    'labels': [d['veiculo'].nome for d in veiculos_multas if d['multas']['valor'] > 0],
                    'data': [d['multas']['valor'] for d in veiculos_multas if d['multas']['valor'] > 0]
                }
                
                # Multas por status
                multas_por_status = {}
                for multa in todas_multas:
                    status = multa.status
                    multas_por_status[status] = multas_por_status.get(status, 0) + 1
                
                if multas_por_status:
                    graficos_data['servicos_por_status'] = {
                        'labels': list(multas_por_status.keys()),
                        'data': list(multas_por_status.values())
                    }
                
                # Multas por mês (últimos 12 meses)
                multas_por_mes = defaultdict(int)
                for multa in todas_multas:
                    if multa.data_infracao:
                        mes_ano = multa.data_infracao.strftime('%Y-%m')
                        multas_por_mes[mes_ano] += 1
                
                if multas_por_mes:
                    meses_ordenados = sorted(multas_por_mes.keys())[-12:]  # Últimos 12 meses
                    graficos_data['tendencia_temporal'] = {
                        'labels': [datetime.strptime(m, '%Y-%m').strftime('%m/%Y') for m in meses_ordenados],
                        'data': [multas_por_mes[m] for m in meses_ordenados]
                    }
                
            elif servico_filtro == 'ipva':
                # Top 10 veículos por IPVA
                veiculos_ipva = sorted(dados_veiculos, key=lambda x: x['ipva']['valor'], reverse=True)[:10]
                graficos_data['veiculos_top'] = {
                    'labels': [d['veiculo'].nome for d in veiculos_ipva if d['ipva']['valor'] > 0],
                    'data': [d['ipva']['valor'] for d in veiculos_ipva if d['ipva']['valor'] > 0]
                }
                
                # IPVA por ano de exercício
                todas_ipvas = []
                for dados in dados_veiculos:
                    todas_ipvas.extend(dados['ipva']['servicos'])
                
                ipva_por_ano = {}
                for ipva in todas_ipvas:
                    ano = ipva.ano_exercicio
                    ipva_por_ano[ano] = ipva_por_ano.get(ano, 0) + ipva.valor_total
                
                if ipva_por_ano:
                    graficos_data['ipva_por_ano'] = {
                        'labels': [str(ano) for ano in sorted(ipva_por_ano.keys())],
                        'data': [ipva_por_ano[ano] for ano in sorted(ipva_por_ano.keys())]
                    }
                
            elif servico_filtro == 'licenciamento':
                # Top 10 veículos por licenciamento
                veiculos_lic = sorted(dados_veiculos, key=lambda x: x['licenciamento']['valor'], reverse=True)[:10]
                graficos_data['veiculos_top'] = {
                    'labels': [d['veiculo'].nome for d in veiculos_lic if d['licenciamento']['valor'] > 0],
                    'data': [d['licenciamento']['valor'] for d in veiculos_lic if d['licenciamento']['valor'] > 0]
                }
                
                # Licenciamentos por ano
                todos_licenciamentos = []
                for dados in dados_veiculos:
                    todos_licenciamentos.extend(dados['licenciamento']['servicos'])
                
                lic_por_ano = {}
                for lic in todos_licenciamentos:
                    ano = lic.ano_exercicio
                    lic_por_ano[ano] = lic_por_ano.get(ano, 0) + lic.valor_total
                
                if lic_por_ano:
                    graficos_data['licenciamento_por_ano'] = {
                        'labels': [str(ano) for ano in sorted(lic_por_ano.keys())],
                        'data': [lic_por_ano[ano] for ano in sorted(lic_por_ano.keys())]
                    }
                
            elif servico_filtro == 'manutencao':
                # Top 10 veículos por manutenção
                veiculos_manut = sorted(dados_veiculos, key=lambda x: x['manutencao']['valor'], reverse=True)[:10]
                graficos_data['veiculos_top'] = {
                    'labels': [d['veiculo'].nome for d in veiculos_manut if d['manutencao']['valor'] > 0],
                    'data': [d['manutencao']['valor'] for d in veiculos_manut if d['manutencao']['valor'] > 0]
                }
                
                # Manutenção por tipo
                todas_manutencoes = []
                for dados in dados_veiculos:
                    todas_manutencoes.extend(dados['manutencao']['servicos'])
                
                manut_por_tipo = {}
                for manut in todas_manutencoes:
                    tipo = manut.tipo_manutencao
                    manut_por_tipo[tipo] = manut_por_tipo.get(tipo, 0) + manut.valor_total
                
                if manut_por_tipo:
                    graficos_data['manutencao_por_tipo'] = {
                        'labels': list(manut_por_tipo.keys()),
                        'data': list(manut_por_tipo.values())
                    }
                
                # Tendência mensal de manutenções
                manut_por_mes = defaultdict(float)
                for manut in todas_manutencoes:
                    if manut.data_servico:
                        mes_ano = manut.data_servico.strftime('%Y-%m')
                        manut_por_mes[mes_ano] += manut.valor_total
                
                if manut_por_mes:
                    meses_ordenados = sorted(manut_por_mes.keys())[-12:]  # Últimos 12 meses
                    graficos_data['tendencia_temporal'] = {
                        'labels': [datetime.strptime(m, '%Y-%m').strftime('%m/%Y') for m in meses_ordenados],
                        'data': [manut_por_mes[m] for m in meses_ordenados]
                    }
        else:
            # Gráficos gerais (quando nenhum serviço específico é filtrado)
            graficos_data = {
                'servicos_por_tipo': {
                    'labels': ['Multas', 'IPVA', 'Licenciamento', 'Manutenção'],
                    'data': [total_multas, total_ipva, total_licenciamento, total_manutencao]
                },
                'valores_por_tipo': {
                    'labels': ['Multas', 'IPVA', 'Licenciamento', 'Manutenção'],
                    'data': [
                        sum(d['multas']['valor'] for d in dados_veiculos),
                        sum(d['ipva']['valor'] for d in dados_veiculos),
                        sum(d['licenciamento']['valor'] for d in dados_veiculos),
                        sum(d['manutencao']['valor'] for d in dados_veiculos)
                    ]
                }
            }
            
            # Gráfico por veículo (top 10)
            if len(dados_veiculos) > 0:
                dados_veiculos_ordenados = sorted(dados_veiculos, key=lambda x: x['total_valor'], reverse=True)[:10]
                graficos_data['veiculos_top'] = {
                    'labels': [d['veiculo'].nome for d in dados_veiculos_ordenados],
                    'data': [d['total_valor'] for d in dados_veiculos_ordenados]
                }
        
        # Estatísticas resumo baseadas no filtro
        if servico_filtro:
            # Estatísticas específicas do serviço
            servico_dados = dados_veiculos[0][servico_filtro] if dados_veiculos else {'quantidade': 0, 'valor': 0}
            total_servicos_filtro = sum(d[servico_filtro]['quantidade'] for d in dados_veiculos)
            total_valor_filtro = sum(d[servico_filtro]['valor'] for d in dados_veiculos)
            
            estatisticas = {
                'total_veiculos': len([d for d in dados_veiculos if d[servico_filtro]['quantidade'] > 0]),
                'total_servicos': total_servicos_filtro,
                'valor_total': total_valor_filtro,
                'media_por_veiculo': total_valor_filtro / len(dados_veiculos) if len(dados_veiculos) > 0 else 0,
                'servico_selecionado': servico_filtro
            }
        else:
            # Estatísticas gerais
            estatisticas = {
                'total_veiculos': len(dados_veiculos),
                'total_servicos': total_multas + total_ipva + total_licenciamento + total_manutencao,
                'valor_total': valor_total_geral,
                'media_por_veiculo': valor_total_geral / len(dados_veiculos) if len(dados_veiculos) > 0 else 0
            }
        
        # Opções para filtros
        tipos_servico = [
            {'value': '', 'label': 'Todos os Serviços'},
            {'value': 'multas', 'label': 'Multas'},
            {'value': 'ipva', 'label': 'IPVA'},
            {'value': 'licenciamento', 'label': 'Licenciamento'},
            {'value': 'manutencao', 'label': 'Manutenção'}
        ]
        
        # Gravar log
        registrar_log('VISUALIZAR', f'Relatório de Veículos e Serviços visualizado')
        
        return render_template(
            'relatorio_veiculos_servicos.html',
            dados_veiculos=dados_veiculos,
            veiculos=veiculos,
            tipos_servico=tipos_servico,
            graficos_data=graficos_data,
            estatisticas=estatisticas,
            veiculo_filtro=veiculo_filtro,
            servico_filtro=servico_filtro
        )
        
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/relatorios/veiculos-servicos/exportar/<string:formato>')
def exportar_relatorio_veiculos_servicos(formato):
    """Exportar relatório de veículos e serviços"""
    if 'user_id' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    # Verificar se é ROOT ou administrador
    if not (is_root_user() or session.get('categoria', '').lower() == 'administrativo'):
        return jsonify({'error': 'Acesso restrito a administradores'}), 403
    
    try:
        # Obter filtros
        veiculo_filtro = request.args.get('veiculo', '')
        servico_filtro = request.args.get('servico', '')
        
        # Buscar veículos
        veiculos = db.session.query(Veiculo).join(CategoriaVeiculo).order_by(
            CategoriaVeiculo.nome, Veiculo.nome
        ).all()
        
        # Aplicar filtro de veículo
        if veiculo_filtro:
            veiculos = [v for v in veiculos if str(v.id_veiculo) == veiculo_filtro]
        
        # Preparar dados para exportação
        headers = ['Veículo', 'Categoria', 'Placa', 'Multas (Qtd)', 'Multas (Valor)', 
                  'IPVA (Qtd)', 'IPVA (Valor)', 'Licenciamento (Qtd)', 'Licenciamento (Valor)',
                  'Manutenção (Qtd)', 'Manutenção (Valor)', 'Total Geral']
        
        data = []
        
        for veiculo in veiculos:
            # Buscar serviços
            multas = MultaVeiculo.query.filter_by(id_veiculo=veiculo.id_veiculo).all()
            ipvas = IpvaVeiculo.query.filter_by(id_veiculo=veiculo.id_veiculo).all()
            licenciamentos = LicenciamentoVeiculo.query.filter_by(id_veiculo=veiculo.id_veiculo).all()
            manutencoes = ManutencaoVeiculo.query.filter_by(id_veiculo=veiculo.id_veiculo).all()
            
            # Calcular valores
            valor_multas = sum(m.valor_pago or m.valor_original for m in multas)
            valor_ipva = sum(i.valor_total for i in ipvas)
            valor_licenciamento = sum(l.valor_total for l in licenciamentos)
            valor_manutencao = sum(m.valor_total for m in manutencoes)
            
            qtd_multas = len(multas)
            qtd_ipva = len(ipvas)
            qtd_licenciamento = len(licenciamentos)
            qtd_manutencao = len(manutencoes)
            
            # Filtrar por tipo de serviço se especificado
            if servico_filtro:
                if servico_filtro == 'multas' and qtd_multas == 0:
                    continue
                elif servico_filtro == 'ipva' and qtd_ipva == 0:
                    continue
                elif servico_filtro == 'licenciamento' and qtd_licenciamento == 0:
                    continue
                elif servico_filtro == 'manutencao' and qtd_manutencao == 0:
                    continue
            
            total_geral = valor_multas + valor_ipva + valor_licenciamento + valor_manutencao
            
            data.append([
                veiculo.nome,
                veiculo.categoria.nome if veiculo.categoria else 'N/A',
                veiculo.placa or 'N/A',
                qtd_multas,
                f"R$ {valor_multas:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                qtd_ipva,
                f"R$ {valor_ipva:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                qtd_licenciamento,
                f"R$ {valor_licenciamento:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                qtd_manutencao,
                f"R$ {valor_manutencao:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {total_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ])
        
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f"relatorio_veiculos_servicos_{today}"
        
        # Título do relatório
        titulo = "Relatório de Veículos e Serviços"
        if veiculo_filtro:
            veiculo_obj = next((v for v in veiculos if str(v.id_veiculo) == veiculo_filtro), None)
            if veiculo_obj:
                titulo += f" - {veiculo_obj.nome}"
        if servico_filtro:
            servicos_map = {'multas': 'Multas', 'ipva': 'IPVA', 'licenciamento': 'Licenciamento', 'manutencao': 'Manutenção'}
            titulo += f" - {servicos_map.get(servico_filtro, 'Todos')}"
        
        # Gravar log
        registrar_log('EXPORTAR', f'Relatório de Veículos e Serviços exportado ({formato.upper()})')
        
        if formato == 'excel':
            return criar_excel_response(headers, data, filename)
        elif formato == 'pdf':
            return criar_pdf_response(headers, data, titulo, filename)
        else:
            return jsonify({'error': 'Formato não suportado'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/administrativo/logs')
def listar_logs():
    """Lista todos os logs do sistema com paginação - acesso apenas para administrativos"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verificar se o usuário é administrativo
    usuario = Usuario.query.get(session['user_id'])
    if not usuario or not usuario.colaborador:
        flash('Erro ao carregar informações do usuário.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Verificar se é administrador
    is_admin = is_admin_user()
    if not is_admin:
        flash('Acesso negado. Esta funcionalidade é restrita a usuários administrativos.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Paginação - 50 registros por página
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Buscar logs ordenados por data/hora decrescente (mais recentes primeiro)
    logs = LogSistema.query.order_by(LogSistema.data_hora.desc()).paginate(
        page=page, 
        per_page=per_page, 
        error_out=False
    )
    
    return render_template('logs_sistema.html', logs=logs, usuario=usuario)

if __name__ == '__main__':
    # Inicializar usuário ROOT na primeira execução
    with app.app_context():
        criar_usuario_root()
    
    # Configuração de porta para Railway e desenvolvimento local
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    
    app.run(host='0.0.0.0', port=port, debug=debug)
